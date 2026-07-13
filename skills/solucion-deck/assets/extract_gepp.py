#!/usr/bin/env python3
# Extract GEPP proposal data from the two rev2 workbooks into JSON for the HTML presentation.
import openpyxl, json, re, unicodedata

SP = "/tmp/claude-1001/-home-mario/1a93f954-d79f-48bf-8b5c-d7a2f3592f1c/scratchpad"

def cellrows(ws):
    rows = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                rows.setdefault(c.row, {})[c.column_letter] = c.value
    return rows

def find_row(rows, prefix, col='A'):
    for r in sorted(rows):
        v = rows[r].get(col)
        if isinstance(v, str) and v.strip().startswith(prefix):
            return r
    return None

def num(v):
    return round(v, 6) if isinstance(v, (int, float)) else None

MONCOLS = list("BCDEFGHIJKLM")

def extract_site(ws):
    rows = cellrows(ws)
    out = {}
    out['title'] = rows[1]['A']
    # --- base data (section 1)
    r0 = find_row(rows, '1. DATOS BASE')
    hdr = r0 + 1
    out['months'] = [rows[hdr].get(c) for c in MONCOLS]
    def series(label):
        for r in range(hdr, hdr + 20):
            v = rows.get(r, {}).get('A', '')
            if isinstance(v, str) and v.startswith(label):
                return [num(rows[r].get(c)) for c in MONCOLS], num(rows[r].get('N'))
        return None, None
    for key, label in [('cbase','Consumo Base'),('cint','Consumo Intermedia'),('cpunta','Consumo Punta'),
                       ('ctotal','Consumo Total'),('dpunta','Dem. máx. en punta'),('dmax','Dem. máx. mensual'),
                       ('gasto','Gasto actual facturado')]:
        s, t = series(label)
        out[key], out[key+'_tot'] = s, t
    # --- section 2: proposed system, label -> [op1, op2]
    r2 = find_row(rows, '2. SISTEMA PROPUESTO')
    r3 = find_row(rows, '3. MODELO MENSUAL')
    sys = {}
    for r in range(r2+1, r3):
        a = rows.get(r, {}).get('A')
        if isinstance(a, str) and a.strip() and not a.startswith('—'):
            sys[a.strip()] = [rows[r].get('C'), rows[r].get('D')]
    out['sys'] = {k: [num(x) if isinstance(x,(int,float)) else x for x in v] for k, v in sys.items()}
    # --- monthly model op1: cols M (ahorro PV total) R (ahorro BESS bruto)
    hdr3 = r3 + 1
    mm = []
    r = hdr3 + 1
    while True:
        b = rows.get(r, {}).get('B')
        if b == 'TOTAL' or b is None: break
        mm.append({'pv': num(rows[r].get('M')), 'bess': num(rows[r].get('R'))})
        r += 1
    out['mm1'] = mm
    # --- projections
    def proj(section):
        rs = find_row(rows, section)
        if rs is None: return None
        h = rs + 1
        res = []
        r = h + 1
        while True:
            b = rows.get(r, {}).get('B')
            if b is None: break
            res.append({'yr': b,
                'pleno': num(rows[r].get('C')), 'autoab': num(rows[r].get('D')),
                'real': num(rows[r].get('E')), 'apv': num(rows[r].get('F')),
                'abess': num(rows[r].get('G')), 'ppa': num(rows[r].get('H')),
                'part': num(rows[r].get('I')), 'neto': num(rows[r].get('J')),
                'acum': num(rows[r].get('K')), 'pct': num(rows[r].get('L'))})
            r += 1
        return res
    out['proj1'] = proj('5. PROYECCIÓN 20 AÑOS — OPCIÓN 1')
    out['proj2'] = proj('6. PROYECCIÓN 20 AÑOS — OPCIÓN 2')
    # --- dispatch profiles
    def dispatch(prefix):
        rs = find_row(rows, prefix)
        if rs is None: return None
        h = rs + 1
        prof = []
        for r in range(h+1, h+25):
            rr = rows.get(r, {})
            prof.append({'h': rr.get('A'), 'per': rr.get('B'), 'carga': num(rr.get('C')),
                         'pv': num(rr.get('D')), 'chg': num(rr.get('F')), 'dis': num(rr.get('G')),
                         'neta': num(rr.get('H'))})
        # lectura block
        lect = {}
        for r in range(h+25, h+35):
            a = rows.get(r, {}).get('A', '')
            e = rows.get(r, {}).get('E')
            if isinstance(a, str) and e is not None:
                lect[a.strip()] = num(e)
        return {'prof': prof, 'lect': lect}
    out['inv'] = dispatch('8. DESPACHO DÍA TÍPICO — INVIERNO')
    out['ver'] = dispatch('9. DESPACHO DÍA TÍPICO — VERANO')
    return out

def extract_resumen(wb):
    ws = wb['Resumen']
    rows = cellrows(ws)
    res = {'rows': [], 'portfolio': {}}
    site = None
    for r in sorted(rows):
        rr = rows[r]
        a, b = rr.get('A'), rr.get('B')
        if a == 'PORTAFOLIO' or (site == 'PORTAFOLIO' and b):
            site = 'PORTAFOLIO'
            if b:
                res['portfolio'][b] = {k: num(rr.get(c)) for k, c in
                    [('kwp','C'),('bkw','D'),('bkwh','E'),('gen','F'),('gasto','G'),('autoab','H'),
                     ('bruto','I'),('ppa','J'),('part','K'),('neto','L')]}
            continue
        if b and isinstance(b, str) and b.startswith('Op'):
            if a: site = a
            res['rows'].append({'site': site, 'op': b, **{k: num(rr.get(c)) for k, c in
                [('kwp','C'),('bkw','D'),('bkwh','E'),('gen','F'),('gasto','G'),('autoab','H'),
                 ('bruto','I'),('ppa','J'),('part','K'),('neto','L'),('pctb','M'),('pctn','N'),
                 ('ppakwh','O'),('tirpv','P'),('tirbess','Q')]}})
    return res

wb4 = openpyxl.load_workbook(f'{SP}/gepp_4sitios.xlsx', data_only=True)
wbp = openpyxl.load_workbook(f'{SP}/gepp_proplasa.xlsx', data_only=True)

data = {'resumen4': extract_resumen(wb4), 'resumenP': extract_resumen(wbp), 'sites': {}}
for key, wb, sn in [('ixt', wb4, 'Ixtlahuacán'), ('aca', wb4, 'Acapulco'), ('can', wb4, 'Cancún'),
                    ('pro', wb4, 'Proplasa'), ('pr1', wbp, 'Preforma 1'), ('pr2', wbp, 'Preforma 2'),
                    ('tap', wbp, 'Tapa')]:
    data['sites'][key] = extract_site(wb[sn])
    print(key, 'ok — sys keys:', len(data['sites'][key]['sys']), '| proj1:', len(data['sites'][key]['proj1'] or []),
          '| inv:', bool(data['sites'][key]['inv']), '| gasto_tot:', data['sites'][key]['gasto_tot'])

json.dump(data, open(f'{SP}/gepp_data.json', 'w'), ensure_ascii=False)
print('\nwrote gepp_data.json', len(json.dumps(data)) // 1024, 'KB')
