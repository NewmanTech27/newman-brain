# -*- coding: utf-8 -*-
"""GEPP — modelo paramétrico de dimensionado BESS (jul-2026, análisis [[2026-07-07-gepp-bess-verano-resizing]]).

Replica EXACTA del motor mensual BESS embebido en los libros de propuesta GEPP
(verificada a 0.000% en R72 y TIR C52 contra los 7 sheets), parametrizada en (P, E):

  N    = MIN(E_útil·días_punta, P·hrs_punta·días_punta, kWh_punta)          [arbitraje]
  Q    = t_cap · MIN(D_punta, E_útil/horas_bloque, P)   horas_bloque: 2h may–sep, 4h oct–abr
  arb  = N·t_punta − (N/RTE)·t_base
  bruto año1 = Σ(arb+Q) · (12−merma)/12
  flujo inversionista @part: y0=−CAPEX; y1..plazo: part·bruto·esc^(t−1)·(1−deg)^(t−1) − O&M·E·esc^(t−1)

La forma cerrada de Q equivale al rasurado plano de la simulación horaria (±0.07% al tamaño
de diseño). Para los sheets Preforma del libro 3-sitios, el Q del motor es el agregado
prorrateado por E — usar q_mode='linear' (escala MIN(1,E/E_diseño)) + cap físico
Q ≤ t_cap·D_punta (Preforma 1 lo requiere: datos kWh-punta vs D-punta inconsistentes).

Uso:
  python tools/gepp_bess_scenarios.py                # verifica vs libros + tabla A/B/C
Escenarios: A=actual (libro), B=verano-2h (E_útil=2h·D̄_punta may–sep), C=eficiente
(mayor E cuyo kWh marginal rinde TIR ≥ WACC al inversionista). Recomendación vigente: B.
"""
import sys, io, re
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

BASE = r"C:\Users\vidan\Desktop\CFE Brain\entregables\propuestas"
BOOK_4S = BASE + r"\GEPP - Propuesta PV+BESS 4 Sitios (2026-07) rev2 - Despacho flat-top + DIST.xlsx"
BLOCK_H = [2, 2, 2, 2, 2, 4, 4, 4, 4, 4, 4, 4]      # may..abr; oct–abr contienen días de punta 4h
COLS = "BCDEFGHIJKLM"

# Supuestos (== hoja Supuestos de los libros; mantener sincronizados)
FX = 17.55; PRECIO_BESS = 0.28          # USD/Wh instalado (sistema 2h/0.5C)
ESC = 0.05; DEG = 0.0125; DOD = 0.96; RTE = 0.96
MERMA = 2; PART = 0.5; PLAZO = 15; HORIZONTE = 20; WACC = 0.12; OM = 180
UTIL = DOD * RTE ** 0.5                 # E_útil = E_nominal · UTIL


def extraer_sitio(ws_f, ws_v):
    def fila(r): return [ws_v[f"{c}{r}"].value for c in COLS]
    s = dict(kwh_punta=fila(7), d_punta=fila(9), dias_punta=fila(11), hrs_punta=fila(12),
             t_base=fila(14), t_punta=fila(16), t_cap=fila(17),
             bess_kw=ws_v["C28"].value, bess_kwh=ws_v["C29"].value,
             r72=ws_v["R72"].value, tir_bess=ws_v["C52"].value)
    qc, qd = [], None
    for r in range(60, 72):
        m = re.match(r"=([\d.]+)\*MIN\(1,\$C\$30/([\d.]+)\)", str(ws_f[f"Q{r}"].value) or "")
        if m: qc.append(float(m.group(1))); qd = float(m.group(2))
    s["q_design"], s["q_div"] = qc, qd
    return s


def irr(fl, lo=-0.9, hi=2.0):
    def npv(r): return sum(f / (1 + r) ** t for t, f in enumerate(fl))
    if npv(lo) * npv(hi) > 0: return None
    for _ in range(100):
        m = (lo + hi) / 2
        if npv(lo) * npv(m) <= 0: hi = m
        else: lo = m
    return (lo + hi) / 2


def modelo(s, E_nom, P=None, q_mode="exact", part=PART, cap_fisico=False):
    """Devuelve dict con bruto, capex, tir, van_inv, van_cli, ciclos para un BESS (P, E_nom)."""
    P = P or E_nom / 2
    E = E_nom * UTIL
    arb = cap = ncum = 0.0
    for i in range(12):
        N = min(E * s["dias_punta"][i], P * s["hrs_punta"][i] * s["dias_punta"][i], s["kwh_punta"][i])
        arb += N * s["t_punta"][i] - N / RTE * s["t_base"][i]
        if q_mode == "exact":
            Q = s["t_cap"][i] * min(s["d_punta"][i], E / BLOCK_H[i], P)
        else:
            Q = s["q_design"][i] * min(1, E / s["q_div"])
            if cap_fisico: Q = min(Q, s["t_cap"][i] * s["d_punta"][i])
        cap += Q; ncum += N
    bruto1 = (arb + cap) * (12 - MERMA) / 12
    capex = E_nom * 1000 * PRECIO_BESS * FX
    inv, cli = [-capex], [0.0]
    for t in range(1, HORIZONTE + 1):
        g = bruto1 * (1 + ESC) ** (t - 1) * (1 - DEG) ** (t - 1)
        pago = g * part if t <= PLAZO else 0.0
        inv.append(pago - OM * E_nom * (1 + ESC) ** (t - 1) if t <= PLAZO else 0.0)
        cli.append(g - pago)
    van = lambda fl: fl[0] + sum(f / (1 + WACC) ** t for t, f in enumerate(fl[1:], 1))
    return dict(E=E_nom, P=P, R_year=arb + cap, bruto1=bruto1, capex=capex, tir=irr(inv),
                van_inv=van(inv), van_cli=sum(f / (1 + WACC) ** t for t, f in enumerate(cli)),
                cli_a1=cli[1], ciclos=ncum / E_nom)


def tam_B(s):
    """Regla verano-2h: E_útil = 2h × demanda punta promedio may–sep; nominal a 100 kWh."""
    return round(2 * sum(s["d_punta"][:5]) / 5 / UTIL / 100) * 100


def tam_C(s):
    """Mayor E (pasos 100 kWh) cuyo kWh marginal rinde TIR ≥ WACC al inversionista."""
    best = 200
    for E in range(200, int(s["bess_kwh"] * 1.6), 100):
        m1, m2 = modelo(s, E), modelo(s, E + 100)
        dif = [-(m2["capex"] - m1["capex"])] + [
            (m2["bruto1"] - m1["bruto1"]) * PART * (1 + ESC) ** (t - 1) * (1 - DEG) ** (t - 1)
            - OM * 100 * (1 + ESC) ** (t - 1) for t in range(1, PLAZO + 1)]
        t = irr(dif)
        if t is not None and t >= WACC: best = E + 100
    return best


if __name__ == "__main__":
    wf = openpyxl.load_workbook(BOOK_4S, data_only=False)
    wv = openpyxl.load_workbook(BOOK_4S, data_only=True)
    print(f"{'sitio':12s} {'esc':2s} {'kWh':>7s} {'TIR inv':>8s} {'VANinv M$':>10s} {'cli a1 M$':>10s} {'VANcli M$':>10s} {'ciclos':>7s}")
    for n in [x for x in wf.sheetnames if x not in ("Resumen", "Supuestos")]:
        s = extraer_sitio(wf[n], wv[n])
        a = modelo(s, s["bess_kwh"], s["bess_kw"])
        dv = abs(a["R_year"] - s["r72"]) / s["r72"] * 100
        assert dv < 0.05, f"{n}: réplica difiere del libro {dv:.2f}% — revisar"
        for k, E in [("A", s["bess_kwh"]), ("B", tam_B(s)), ("C", tam_C(s))]:
            m = modelo(s, E)
            print(f"{n:12s} {k:2s} {m['E']:7,.0f} {m['tir']*100:7.2f}% {m['van_inv']/1e6:10.2f} "
                  f"{m['cli_a1']/1e6:10.2f} {m['van_cli']/1e6:10.2f} {m['ciclos']:7.0f}")
    print("\n(A verificado vs libro; ver análisis 2026-07-07-gepp-bess-verano-resizing y el Comparativo xlsx)")
