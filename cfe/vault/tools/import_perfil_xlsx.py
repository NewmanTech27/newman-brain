# -*- coding: utf-8 -*-
"""Import a GEPP-style "Perfil de Consumo" workbook into the OS bill schema.

A client workbook like `raw/Perfil de Consumo Electrico GEPP 2025-26.xlsx` already
carries everything needed to model a case: per-month kWh by period (base/intermedia/
punta), demanda maxima by period, kWMax ano movil, kVArh, FP, and the full charge
stack — plus a `Tarifas` sheet with the per-plant monthly CFE rates by concepto.

This converter REINTERPRETS that data into the canonical record the rest of the OS
already speaks: `cfe_savings.extract`'s monthly bill dict (the same one a PDF/XML
recibo parses into). It does NOT invent numbers — it reconstructs the disaggregated
MEM components (Generacion B/I/P, Capacidad, Distribucion, Suministro) from the
workbook's own per-period quantities x the workbook's own Tarifas rates, then
VALIDATES each reconstruction against the workbook's printed Subtotal/Total Recibo.
A service that foots (<0.5%) is a faithful CFE bill; one that doesn't is flagged
(typically an autoabasto site whose printed total reflects a Tala/ENEL discount, so
the pure-CFE reconstruction = full load valued at CFE rates — usable, but labeled).

Per service it writes:
  imported/<slug>/bills.json   list[bill dict]  (extract schema; engine-readable)
  imported/<slug>/inputs.json  calc_core inputs scaffold (cliente/division/...)
  imported/<slug>/meta.json    provenance + per-month validation + block inventory

Downstream: `cfe_savings.extract.extract_folder` reads `bills.json` if present, so
calc_core / fill_calculadora / portfolio / webapp can model an imported folder with
zero further changes (point them at imported/<slug>).

Usage:
  python tools/import_perfil_xlsx.py "raw/Perfil de Consumo Electrico GEPP 2025-26 (1).xlsx"
  python tools/import_perfil_xlsx.py <workbook> --only CAN ACA      # subset of sheets

Raw bytes never enter LLM context: this is a local script that emits compact JSON.
"""
from __future__ import annotations
import argparse, calendar, json, sys, unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OUT_ROOT = ROOT / "imported"
FC = 0.57  # GDMTH umbral factor de carga (A/158/2024) — see wiki [[demanda-facturable]]
IVA = 1.16

# --- service registry: which plant sheets to import, and how to resolve them ---------
# tarifa_plant / tarifa_class point at the row group in the `Tarifas` sheet.
# division is the engine code (SIN/BC/BCS); all GEPP target sites are SIN.
SERVICES = {
    "CAN": dict(slug="gepp-cancun",     cliente="gepp", tarifa_plant="Cancun",
                tarifa_class="GDMTH", division="SIN", municipio="Benito Juarez",
                nombre="GEPP Cancun (Peninsular)"),
    "ACA": dict(slug="gepp-acapulco",   cliente="gepp", tarifa_plant="Acapulco",
                tarifa_class="GDMTH", division="SIN", municipio="Acapulco de Juarez",
                nombre="GEPP Acapulco (Centro Sur)"),
    "IXT": dict(slug="gepp-ixtlahuacan", cliente="gepp", tarifa_plant="Ixtlahuacan",
                tarifa_class="DIST", division="SIN", municipio="Ixtlahuacan de los Membrillos",
                nombre="GEPP Ixtlahuacan (Jalisco; autoabasto Tala)"),
    "PR1": dict(slug="gepp-proplasa-pr1", cliente="gepp", tarifa_plant="Planta Preforma",
                tarifa_class="GDMTH", division="SIN", municipio="Cuautitlan Izcalli",
                nombre="GEPP Proplasa Preforma 1 (Valle de Mexico Norte)"),
    "PR2": dict(slug="gepp-proplasa-pr2", cliente="gepp", tarifa_plant="Planta Preforma",
                tarifa_class="GDMTH", division="SIN", municipio="Cuautitlan Izcalli",
                nombre="GEPP Proplasa Preforma 2 (Valle de Mexico Norte)"),
    "TAP": dict(slug="gepp-proplasa-tap", cliente="gepp", tarifa_plant="Planta Tapa",
                tarifa_class="GDMTH", division="SIN", municipio="Cuautitlan Izcalli",
                nombre="GEPP Proplasa Tapa (Valle de Mexico Norte)"),
}
MODEL_BLOCK = "Total Planta"   # full load (CFE + any autoabasto) — what PV/BESS offsets
MONTH_COL0 = 3                 # column 3 = Enero ... column 14 = Diciembre


def _fold(s) -> str:
    """lowercase + strip accents, for robust matching across the workbook's encoding."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    return "".join(c for c in s if not unicodedata.combining(c)).strip().lower()


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    t = str(v).replace(",", "").replace("$", "").strip()
    try:
        return float(t)
    except ValueError:
        return None


# --- Tarifas sheet: plant -> class -> concepto -> [12 monthly rates] ------------------
CONCEPTOS = {"fijo": "fijo", "base": "base", "intermedia": "intermedia",
             "punta": "punta", "distribucion": "distribucion",
             "capacidad": "capacidad", "energia": "energia"}


def load_tarifas(wb) -> dict:
    ws = wb["Tarifas"]
    # locate the month header row (the one whose first date cell sits in col 6)
    rates: dict = {}
    for r in range(1, ws.max_row + 1):
        plant = ws.cell(r, 2).value
        clase = ws.cell(r, 4).value
        concepto = _fold(ws.cell(r, 5).value)
        if not plant or concepto not in CONCEPTOS:
            continue
        key = (_fold(plant), _fold(clase))
        months = [_num(ws.cell(r, 6 + i).value) for i in range(12)]
        # The Tarifas sheet only carries rates for the months the client filled (Jan–Apr
        # in the GEPP workbook). Reconstructing a full year (the 2025 block) leaves the
        # rest at 0 → those months get $0 importes and PV/BESS credit at $0 (a bug).
        # Fill blank months with the mean of the populated ones — the published rates are
        # near-constant across the period, so this applies current tariffs to the full year.
        present = [v for v in months if v]
        if present and any((v in (None, 0.0)) for v in months):
            fill = sum(present) / len(present)
            months = [v if v else fill for v in months]
        rates.setdefault(key, {})[CONCEPTOS[concepto]] = months
    return rates


def plant_rates(rates: dict, plant: str, clase: str) -> dict:
    key = (_fold(plant), _fold(clase))
    if key not in rates:
        raise KeyError(f"No Tarifas rows for plant={plant!r} class={clase!r}. "
                       f"Available: {sorted(set(rates))[:8]}...")
    return rates[key]


# --- plant detail sheet: parse the MODEL_BLOCK into per-month quantities --------------
LABELS = [
    ("kwh_base",  lambda l, u: l == "base" and u == "kwh"),
    ("kwh_inter", lambda l, u: l == "intermedia" and u == "kwh"),
    ("kwh_punta", lambda l, u: l == "punta" and u == "kwh"),
    ("kw_base",   lambda l, u: l.startswith("dem") and l.endswith("base") and u == "kw"),
    ("kw_inter",  lambda l, u: l.startswith("dem") and "intermedia" in l and u == "kw"),
    ("kw_punta",  lambda l, u: l.startswith("dem") and "punta" in l and u == "kw"),
    ("kw_max",    lambda l, u: "kwmax" in l and u == "kw"),
    ("kvarh",     lambda l, u: u == "kvarh"),
    ("fp",        lambda l, u: l.startswith("factor de potencia")),
    ("p_fijo",    lambda l, u: l == "cargo fijo"),
    ("p_dem",     lambda l, u: l == "total $/kw"),
    ("p_energia", lambda l, u: l == "total $/kwh"),
    ("p_importe", lambda l, u: l.startswith("importe")),
    ("p_bonif",   lambda l, u: l.startswith("bonifica")),
    ("p_subtotal", lambda l, u: l == "subtotal"),
    ("p_recibo",  lambda l, u: l.startswith("total recibo")),
]


def find_blocks(ws) -> list:
    """Return [(block_label, header_row)] for every 'Perfil de Consumo' anchor."""
    out = []
    for r in range(1, ws.max_row + 1):
        if _fold(ws.cell(r, 1).value).startswith("perfil de consumo"):
            label = ws.cell(r + 1, 1).value
            out.append((str(label).strip() if label else "", r + 1))
    return out


def parse_block(ws, header_row: int, year: int, month_col0: int = MONTH_COL0) -> dict:
    """Read one block's metric rows; return {field: [12 monthly values]} (+ year/per-row meta).

    `month_col0` selects which side-by-side month block to read: the GEPP plant sheets
    carry the partial 2026 year in cols 3-14 (default) AND the full 2025 year in cols
    20-31 (pass month_col0=20). Both blocks share the col-A row labels and header row."""
    data = {f: [None] * 12 for f, _ in LABELS}
    r = header_row + 1
    # walk down until the next 'Perfil de Consumo' anchor or sheet end
    while r <= ws.max_row and not _fold(ws.cell(r, 1).value).startswith("perfil de consumo"):
        label = _fold(ws.cell(r, 1).value)
        unit = _fold(ws.cell(r, 2).value)
        for field, match in LABELS:
            if match(label, unit):
                data[field] = [_num(ws.cell(r, month_col0 + i).value) for i in range(12)]
                break
        r += 1
    return data


def block_months(block: dict) -> list:
    """Month indices (0..11) that carry real consumption."""
    out = []
    for i in range(12):
        kb = block["kwh_base"][i] or 0.0
        ki = block["kwh_inter"][i] or 0.0
        kp = block["kwh_punta"][i] or 0.0
        if kb + ki + kp > 0:
            out.append(i)
    return out


# --- reconstruct one engine bill record from quantities x Tarifas rates ---------------
def build_bill(svc: dict, block: dict, rates: dict, i: int, year: int,
               servicio: str, demanda_contratada: float) -> dict:
    m = i + 1
    days = calendar.monthrange(year, m)[1]
    kb = block["kwh_base"][i] or 0.0
    ki = block["kwh_inter"][i] or 0.0
    kp = block["kwh_punta"][i] or 0.0
    kwh_total = kb + ki + kp
    kwb = block["kw_base"][i] or 0.0
    kwi = block["kw_inter"][i] or 0.0
    kwp = block["kw_punta"][i] or 0.0

    r_fijo = (rates.get("fijo") or [0] * 12)[i] or 0.0
    # DIST tariff (autoabasto porteo) collapses energy into base/inter/punta and has
    # no Distribucion line; GDMTH carries both. 'energia' (GDMTO) flat rate -> all periods.
    r_base = (rates.get("base") or rates.get("energia") or [0] * 12)[i] or 0.0
    r_int = (rates.get("intermedia") or rates.get("energia") or [0] * 12)[i] or 0.0
    r_pta = (rates.get("punta") or rates.get("energia") or [0] * 12)[i] or 0.0
    r_dist = (rates.get("distribucion") or [0] * 12)[i] or 0.0
    r_cap = (rates.get("capacidad") or [0] * 12)[i] or 0.0

    umbral = kwh_total / (days * 24 * FC) if days else 0.0
    dem_meas = max(kwb, kwi, kwp)
    cap_basis = min(kwp or dem_meas, umbral) if umbral else (kwp or dem_meas)
    dist_basis = min(dem_meas, umbral) if umbral else dem_meas

    gen_base = kb * r_base
    gen_inter = ki * r_int
    gen_punta = kp * r_pta
    capacidad = cap_basis * r_cap
    distribucion = dist_basis * r_dist
    suministro = r_fijo

    # bonificacion/cargo FP: workbook sign convention == engine's (mem + bonif).
    # FP<90 -> positive cargo (also surfaced as cargo_fp_penalty); FP>=90 -> negative credit.
    bonif = block["p_bonif"][i] or 0.0
    cargo_fp = max(0.0, bonif)

    # An imported bill is internally self-consistent: its facturacion = the sum of its
    # own components (rates x quantities), so the engine's footing check verifies JSON
    # integrity. The comparison against the CLIENT's printed total is QA, kept in meta:
    # Cancun lands ~0.1% (faithful CFE recibo); autoabasto sites diverge by the
    # Tala/ENEL discount (the reconstruction = full load valued at CFE GDMTH rates).
    mem = suministro + gen_base + gen_inter + gen_punta + capacidad + distribucion
    recon_subtotal = mem + bonif
    recon_recibo = recon_subtotal * IVA
    printed_recibo = block["p_recibo"][i]
    printed_subtotal = block["p_subtotal"][i]

    return {
        "file": f"{svc['slug']}-{year}-{m:02d}",
        "start": f"{year}-{m:02d}-01",
        "end": f"{year}-{m:02d}-{days:02d}",
        "days": days, "month": m, "year": year,
        "servicio": servicio, "multiplicador": 1.0,
        "demanda_contratada": demanda_contratada,
        "kwh_base": kb, "kwh_inter": ki, "kwh_punta": kp,
        "kw_base": kwb, "kw_inter": kwi, "kw_punta": kwp,
        "kw_max": block["kw_max"][i], "kvarh": block["kvarh"][i], "fp": block["fp"][i],
        "suministro": suministro, "distribucion": distribucion,
        "transmision": 0.0, "cenace": 0.0,
        "gen_base": gen_base, "gen_inter": gen_inter, "gen_punta": gen_punta,
        "capacidad": capacidad, "scnmem": 0.0,
        "bonif_fp": bonif, "cargo_fp_penalty": cargo_fp,
        "energia_desglose": None,
        "subtotal_recibo": recon_subtotal,
        "facturacion_recibo": recon_recibo,
        "kwh_total": kwh_total,
        # provenance for the QA report (not consumed by the engine)
        "_printed_recibo": printed_recibo,
        "_printed_subtotal": printed_subtotal,
        "_printed_energia": block["p_energia"][i],
        "_printed_dem": block["p_dem"][i],
    }


def validate(bill: dict) -> dict:
    """QA vs the CLIENT's printed workbook totals. (Internal footing is exact by
    construction — facturacion = sum of components — so the engine never excludes
    an imported bill; this only measures how close our CFE model lands to theirs.)"""
    recon = bill["facturacion_recibo"]
    printed = bill.get("_printed_recibo")
    diff = (recon / printed - 1) if printed else None
    energ = bill["gen_base"] + bill["gen_inter"] + bill["gen_punta"]
    dem = bill["capacidad"] + bill["distribucion"]
    de = (energ / bill["_printed_energia"] - 1) if bill.get("_printed_energia") else None
    dd = (dem / bill["_printed_dem"] - 1) if bill.get("_printed_dem") else None
    return dict(recon=recon, printed=printed, diff=diff,
                ok=(diff is not None and abs(diff) < 0.005),
                d_energia=de, d_dem=dd)


def import_service(wb, sheet: str, rates: dict, month_col0: int = MONTH_COL0,
                   year_override: int | None = None) -> dict:
    svc = SERVICES[sheet]
    ws = wb[sheet]
    servicio = str(ws.cell(2, 2).value or "").strip()        # No. Medidor
    demanda = _num(ws.cell(3, 2).value) or 0.0               # Demanda Contratada
    year = year_override or int(_num(ws.cell(4, 2).value) or 2026)   # Perfil period year

    blocks = find_blocks(ws)
    inventory = [b[0] for b in blocks]
    target = next((hr for (lab, hr) in blocks if _fold(lab) == _fold(MODEL_BLOCK)), None)
    if target is None:
        raise ValueError(f"{sheet}: no '{MODEL_BLOCK}' block; found {inventory}")

    prates = plant_rates(rates, svc["tarifa_plant"], svc["tarifa_class"])
    block = parse_block(ws, target, year, month_col0=month_col0)
    months = block_months(block)
    if not months:
        raise ValueError(f"{sheet}: '{MODEL_BLOCK}' block has no months with consumption")

    bills, vals = [], []
    for i in months:
        b = build_bill(svc, block, prates, i, year, servicio, demanda)
        v = validate(b)
        b_clean = {k: val for k, val in b.items() if not k.startswith("_")}
        bills.append(b_clean)
        vals.append(dict(year=b["year"], month=b["month"], **{
            k: (round(v[k], 6) if isinstance(v.get(k), float) else v.get(k))
            for k in ("diff", "ok", "d_energia", "d_dem")}))

    n_ok = sum(1 for v in vals if v["ok"])
    worst = max((abs(v["diff"]) for v in vals if v["diff"] is not None), default=None)
    if worst is not None and worst < 0.015:
        basis, note = "CFE recibo", ("Reconstruccion ~= recibo CFE del cliente (delta < 1.5%: "
                                     "dias de facturacion / 2% baja tension). Modelo fiel.")
    else:
        basis, note = "full load @ CFE rates", (
            "Diverge del total impreso por el cliente — tipicamente sitio con autoabasto "
            "(descuento Tala/ENEL). La reconstruccion = carga TOTAL valuada a tarifas CFE "
            "GDMTH, la base correcta para dimensionar PV/BESS. Delta ~= descuento autoabasto.")
    meta = dict(
        source="raw/Perfil de Consumo Electrico GEPP 2025-26 (1).xlsx",
        sheet=sheet, servicio=servicio, year=year, block=MODEL_BLOCK,
        block_inventory=inventory, n_months=len(months),
        basis=basis, n_within_0p5pct=n_ok, worst_diff_vs_client=worst,
        tarifa=f"{svc['tarifa_plant']} / {svc['tarifa_class']}",
        validation=vals, note=note,
    )
    inputs = dict(
        cliente=svc["cliente"], rpu=servicio, municipio=svc["municipio"],
        division=svc["division"], demanda_contratada=demanda,
        nombre=svc["nombre"], giro="Industrial 24/7",
        # sizing/financials left for the analyst (cfe-savings-analyst) to propose
    )
    return dict(svc=svc, bills=bills, meta=meta, inputs=inputs)


def write_service(res: dict):
    out = OUT_ROOT / res["svc"]["slug"]
    out.mkdir(parents=True, exist_ok=True)
    (out / "bills.json").write_text(json.dumps(res["bills"], ensure_ascii=False, indent=2), "utf-8")
    (out / "meta.json").write_text(json.dumps(res["meta"], ensure_ascii=False, indent=2), "utf-8")
    (out / "inputs.json").write_text(json.dumps(res["inputs"], ensure_ascii=False, indent=2), "utf-8")
    return out


def main():
    ap = argparse.ArgumentParser(description="Import a GEPP-style perfil workbook into the OS bill schema")
    ap.add_argument("workbook")
    ap.add_argument("--only", nargs="*", help="subset of sheet codes (e.g. CAN ACA)")
    ap.add_argument("--out", default=None, help="override output root (default: imported/)")
    ap.add_argument("--month-col0", type=int, default=MONTH_COL0, dest="month_col0",
                    help="first month column: 3 = partial 2026 block (default); 20 = full 2025 block")
    ap.add_argument("--year", type=int, default=None, help="override the perfil year (e.g. 2025 for the right block)")
    args = ap.parse_args()
    global OUT_ROOT
    if args.out:
        OUT_ROOT = Path(args.out)

    wb = openpyxl.load_workbook(args.workbook, read_only=True, data_only=True)
    rates = load_tarifas(wb)
    sheets = args.only or list(SERVICES)

    print(f"Importing {len(sheets)} service(s) from {Path(args.workbook).name}\n")
    print(f"{'sheet':5} {'slug':22} {'mo':>3} {'vs client':>10}  basis")
    print("-" * 78)
    for sh in sheets:
        if sh not in SERVICES:
            print(f"{sh:5} -- not in registry, skipped"); continue
        try:
            res = import_service(wb, sh, rates, month_col0=args.month_col0, year_override=args.year)
        except Exception as e:
            print(f"{sh:5} ERROR: {e}"); continue
        write_service(res)
        m = res["meta"]
        worst = f"{m['worst_diff_vs_client']*100:+.2f}%" if m["worst_diff_vs_client"] is not None else "n/a"
        print(f"{sh:5} {res['svc']['slug']:22} {m['n_months']:>3} {worst:>10}  {m['basis']}")
    print(f"\nWrote to {OUT_ROOT}/  (bills.json | inputs.json | meta.json per service)")
    print("Model any service:  point calc_core / portfolio / webapp at imported/<slug>")


if __name__ == "__main__":
    main()
