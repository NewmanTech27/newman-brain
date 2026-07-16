# -*- coding: utf-8 -*-
"""Auto-fill the generic calculator from a folder of CFE bills (PDF or CFDI XML).

Usage:
  python tools/fill_calculadora.py raw/bills/780881200029
  python tools/fill_calculadora.py raw/bills/456220800389 --giro "Industrial 24/7" --kwp 700 --bess-kwh 1700 --bess-kw 850

What it auto-derives (and writes into a COPY of the generic calculator):
  BILLS  : 12 months of kWh/kW/importes/bonif-cargo FP/dias (validated to foot first)
  INPUTS : RPU, demanda contratada, division (CP->estado), municipio,
           monthly solar yield (CP -> municipio -> Solar_index_geografico),
           % autoconsumo derivado (curva x campana solar) si kWp >= 700
What it never guesses: giro, sizing, deal terms — CLI flags or edit in Excel.
Doctrine: a bill that doesn't foot (>0.5%) stops the run (use --force to override).
Output: "entregables/calculadoras/Calculadora - <RPU>.xlsx" + expected results printed via calc_core.
"""
from __future__ import annotations
import argparse, re, shutil, sys
from datetime import date
from pathlib import Path

import openpyxl

TOOLS = Path(__file__).resolve().parent
ROOT = TOOLS.parent
sys.path.insert(0, str(TOOLS))
from cfe_savings import extract
import calc_core, solar_lookup, load_curves

TEMPLATE = ROOT / "entregables" / "calculadoras" / "Calculadora Ahorros PV-BESS (generica).xlsx"


def sniff_cp_pdf(path: str) -> str | None:
    """Best-effort CP from a CFE PDF (raw bytes stay here)."""
    import pdfplumber
    with pdfplumber.open(path) as pdf:
        text = pdf.pages[0].extract_text() or ""
    m = re.search(r"C\.?P\.?\s*:?\s*(\d{5})", text)
    return m.group(1) if m else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("bills_dir")
    ap.add_argument("--out", default=None)
    ap.add_argument("--giro", default="Otro")
    ap.add_argument("--kwp", type=float, default=None)
    ap.add_argument("--bess-kwh", type=float, default=None)
    ap.add_argument("--bess-kw", type=float, default=None)
    ap.add_argument("--division", default=None)
    ap.add_argument("--force", action="store_true")
    a = ap.parse_args()

    bills = extract.extract_folder(a.bills_dir)
    if not bills:
        sys.exit(f"No se encontraron PDFs/XMLs en {a.bills_dir}")
    if len(bills) > 12:
        bills = bills[-12:]
        print("AVISO: más de 12 recibos — usando los 12 más recientes")

    # 1. validate footing (doctrine: stop on errors)
    bad = []
    for b in bills:
        v = calc_core.validate_bill(b)
        tag = "OK " if v["ok"] else "FAIL"
        print(f"{tag} {b['year']}-{b['month']:02d}  MEM {v['mem']:>13,.2f}  bonif/cargo {v['bonif']:>11,.2f}  "
              f"calc {v['calc_total']:>13,.2f}  recibo {v['recibo'] or 0:>13,.2f}  "
              f"diff {100*(v['diff'] or 0):+.3f}%")
        if not v["ok"]:
            bad.append(b["file"])
    if bad and not a.force:
        sys.exit(f"\nDETENIDO: {len(bad)} recibo(s) no cuadran (>0.5%): {bad}\n"
                 "Posible error de CFE o de parseo — revisar antes de continuar (--force para ignorar).")

    rpu = next((b.get("servicio") for b in bills if b.get("servicio")), "")
    dcon = max((b.get("demanda_contratada") or 0 for b in bills), default=0)

    # 2. CP -> municipio -> yield -> division
    cp = next((b.get("cp") for b in bills if b.get("cp")), None)
    if not cp:
        import glob, os
        for pdf in sorted(glob.glob(os.path.join(a.bills_dir, "*.pdf"))):
            cp = sniff_cp_pdf(pdf)
            if cp:
                break
    yld = solar_lookup.cp_to_yield(cp) if cp else None
    division = a.division or (yld["division"] if yld else "SIN")
    municipio = f"{yld['municipio']}, {yld['estado']}" if yld else ""
    if yld:
        print(f"\nCP {cp} -> {yld['match']} -> {yld['anual']:,.0f} kWh/kWp/año (prefactibilidad)")
    else:
        print("\nAVISO: sin CP detectable — rendimiento solar queda el del template; editar en INPUTS")

    # 3. fill a copy of the calculator
    out = Path(a.out) if a.out else ROOT / "entregables" / "calculadoras" / f"Calculadora - {rpu or Path(a.bills_dir).name}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(TEMPLATE, out)
    wb = openpyxl.load_workbook(out)
    wsB, wsI = wb["BILLS"], wb["INPUTS"]
    for r in range(2, 14):  # clear demo rows
        for c in [1, 2, 4, 5, 6, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 21, 24]:
            wsB.cell(row=r, column=c, value=None)
    for i, b in enumerate(bills):
        r = 2 + i
        wsB.cell(row=r, column=1, value=date(b["year"], b["month"], 1)).number_format = "mmm-yy"
        vals = {2: b["days"], 4: b["kwh_base"], 5: b["kwh_inter"], 6: b["kwh_punta"],
                8: b["kw_base"], 9: b["kw_inter"], 10: b["kw_punta"],
                11: b["suministro"], 12: b["distribucion"], 13: b["transmision"],
                14: b["cenace"], 15: b["gen_base"], 16: b["gen_inter"], 17: b["gen_punta"],
                18: b["capacidad"], 19: b["scnmem"], 21: b["bonif_fp"],
                24: b["facturacion_recibo"]}
        for c, v in vals.items():
            wsB.cell(row=r, column=c, value=v if v is not None else 0)

    def set_input(label, value):
        for r in range(2, 60):
            if wsI.cell(row=r, column=1).value == label:
                wsI.cell(row=r, column=2, value=value)
                return
        raise KeyError(label)
    set_input("Nombre del cliente", f"RPU {rpu}")
    set_input("RPU / No. de servicio", rpu)
    set_input("Municipio / Estado", municipio)
    set_input("Division tarifaria", division)
    set_input("Giro del cliente", a.giro)
    if dcon:
        set_input("Demanda contratada (kW)", dcon)
    if a.kwp is not None:
        set_input("Capacidad PV (kWp)", a.kwp)
    if a.bess_kwh is not None:
        set_input("Capacidad nominal (kWh)", a.bess_kwh)
    if a.bess_kw is not None:
        set_input("Potencia (kW)", a.bess_kw)
    if yld:
        for m in range(1, 13):
            wsI.cell(row=2 + m, column=6, value=yld["monthly"][m])

    # 4. curve fit (diagnóstico) + % autoconsumo derivado si kWp >= 700
    kwp = a.kwp if a.kwp is not None else 194.48
    bkwh = a.bess_kwh if a.bess_kwh is not None else 2940.0
    bkw = a.bess_kw if a.bess_kw is not None else 1503.0
    ym = yld["monthly"] if yld else None
    if ym is None:  # template demo yield
        twb = openpyxl.load_workbook(TEMPLATE)
        ym = {m: twb["INPUTS"].cell(row=2 + m, column=6).value for m in range(1, 13)}

    fit = load_curves.fit_bills(bills, division)
    print("\nAjuste de curva vs recibos (MAE):" + "".join(
        f"\n  {r['giro']:<20} {r['mae']:.1%}" for r in fit["ranking"] if r["mae"] is not None)
        + f"\n  -> mejor ajuste: {fit['best']}")
    pct_flat = None
    if kwp >= 700:
        curva = a.giro if a.giro in load_curves.CURVES else (fit["best"] or "Otro")
        pcts = []
        for b in bills:
            g = kwp * ym.get(b["month"], 0)
            pcts.append((load_curves.autoconsumo_pct(curva, b["year"], b["month"],
                                                     b["kwh_total"], g), g))
        gtot = sum(g for _, g in pcts)
        pct_flat = sum(p * g for p, g in pcts) / gtot if gtot else 1.0
        set_input("Override % autoconsumo (vacio = default)", round(pct_flat, 4))
        print(f"% autoconsumo derivado (curva '{curva}' x campana solar, prom. ponderado): {pct_flat:.1%} — escrito como override en el Excel")
    wb.save(out)

    # 5. expected results via the shared core (mismo pct que el Excel)
    res = calc_core.compute(dict(kwp=kwp, bess_kwh=bkwh, bess_kw=bkw, division=division,
                                 giro=a.giro, demanda_contratada=dcon or None,
                                 pct_autoconsumo=pct_flat,
                                 yield_monthly=ym, rpu=rpu), bills)
    an, fin = res["annual"], res["finance"]
    print(f"\n=== ESPERADO al abrir el Excel (ano 0, bruto) — {len(bills)} meses ===")
    print(f"Factura ANTES real: {an['antes']:,.2f}")
    for k, lbl in [("pv_only", "PV solo"), ("bess_only", "BESS solo"), ("hibrido", "Hibrido")]:
        f = fin["PV" if k == "pv_only" else "BESS" if k == "bess_only" else "Hibrido"]
        tir = f"{f['tir_proyecto']:.1%}" if f["tir_proyecto"] is not None else "n/a"
        print(f"{lbl:>10}: {an[k]:>13,.2f} ({an[k]/an['antes']:.1%})  TIR proy {tir}  payback {f['payback_proyecto']}")
    if an["fp_cargo"] > 0:
        print(f"+ Correccion FP disponible: {an['fp_cargo']:,.2f}/ano (palanca separada)")
    print(f"Regimen: {res['regimen']['regimen']}")
    for lvl, msg in res["regimen"]["alerts"]:
        print(f"  [{lvl}] {msg}")
    print(f"\nEscrito: {out}")


if __name__ == "__main__":
    main()
