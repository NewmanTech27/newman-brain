# -*- coding: utf-8 -*-
"""Standard "Solución Energética" Excel book — the default calculadora deliverable.

Generalizes the GEPP book (entregables/calculadoras/GEPP - Solucion Energetica ...)
into ONE parameterized builder. Feed it one or more projects:
  - a CFE bills folder  raw/bills/<RPU>/      (PDF or CFDI XML — has CP -> yield)
  - an imported folder  imported/<slug>/      (bills.json from a consumption workbook)
Single project  -> Resumen Ejecutivo + Supuestos + one site sheet.
Group (>=2 projects sharing `cliente`) -> the above per project PLUS a roll-up:
  Comparativo Comercial + Proyección 20 años (agregada).

Fidelity = HYBRID (CLAUDE.md decision 2026-06-18):
  - The monthly savings motor (umbral, BESS shave, arbitraje) is computed by the
    deterministic engine (calc_core.compute — golden 18/18) and written as VALUES.
  - Supuestos assumptions, CAPEX, the 20-year projection, =IRR()/=NPV(), and all
    cross-sheet roll-ups are LIVE FORMULAS, so the levers a decision-maker edits
    (PPA $/kWh, PPA term, EPC costs, escalations, comisión, WACC) recalc on edit.
  - Editing SIZING (PV kWp / BESS / FP) in Supuestos requires a re-run (the savings
    values are static); the sheet labels those cells accordingly.

The engine is never modified. Numbers in this book reconcile to calc_core to the peso.

CLI:
  python tools/make_project_book.py raw/bills/780881200029
  python tools/make_project_book.py imported/gepp-* --cliente gepp
  python tools/make_project_book.py raw/bills/<RPU> --kwp 700 --bess-kwh 1700 --bess-kw 850 --fp
  # per-project sizing/yield overrides for curated group books:
  python tools/make_project_book.py imported/gepp-cancun imported/gepp-acapulco \
      --cliente gepp --config gepp_sizing.json
"""
from __future__ import annotations
import argparse, glob, json, math, sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

TOOLS = Path(__file__).resolve().parent
ROOT = TOOLS.parent
sys.path.insert(0, str(TOOLS))
from cfe_savings import extract, sizing
import calc_core, ppa_pricer, solar_lookup, optimize_sizing

MES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

# ---------------------------------------------------------------- styling
BLUE = PatternFill("solid", fgColor="DCE6F1")      # editable input (financial — safe to edit live)
AMBER = PatternFill("solid", fgColor="FDE9D9")     # editable input (sizing — re-run engine after edit)
HEAD = PatternFill("solid", fgColor="1F4E5F")      # section header band
SUB = PatternFill("solid", fgColor="D9D9D9")       # column sub-header
TOTAL = PatternFill("solid", fgColor="F2F2F2")
WHITE_BOLD = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '#,##0'
MONEY2 = '#,##0.00'
PCT = '0.0%'
NUM = '#,##0'


def _hdr(ws, row, text, span=11):
    c = ws.cell(row=row, column=1, value=text)
    c.fill, c.font = HEAD, WHITE_BOLD
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = HEAD


def _money(ws, row, col, val, fmt=MONEY):
    c = ws.cell(row=row, column=col, value=val)
    c.number_format = fmt
    return c


# ---------------------------------------------------------------- project load

def _load_inputs(folder: Path) -> dict:
    f = folder / "inputs.json"
    return json.loads(f.read_text(encoding="utf-8")) if f.exists() else {}


def load_project(path: str, config: dict, cli: dict) -> dict | None:
    folder = Path(path)
    bills = extract.extract_folder(str(folder))
    if not bills:
        print(f"  [SKIP] {folder.name}: sin recibos/bills.json")
        return None
    if len(bills) > 12:
        bills = bills[-12:]
    # footing validation (doctrine: exclude non-footing bills)
    good = [b for b in bills if calc_core.validate_bill(b)["ok"]]
    bad = [f"{b['year']}-{b['month']:02d}" for b in bills if not calc_core.validate_bill(b)["ok"]]
    if bad and not cli.get("force"):
        print(f"  [WARN] {folder.name}: {len(bad)} recibo(s) no cuadran {bad} — excluidos")
    bills = good or bills
    if not bills:
        print(f"  [SKIP] {folder.name}: ningún recibo cuadra")
        return None

    inp = _load_inputs(folder)
    cfg = config.get(folder.name, {}) or config.get(inp.get("rpu", ""), {})
    rpu = inp.get("rpu") or next((b.get("servicio") for b in bills if b.get("servicio")), folder.name)
    name = cfg.get("nombre") or inp.get("nombre") or f"RPU {rpu}"
    division = cfg.get("division") or inp.get("division") or "SIN"
    giro = cfg.get("giro") or inp.get("giro") or "Otro"
    dcon = cfg.get("demanda_contratada") or inp.get("demanda_contratada") \
        or max((b.get("demanda_contratada") or 0 for b in bills), default=0)
    cliente = cfg.get("cliente") or inp.get("cliente") or cli.get("cliente") or ""

    ym, ysrc = optimize_sizing.resolve_yield({**inp, **cfg}, bills)

    # explicit sizing (CLI > config > inputs.json flat/nested) wins; otherwise the
    # max-NPV sweep decides — NO automatic 0.7 MW cap (CLAUDE.md 2026-06-19).
    sysd = inp.get("system") or {}
    def explicit(*keys):
        for k in keys:
            for src in (cli, cfg, inp, sysd):
                if src.get(k) is not None:
                    return float(src[k])
        return None
    kwp_x = explicit("kwp", "pv_kwp")
    bkwh_x = explicit("bess_kwh", "bess_nominal_kwh")
    bkw_x = explicit("bess_kw", "bess_power_kw")
    area = cfg.get("area_m2") or inp.get("area_m2")
    if kwp_x is None:
        base_inp = dict(division=division, giro=giro, yield_monthly=ym,
                        demanda_contratada=dcon or None)
        opt = optimize_sizing.propose(bills, base_inp, area_m2=area)["recomendado"]
        kwp = opt["kwp"]
        bess_kwh = bkwh_x if bkwh_x is not None else float(opt["bess_kwh"])
        bess_kw = bkw_x if bkw_x is not None else float(opt["bess_kw"])
        sizing_src = f"sweep NPV ({opt['regimen']})"
    else:
        be = sizing.propose_bess(bills, division, None)
        kwp = kwp_x
        bess_kwh = bkwh_x if bkwh_x is not None else float(be["proposed_nominal_kwh"])
        bess_kw = bkw_x if bkw_x is not None else float(be["proposed_power_kw"])
        sizing_src = "explícito"
    fp = bool(cli.get("fp") or cfg.get("fp_correction") or inp.get("fp_correction"))

    engine_inputs = dict(
        cliente=cliente, rpu=rpu, municipio=inp.get("municipio", ""), division=division,
        giro=giro, demanda_contratada=dcon or None, fp_correction=fp,
        kwp=kwp, bess_kwh=bess_kwh, bess_kw=bess_kw, yield_monthly=ym,
        ppa=cfg.get("ppa"), ppa_yrs=cfg.get("ppa_yrs"), comision=cfg.get("comision"),
        # 2026-term overrides (additive; defaults hold when key absent)
        epc_usd_wp=cfg.get("epc_usd_wp"), fx=cfg.get("fx"),
        falla_meses=cfg.get("falla_meses"), esc_ppa=cfg.get("esc_ppa"),
        escenario="Hibrido",
    )
    res = calc_core.compute({k: v for k, v in engine_inputs.items() if v is not None}, bills)

    # seed a PPA rate that hits the target financier IRR (live cell default)
    try:
        ppa_seed = ppa_pricer.solve_ppa_rate(
            res["inputs"], res["annual"], "Hibrido",
            cli.get("target_irr", 0.18), res["inputs"]["ppa_yrs"],
            res["inputs"]["esc_ppa"], res["inputs"]["comision"]) or res["inputs"]["ppa"]
    except Exception:
        ppa_seed = res["inputs"]["ppa"]

    p = res["inputs"]
    return dict(folder=folder.name, rpu=rpu, name=name, cliente=cliente, division=division,
                giro=giro, dcon=dcon, n_months=len(bills), yield_src=ysrc, sizing_src=sizing_src,
                kwp=kwp, bess_kwh=bess_kwh, bess_kw=bess_kw, fp=fp,
                capacitor=float(cfg.get("capacitor", 0.0)), ppa_seed=round(ppa_seed, 4),
                ppa_yrs=int(p["ppa_yrs"]), comision=float(p["comision"]),
                epc_pv=p["epc_usd_wp"], epc_bess=p["epc_bess"], res=res, p=p)


# ---------------------------------------------------------------- Supuestos

GLOBALS = [  # (label, key in inputs p, fmt)
    ("EPC PV (USD/Wp)", "epc_usd_wp", MONEY2),
    ("Tipo de cambio (MXN/USD)", "fx", MONEY2),
    ("EPC BESS (USD/kWh)", "epc_bess", MONEY),
    ("WACC / descuento", "wacc", PCT),
    ("Escalación CFE / año", "esc_cfe", PCT),
    ("Escalación PPA / año", "esc_ppa", PCT),
    ("Disponibilidad (factor SaaS)", "falla", PCT),
    ("Degradación PV / año", "pv_degr", PCT),
    ("Degradación BESS / año", "bess_degr", PCT),
    ("Comisión BESS financiador", "comision", PCT),
    ("Horizonte (años)", "horizon", NUM),
    ("FC umbral (demanda facturable)", "fc", '0.00'),
    ("DoD batería", "dod", PCT),
    ("RTE batería (round-trip)", "rte", PCT),
    ("O&M PV (MXN/kWp-año)", "_ompv_ratio", MONEY),
    ("O&M BESS (MXN/kWh-año)", "_ombess_ratio", MONEY),
    ("Fianza (sobre pago PPA)", "fianza", PCT),
    ("Plazo BESS / comisión (años)", "bess_yrs", NUM),
]
GROW0 = 4  # first global row

MATRIX_HDR = ["Proyecto", "PV kWp", "BESS kW", "BESS kWh", "FP corr",
              "PV USD/Wp", "BESS USD/kWh", "Capacitor $", "PPA $/kWh",
              "PPA plazo", "Comisión %", "TIR compra", "TIR financiador",
              "Ahorro bruto/año $"]
MROW0 = GROW0 + len(GLOBALS) + 3  # first matrix data row offset (header sits 2 above)


def build_supuestos(wb, projects):
    ws = wb.create_sheet("Supuestos")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    for col in "BCDEFGHIJKLMN":
        ws.column_dimensions[col].width = 14
    ws["A1"] = "SUPUESTOS — Panel de control en vivo"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = ("Celdas AZUL = financieras (editar recalcula todo en vivo). "
                "Celdas ÁMBAR = dimensionamiento (cambiar requiere re-correr el motor).")
    ws["A2"].font = Font(italic=True, size=9)

    _hdr(ws, GROW0 - 1, "A. SUPUESTOS GLOBALES", span=3)
    g = projects[0]["p"]
    for i, (lbl, key, fmt) in enumerate(GLOBALS):
        r = GROW0 + i
        ws.cell(row=r, column=1, value=lbl)
        if key == "_ompv_ratio":
            val = 310.0
        elif key == "_ombess_ratio":
            val = 180.0
        else:
            val = g[key]
        c = ws.cell(row=r, column=2, value=val)
        c.number_format, c.fill, c.border = fmt, BLUE, BORDER

    # per-project matrix
    mhdr = MROW0 - 1
    _hdr(ws, MROW0 - 2, "B. VARIABLES POR PROYECTO  (AZUL = financiero · ÁMBAR = dimensionamiento)",
         span=len(MATRIX_HDR))
    for j, h in enumerate(MATRIX_HDR):
        c = ws.cell(row=mhdr, column=1 + j, value=h)
        c.fill, c.font, c.alignment = SUB, BOLD, Alignment(wrap_text=True, vertical="center")
    rowmap = {}
    for k, proj in enumerate(projects):
        r = MROW0 + k
        rowmap[proj["folder"]] = r
        sheet = proj["sheet"]
        ws.cell(row=r, column=1, value=proj["name"])
        vals = [proj["kwp"], proj["bess_kw"], proj["bess_kwh"],
                "Sí" if proj["fp"] else "No", proj["epc_pv"], proj["epc_bess"],
                proj["capacitor"], proj["ppa_seed"], proj["ppa_yrs"], proj["comision"]]
        fills = [AMBER, AMBER, AMBER, AMBER, BLUE, BLUE, BLUE, BLUE, BLUE, BLUE]
        fmts = [NUM, NUM, NUM, None, MONEY2, MONEY, MONEY, MONEY2, NUM, PCT]
        for j, (v, fl, fm) in enumerate(zip(vals, fills, fmts)):
            c = ws.cell(row=r, column=2 + j, value=v)
            c.fill, c.border = fl, BORDER
            if fm:
                c.number_format = fm
        # back-references to the site sheet results
        q = f"'{sheet}'"
        ws.cell(row=r, column=12, value=f"={q}!$B${TIRC}").number_format = PCT
        ws.cell(row=r, column=13, value=f"={q}!$B${TIRF}").number_format = PCT
        ws.cell(row=r, column=14, value=f"={q}!$J${RES_TOTAL}").number_format = MONEY
    # PORTAFOLIO total row
    rt = MROW0 + len(projects)
    ws.cell(row=rt, column=1, value="PORTAFOLIO").font = BOLD
    lo, hi = MROW0, MROW0 + len(projects) - 1
    for col in (2, 3, 4, 8, 14):
        cc = get_column_letter(col)
        c = ws.cell(row=rt, column=col, value=f"=SUM({cc}{lo}:{cc}{hi})")
        c.number_format, c.font = (NUM if col in (2, 3, 4) else MONEY), BOLD
        c.fill = TOTAL
    return ws, rowmap


# ---------------------------------------------------------------- site sheet
# fixed row layout (cross-sheet formulas depend on these constants)
P_PARAM = 6          # parameter row (refs into Supuestos)
C1 = 8               # §1 cost/savings header
CAP = 18             # §2 CAPEX header
CAP_TOTAL = 22       # CAPEX TOTAL row
RES_HDR = 24         # §3 Resumen mensual header
RES_COL = 25         # column headers
RES_M0 = 26          # first month row (26..37)
RES_TOTAL = 38       # TOTAL row
PRJ_NOTE = 39        # ann_factor note (B39 = annualization factor)
PRJ_BAND = 40        # §4 section band
PRJ_HDR = 41         # projection column headers
PRJ_Y0 = 42          # year 0
PRJ_Y1 = 43          # year 1 (..62 = year 20)
TIRC = 64            # B64 = TIR compra
TIRF = 65            # B65 = TIR financiador
CUMCLI = 66          # B66 = acumulado cliente PPA 20a
NARR = 68            # §5 narrative


def build_site_sheet(wb, proj, sup_row):
    ws = wb.create_sheet(proj["sheet"])
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    for col in "BCDEFGHIJK":
        ws.column_dimensions[col].width = 13
    S = "Supuestos"
    sr = sup_row  # this project's row in the Supuestos matrix
    res, p = proj["res"], proj["p"]
    ann = res["annual"]
    monthly = res["monthly"]
    n = proj["n_months"]
    annf = 12.0 / n if n else 1.0

    ws["A1"] = f"{proj['name']} — Memoria de cálculo y propuesta energética"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (f"RPU {proj['rpu']} · GDMTH {proj['division']} · "
                f"Demanda contratada {proj['dcon']:,.0f} kW · {proj['n_months']} meses de recibo")
    ws["A2"].font = Font(italic=True, size=9)
    reg = res["regimen"]["regimen"]
    ws["A3"] = f"Régimen: {reg}"
    ws["A3"].font = Font(size=9, color="1F4E5F")

    # parameters (live refs to Supuestos)
    _hdr(ws, P_PARAM - 1, "PARÁMETROS (desde Supuestos — edítalos ahí)")
    pc = 1
    for lbl, frm, fmt in [("PV kWp", f"={S}!$B${sr}", NUM),
                          ("BESS kW", f"={S}!$C${sr}", NUM),
                          ("BESS kWh", f"={S}!$D${sr}", NUM),
                          ("FP corrección", f"={S}!$E${sr}", None),
                          ("Entrega útil kWh", f"={S}!$D${sr}*{S}!$B${GROW0+12}*SQRT({S}!$B${GROW0+13})", NUM)]:
        ws.cell(row=P_PARAM, column=pc, value=lbl).font = BOLD
        c = ws.cell(row=P_PARAM, column=pc + 1, value=frm)
        if fmt:
            c.number_format = fmt
        pc += 2

    # §1 cost & savings
    _hdr(ws, C1, "1. ESTRUCTURA DE COSTO Y AHORRO (pre-IVA, año base · valores del motor)")
    for j, h in enumerate(["Concepto", "$ / año", "Nota"]):
        ws.cell(row=C1 + 1, column=1 + j, value=h).fill = SUB
        ws.cell(row=C1 + 1, column=1 + j).font = BOLD
    fp_year = ann["fp_cargo"] if proj["fp"] else 0.0
    rows1 = [
        ("Tarifa Actual (CFE pleno, sin proyecto)", ann["antes"] * annf, "Σ recibos × anualización"),
        ("(−) Ahorro PV (energía + demanda)", (ann["ah_pv"] + ann["dem_pv"]) * annf, "intermedio + umbral"),
        ("(−) Ahorro BESS (punta + arbitraje)", (ann["hibrido"] - ann["ah_pv"] - ann["dem_pv"]) * annf, "recorte punta + arbitraje (incl. claw)"),
        ("(−) Ahorro FP (corrección)", fp_year * annf, "multa factor de potencia" if proj["fp"] else "no aplicada"),
    ]
    for i, (lbl, val, nota) in enumerate(rows1):
        r = C1 + 2 + i
        ws.cell(row=r, column=1, value=lbl)
        _money(ws, r, 2, val)
        ws.cell(row=r, column=3, value=nota).font = Font(size=8, italic=True)
    rb = C1 + 2  # tarifa actual row
    ws.cell(row=C1 + 6, column=1, value="(=) Costo neto con proyecto").font = BOLD
    _money(ws, C1 + 6, 2, f"=B{rb}-B{rb+1}-B{rb+2}-B{rb+3}").font = BOLD
    ws.cell(row=C1 + 7, column=1, value="Ahorro proyecto / año").font = BOLD
    _money(ws, C1 + 7, 2, f"=B{rb+1}+B{rb+2}+B{rb+3}").font = BOLD
    ws.cell(row=C1 + 8, column=1, value="Ahorro %").font = BOLD
    ws.cell(row=C1 + 8, column=2, value=f"=B{C1+7}/B{rb}").number_format = PCT

    # §2 CAPEX (live from Supuestos)
    _hdr(ws, CAP, "2. CAPEX (fórmulas vivas desde Supuestos)")
    ws.cell(row=CAP + 1, column=1, value="CAPEX PV")
    _money(ws, CAP + 1, 2, f"={S}!$B${sr}*1000*{S}!$F${sr}*{S}!$B${GROW0+1}")
    ws.cell(row=CAP + 2, column=1, value="CAPEX BESS")
    _money(ws, CAP + 2, 2, f"={S}!$D${sr}*{S}!$G${sr}*{S}!$B${GROW0+1}")
    ws.cell(row=CAP + 3, column=1, value="CAPEX capacitor (FP)")
    _money(ws, CAP + 3, 2, f"={S}!$H${sr}")
    ws.cell(row=CAP_TOTAL, column=1, value="CAPEX TOTAL").font = BOLD
    _money(ws, CAP_TOTAL, 2, f"=B{CAP+1}+B{CAP+2}+B{CAP+3}").font = BOLD

    # §3 Resumen mensual (engine VALUES)
    _hdr(ws, RES_HDR, "3. RESUMEN MENSUAL — Tarifa Actual vs Con Proyecto (valores del motor)")
    cols = ["Mes", "kWh Consumo", "Gen PV kWh", "BESS desc kWh", "Tarifa Actual $",
            "Ah PV $", "Ah BESS $", "Ah FP $", "Costo c/proy $", "Ahorro $", "Ahorro %"]
    for j, h in enumerate(cols):
        c = ws.cell(row=RES_COL, column=1 + j, value=h)
        c.fill, c.font, c.alignment = SUB, BOLD, Alignment(wrap_text=True)
    for i in range(12):
        r = RES_M0 + i
        if i < len(monthly):
            m = monthly[i]
            ah_pv = m["ah_pv"] + m["dem_pv"]
            ah_bess = m["hibrido"] - m["ah_pv"] - m["dem_pv"]  # incl claw
            ah_fp = m["fp_cargo"] if proj["fp"] else 0.0
            ahorro = m["hibrido"] + ah_fp
            ws.cell(row=r, column=1, value=f"{MES[m['month']-1]}-{str(m['year'])[2:]}")
            _money(ws, r, 2, m["kwh_total"], NUM)
            _money(ws, r, 3, m["gen"], NUM)
            _money(ws, r, 4, m["desc"], NUM)
            _money(ws, r, 5, m["antes"])
            _money(ws, r, 6, ah_pv)
            _money(ws, r, 7, ah_bess)
            _money(ws, r, 8, ah_fp)
            _money(ws, r, 9, m["antes"] - ahorro)
            _money(ws, r, 10, ahorro)
            ws.cell(row=r, column=11, value=f"=J{r}/E{r}").number_format = PCT
    # TOTAL row (live sums)
    ws.cell(row=RES_TOTAL, column=1, value="TOTAL").font = BOLD
    for col in range(2, 11):
        cc = get_column_letter(col)
        c = ws.cell(row=RES_TOTAL, column=col, value=f"=SUM({cc}{RES_M0}:{cc}{RES_M0+11})")
        c.number_format, c.font, c.fill = (NUM if col <= 4 else MONEY), BOLD, TOTAL
    c = ws.cell(row=RES_TOTAL, column=11, value=f"=J{RES_TOTAL}/E{RES_TOTAL}")
    c.number_format, c.font, c.fill = PCT, BOLD, TOTAL

    # §4 projection (live formulas — mirrors calc_core._finance)
    ws.cell(row=PRJ_NOTE, column=1, value="Factor anualización (12/meses):").font = Font(size=9, italic=True)
    ws.cell(row=PRJ_NOTE, column=2, value=round(annf, 4)).number_format = '0.000'
    if n < 12:
        ws.cell(row=PRJ_NOTE, column=3,
                value=f"Solo {n} meses de recibo → bases año 1 anualizadas ×{annf:.2f} (riesgo estacionalidad)")
        ws.cell(row=PRJ_NOTE, column=3).font = Font(size=8, color="C00000")
    phdr = ["Año", "Tarifa Actual $", "Ah PV $", "Ah BESS $", "Ah FP $", "Costo c/proy $",
            "Flujo Compra $", "Pago PPA $", "Comisión $", "Flujo Cliente PPA $", "Flujo Financiador $"]
    _hdr(ws, PRJ_BAND, "4. PROYECCIÓN 20 AÑOS — Comprar vs PPA/Ahorro compartido")
    for j, h in enumerate(phdr):
        c = ws.cell(row=PRJ_HDR, column=1 + j, value=h)
        c.fill, c.font, c.alignment = SUB, BOLD, Alignment(wrap_text=True, vertical="center")

    # cell refs into Supuestos globals
    g = lambda off: f"{S}!$B${GROW0+off}"
    wacc, esccfe, escppa = g(3), g(4), g(5)
    disp, dpv_r, dbe_r = g(6), g(7), g(8)
    ompv_r, ombe_r, fianza_r, bessyrs = g(14), g(15), g(16), g(17)
    ppa_c, ppayrs_c, comm_c = f"{S}!$I${sr}", f"{S}!$J${sr}", f"{S}!$K${sr}"
    af = f"$B${PRJ_NOTE}"  # ann factor
    # year-1 bases (annualized) from the Resumen TOTAL row
    b_tar = f"$E${RES_TOTAL}*{af}"
    b_pv = f"$F${RES_TOTAL}*{af}"
    b_bess = f"$G${RES_TOTAL}*{af}"
    b_fp = f"$H${RES_TOTAL}*{af}"
    b_gen = f"$C${RES_TOTAL}*{af}"
    kwp_c = f"{S}!$B${sr}"
    bkwh_c = f"{S}!$D${sr}"

    # year 0
    ws.cell(row=PRJ_Y0, column=1, value=0)
    _money(ws, PRJ_Y0, 7, f"=-B{CAP_TOTAL}")
    _money(ws, PRJ_Y0, 11, f"=-B{CAP_TOTAL}")
    for yy in range(1, 21):
        r = PRJ_Y1 + (yy - 1)
        e = f"(1+{esccfe})^{yy}"
        dpv = f"(1-{dpv_r})^{yy}"
        dbe = f"(1-{dbe_r})^{yy}"
        epp = f"(1+{escppa})^{yy-1}"
        eom = f"(1+{esccfe})^{yy-1}"
        # O&M gated exactly like calc_core._finance (PV over PPA term, BESS over bess term)
        om_term = (f"IF({yy}<={ppayrs_c},{ompv_r}*{kwp_c}*{eom},0)"
                   f"+IF({yy}<={bessyrs},{ombe_r}*{bkwh_c}*{eom},0)")
        ws.cell(row=r, column=1, value=yy)
        _money(ws, r, 2, f"={b_tar}*{e}")                                   # Tarifa Actual
        _money(ws, r, 3, f"={b_pv}*{e}*{dpv}")                              # Ah PV
        _money(ws, r, 4, f"={b_bess}*{e}*{dbe}*{disp}")                     # Ah BESS (net of avail)
        _money(ws, r, 5, f"={b_fp}*{e}")                                    # Ah FP
        _money(ws, r, 6, f"=B{r}-C{r}-D{r}-E{r}")                           # Costo c/proy
        _money(ws, r, 7, f"=C{r}+D{r}+E{r}-({om_term})")                    # Flujo Compra (proyecto)
        # Pago PPA only while yy<=PPA plazo
        _money(ws, r, 8, f"=IF({yy}<={ppayrs_c},{b_gen}*{dpv}*{ppa_c}*{epp},0)")
        _money(ws, r, 9, f"=IF({yy}<={bessyrs},D{r}*{comm_c},0)")           # Comisión on BESS savings
        _money(ws, r, 10, f"=C{r}+D{r}+E{r}-H{r}-I{r}")                     # Flujo Cliente PPA
        _money(ws, r, 11, f"=H{r}+I{r}-({om_term})-H{r}*{fianza_r}")        # Flujo Financiador
    # IRR / NPV
    ws.cell(row=TIRC, column=1, value="TIR compra (proyecto):").font = BOLD
    ws.cell(row=TIRC, column=2, value=f"=IRR(G{PRJ_Y0}:G{PRJ_Y1+19})").number_format = PCT
    ws.cell(row=TIRC, column=3, value="VPN compra @WACC:").font = Font(size=9)
    _money(ws, TIRC, 4, f"=G{PRJ_Y0}+NPV({wacc},G{PRJ_Y1}:G{PRJ_Y1+19})")
    ws.cell(row=TIRF, column=1, value="TIR financiador:").font = BOLD
    ws.cell(row=TIRF, column=2, value=f"=IRR(K{PRJ_Y0}:K{PRJ_Y1+19})").number_format = PCT
    ws.cell(row=TIRF, column=3, value="VPN cliente PPA @WACC:").font = Font(size=9)
    _money(ws, TIRF, 4, f"=NPV({wacc},J{PRJ_Y1}:J{PRJ_Y1+19})")
    ws.cell(row=CUMCLI, column=1, value="Acumulado cliente PPA 20a:").font = BOLD
    _money(ws, CUMCLI, 2, f"=SUM(J{PRJ_Y1}:J{PRJ_Y1+19})").font = BOLD
    ws.cell(row=CUMCLI, column=3, value="Acumulado compra 20a:").font = Font(size=9)
    _money(ws, CUMCLI, 4, f"=SUM(G{PRJ_Y0}:G{PRJ_Y1+19})")

    # §5 why this sizing is optimal
    _hdr(ws, NARR, "5. POR QUÉ ESTE DIMENSIONAMIENTO ES ÓPTIMO")
    exento = proj["kwp"] < 700
    narr = []
    narr.append(
        f"PV {proj['kwp']:,.0f} kWp — "
        + ("generador exento <0.7 MW (medición neta, sin permiso CNE)." if exento
           else "autoconsumo ≥0.7 MW (permiso simplificado CNE, excedentes solo a CFE).")
        + f" En {proj['division']} la generación solar cae en intermedio → recorta energía intermedia; "
        "no genera en punta, por lo que el BESS es la palanca de demanda.")
    narr.append(
        f"BESS {proj['bess_kw']:,.0f} kW / {proj['bess_kwh']:,.0f} kWh — recorta el pico de punta "
        "(arbitraje carga en base, descarga en punta, topado al kWh punta del recibo).")
    narr.append(
        f"Dimensionado al consumo (rend. {proj['yield_src']}); excedentes bajo medición neta expiran "
        "en 12 meses, por eso no se sobredimensiona el PV.")
    if str(proj.get("sizing_src", "")).startswith("sweep"):
        narr.append(
            f"PV/BESS dimensionados por **barrido de máximo NPV del cliente** (NO se topó a 0.7 MW): "
            f"el óptimo cae en {proj['kwp']:,.0f} kWp / {proj['bess_kw']:,.0f} kW de BESS ({proj['sizing_src']}). "
            "Más allá del punto de autoconsumo solar el excedente se valúa ~$0 y el NPV cae; el barrido "
            "compara explícitamente quedarse exento <0.7 MW vs cruzar a autoconsumo.")
    if proj["fp"]:
        narr.append("Corrección de factor de potencia incluida — palanca separada, capacitores a cargo "
                    "del financiador en el esquema PPA (el cliente conserva 100% del ahorro FP).")
    narr.append("Ahorro mensual = valores del motor determinista (golden 18/18). Proyección, CAPEX, "
                "TIR y flujos PPA = fórmulas vivas; edita los Supuestos financieros y recalcula.")
    for i, t in enumerate(narr):
        c = ws.cell(row=NARR + 1 + i, column=1, value="• " + t)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=NARR + 1 + i, start_column=1, end_row=NARR + 1 + i, end_column=11)
    return ws


# ---------------------------------------------------------------- roll-ups

def build_resumen(wb, projects):
    ws = wb.create_sheet("Resumen Ejecutivo", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    for col in "BCDEFGHIJK":
        ws.column_dimensions[col].width = 15
    ws["A1"] = ("Solución Energética — "
                + (projects[0]["cliente"].upper() if projects[0]["cliente"] else projects[0]["name"]))
    ws["A1"].font = Font(bold=True, size=15)
    ws["A2"] = "Resumen ejecutivo · pre-IVA · valores del motor (golden 18/18) · finanzas en vivo desde Supuestos"
    ws["A2"].font = Font(italic=True, size=9)

    _hdr(ws, 4, "RECOMENDACIÓN POR PROYECTO")
    cols = ["Sitio", "Tarifa Actual $/año", "Ahorro bruto $/año", "Ahorro %",
            "Inversión $", "TIR compra", "PPA $/kWh", "Benef. cliente PPA a1 $"]
    for j, h in enumerate(cols):
        c = ws.cell(row=5, column=1 + j, value=h)
        c.fill, c.font, c.alignment = SUB, BOLD, Alignment(wrap_text=True)
    r0 = 6
    for k, proj in enumerate(projects):
        r = r0 + k
        q = f"'{proj['sheet']}'"
        ws.cell(row=r, column=1, value=proj["name"])
        _money(ws, r, 2, f"={q}!$E${RES_TOTAL}*{q}!$B${PRJ_NOTE}")
        _money(ws, r, 3, f"={q}!$J${RES_TOTAL}*{q}!$B${PRJ_NOTE}")
        ws.cell(row=r, column=4, value=f"=C{r}/B{r}").number_format = PCT
        _money(ws, r, 5, f"={q}!$B${CAP_TOTAL}")
        ws.cell(row=r, column=6, value=f"={q}!$B${TIRC}").number_format = PCT
        ws.cell(row=r, column=7, value=f"=Supuestos!$I${MROW0+k}").number_format = MONEY2
        _money(ws, r, 8, f"={q}!$J${PRJ_Y1}")
    rt = r0 + len(projects)
    ws.cell(row=rt, column=1, value="PORTAFOLIO").font = BOLD
    for col in (2, 3, 5, 8):
        cc = get_column_letter(col)
        c = ws.cell(row=rt, column=col, value=f"=SUM({cc}{r0}:{cc}{rt-1})")
        c.number_format, c.font, c.fill = MONEY, BOLD, TOTAL
    c = ws.cell(row=rt, column=4, value=f"=C{rt}/B{rt}")
    c.number_format, c.font, c.fill = PCT, BOLD, TOTAL

    # KPI cards
    _hdr(ws, rt + 2, "INDICADORES CLAVE")
    kpis = [("Gasto eléctrico actual $/año", f"=B{rt}"),
            ("Ahorro proyecto $/año", f"=C{rt}"),
            ("Ahorro % del gasto", f"=C{rt}/B{rt}"),
            ("Inversión total (compra) $", f"=E{rt}")]
    for j, (lbl, frm) in enumerate(kpis):
        ws.cell(row=rt + 3, column=1 + j * 2, value=lbl).font = Font(size=9, bold=True)
        c = ws.cell(row=rt + 4, column=1 + j * 2, value=frm)
        c.number_format = PCT if "%" in lbl else MONEY
        c.font = Font(size=12, bold=True, color="1F4E5F")
    return ws


def build_comparativo(wb, projects):
    ws = wb.create_sheet("Comparativo Comercial")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 24
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 16
    ws["A1"] = "Comparativo Comercial — Comprar vs PPA/Ahorro compartido"
    ws["A1"].font = Font(bold=True, size=13)
    _hdr(ws, 3, "RESULTADO POR PROYECTO (referencias vivas a cada hoja)", span=8)
    cols = ["Proyecto", "Compra benef/año $", "Compra inversión $", "Compra TIR",
            "PPA benef cliente a1 $", "PPA TIR financiador", "Ahorro bruto/año $", "Ahorro %"]
    for j, h in enumerate(cols):
        c = ws.cell(row=4, column=1 + j, value=h)
        c.fill, c.font, c.alignment = SUB, BOLD, Alignment(wrap_text=True)
    r0 = 5
    for k, proj in enumerate(projects):
        r = r0 + k
        q = f"'{proj['sheet']}'"
        ws.cell(row=r, column=1, value=proj["name"])
        _money(ws, r, 2, f"={q}!$C${PRJ_Y1}+{q}!$D${PRJ_Y1}+{q}!$E${PRJ_Y1}-({q}!$C${PRJ_Y1}+{q}!$D${PRJ_Y1}+{q}!$E${PRJ_Y1}-{q}!$G${PRJ_Y1})")
        _money(ws, r, 3, f"={q}!$B${CAP_TOTAL}")
        ws.cell(row=r, column=4, value=f"={q}!$B${TIRC}").number_format = PCT
        _money(ws, r, 5, f"={q}!$J${PRJ_Y1}")
        ws.cell(row=r, column=6, value=f"={q}!$B${TIRF}").number_format = PCT
        _money(ws, r, 7, f"={q}!$J${RES_TOTAL}*{q}!$B${PRJ_NOTE}")
        ws.cell(row=r, column=8, value=f"=G{r}/({q}!$E${RES_TOTAL}*{q}!$B${PRJ_NOTE})").number_format = PCT
    rt = r0 + len(projects)
    ws.cell(row=rt, column=1, value="PORTAFOLIO").font = BOLD
    for col in (2, 3, 5, 7):
        cc = get_column_letter(col)
        c = ws.cell(row=rt, column=col, value=f"=SUM({cc}{r0}:{cc}{rt-1})")
        c.number_format, c.font, c.fill = MONEY, BOLD, TOTAL
    ws.cell(row=rt + 2, column=1,
            value="Comprar maximiza el ahorro (TIR proyecto supera el WACC); PPA = cero inversión, "
                  "el financiador funda el CAPEX y cobra PPA+comisión a su TIR objetivo.").font = Font(italic=True, size=9)
    return ws


def build_proyeccion(wb, projects):
    ws = wb.create_sheet("Proyección 20 años")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 8
    for col in "BCDEFGHI":
        ws.column_dimensions[col].width = 18
    ws["A1"] = "Proyección 20 años — Portafolio (suma de hojas por proyecto)"
    ws["A1"].font = Font(bold=True, size=13)
    cols = ["Año", "Tarifa Actual $", "Ahorro total $", "Costo c/proy $",
            "Flujo Compra $", "Flujo Cliente PPA $", "Acum. Compra $", "Acum. Cliente PPA $"]
    for j, h in enumerate(cols):
        c = ws.cell(row=3, column=1 + j, value=h)
        c.fill, c.font, c.alignment = SUB, BOLD, Alignment(wrap_text=True)
    sheets = [f"'{p['sheet']}'" for p in projects]
    for yy in range(1, 21):
        r = 3 + yy
        src = PRJ_Y1 + (yy - 1)
        ws.cell(row=r, column=1, value=yy)
        _money(ws, r, 2, "=" + "+".join(f"{q}!B{src}" for q in sheets))
        _money(ws, r, 3, "=" + "+".join(f"({q}!C{src}+{q}!D{src}+{q}!E{src})" for q in sheets))
        _money(ws, r, 4, f"=B{r}-C{r}")
        _money(ws, r, 5, "=" + "+".join(f"{q}!G{src}" for q in sheets))
        _money(ws, r, 6, "=" + "+".join(f"{q}!J{src}" for q in sheets))
        _money(ws, r, 7, f"=G{r-1 if yy > 1 else r}+E{r}" if yy > 1 else f"=E{r}")
        _money(ws, r, 8, f"=H{r-1 if yy > 1 else r}+F{r}" if yy > 1 else f"=F{r}")
    return ws


# ---------------------------------------------------------------- main

def _slug_sheet(name: str, used: set) -> str:
    s = name[:28].replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "").replace(":", "")
    base, i = s, 1
    while s in used:
        i += 1
        s = f"{base[:26]} {i}"
    used.add(s)
    return s


def main():
    ap = argparse.ArgumentParser(description="Build the standard Solución Energética Excel book")
    ap.add_argument("projects", nargs="+", help="bills folders (raw/bills/<RPU>) or imported/<slug> folders")
    ap.add_argument("--cliente", default=None)
    ap.add_argument("--out", default=None)
    ap.add_argument("--config", default=None, help="JSON: {folder|rpu: {kwp,bess_kwh,bess_kw,fp_correction,ppa,ppa_yrs,comision,capacitor,nombre,...}}")
    ap.add_argument("--kwp", type=float, default=None)
    ap.add_argument("--bess-kwh", type=float, default=None, dest="bess_kwh")
    ap.add_argument("--bess-kw", type=float, default=None, dest="bess_kw")
    ap.add_argument("--fp", action="store_true")
    ap.add_argument("--target-irr", type=float, default=0.18, dest="target_irr")
    ap.add_argument("--force", action="store_true")
    a = ap.parse_args()

    # expand globs (PowerShell may not)
    paths = []
    for pat in a.projects:
        hits = glob.glob(pat)
        paths.extend(hits if hits else [pat])
    config = json.loads(Path(a.config).read_text(encoding="utf-8")) if a.config else {}
    cli = dict(cliente=a.cliente, kwp=a.kwp, bess_kwh=a.bess_kwh, bess_kw=a.bess_kw,
               fp=a.fp, target_irr=a.target_irr, force=a.force)

    print(f"Cargando {len(paths)} proyecto(s)...")
    projects = []
    for path in paths:
        proj = load_project(path, config, cli)
        if proj:
            projects.append(proj)
            an = proj["res"]["annual"]
            print(f"  [OK] {proj['name']}  PV {proj['kwp']:,.0f} kWp · BESS {proj['bess_kw']:,.0f} kW "
                  f"[{proj['sizing_src']}] · ahorro híbrido ${an['hibrido']:,.0f} "
                  f"({an['hibrido']/an['antes']:.1%}) · {proj['n_months']} meses")
    if not projects:
        sys.exit("No hay proyectos válidos para construir el libro.")

    group = len(projects) > 1
    used = set()
    for proj in projects:
        proj["sheet"] = _slug_sheet(proj["name"], used)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sup, rowmap = build_supuestos(wb, projects)
    for proj in projects:
        build_site_sheet(wb, proj, rowmap[proj["folder"]])
    build_resumen(wb, projects)
    if group:
        build_comparativo(wb, projects)
        build_proyeccion(wb, projects)
    # creation order already yields: Resumen, Supuestos, sites..., [Comparativo, Proyección]

    cli_name = (a.cliente or projects[0]["cliente"] or "").strip()
    if a.out:
        out = Path(a.out)
    elif group and cli_name:
        out = ROOT / "entregables" / "calculadoras" / f"{cli_name.upper()} - Solucion Energetica.xlsx"
    else:
        out = ROOT / "entregables" / "calculadoras" / f"{projects[0]['rpu']} - Solucion Energetica.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"\nEscrito: {out}")
    print(f"Hojas: {wb.sheetnames}")


if __name__ == "__main__":
    main()
