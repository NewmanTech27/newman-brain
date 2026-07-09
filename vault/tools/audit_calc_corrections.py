# Quantify the corrections for the 780881200029 calculadora audit:
#  (1) Arbitrage corrected: discharge only on punta weekdays (Mon-Fri minus festivos),
#      capped at the bill's actual punta kWh (can't displace energy that doesn't exist).
#  (2) FP bonificacion claw-back: the ~1.2% bonificacion shrinks with the bill.
# Reads bill volumes + implied rates from TARIFF_RATES; bonif from the real PDFs (engine extractor).
import openpyxl, sys, calendar, math
from datetime import date

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, r"C:\Users\vidan\Desktop\CFE Brain\tools")
from cfe_savings import extract

WB = r"C:\Users\vidan\Desktop\CFE Brain\raw\data\780881200029.xlsx"
USABLE_DAY = 2940 * 0.96 * math.sqrt(0.96)   # 2,765.38 kWh deliverable per cycle
RTE = 0.96

def festivos(y):
    # CFE/statutory holidays treated as non-punta: Jan 1, 1st Mon Feb, 3rd Mon Mar,
    # May 1, Sep 16, 3rd Mon Nov, Dec 25.
    def nth_monday(year, month, n):
        d = date(year, month, 1)
        off = (7 - d.weekday()) % 7  # days to first Monday
        return date(year, month, 1 + off + 7 * (n - 1))
    return {date(y,1,1), nth_monday(y,2,1), nth_monday(y,3,3), date(y,5,1),
            date(y,9,16), nth_monday(y,11,3), date(y,12,25)}

def punta_weekdays(y, m):
    fest = festivos(y)
    n = 0
    for d in range(1, calendar.monthrange(y, m)[1] + 1):
        dt = date(y, m, d)
        if dt.weekday() < 5 and dt not in fest:
            n += 1
    return n

wb = openpyxl.load_workbook(WB, data_only=True, read_only=True)
ws = wb["TARIFF_RATES"]

# bonif per billing month from the PDFs
from pathlib import Path
MONTH_KEY = {"ABR":(2025,4),"MAY":(2025,5),"JUN":(2025,6),"JUL":(2025,7),"AGO":(2025,8),
             "SEP":(2025,9),"OCT":(2025,10),"NOV":(2025,11),"DIC":(2025,12),
             "ENE":(2026,1),"FEB":(2026,2),"MAR":(2026,3)}
bonif = {}
for pdf in Path(r"C:\Users\vidan\Desktop\CFE Brain\raw\bills\780881200029").glob("*.pdf"):
    tag = pdf.stem.split()[-1][:3]
    b = extract.parse_bill(str(pdf))
    bonif[MONTH_KEY[tag]] = b.get("bonif_fp", 0.0) or 0.0

print("Mes      | wd | punta kWh | desc VIEJO | desc CORR | arb VIEJO $ | arb CORR $ | bonif% | clawback $")
tot = dict(old_d=0.0, new_d=0.0, old_a=0.0, new_a=0.0, claw=0.0, dem=0.0, pv=0.0, subtotal=0.0)
for i in range(12):
    r1, r3 = 6 + i, 36 + i   # consumption row, rates row
    period = ws.cell(row=r1, column=1).value
    y, m = period.year, period.month
    days = ws.cell(row=r1, column=2).value
    kwh_p = ws.cell(row=r1, column=9).value
    rate_p = ws.cell(row=r3, column=13).value   # M: $/kWh punta TOTAL
    rate_b = ws.cell(row=r3, column=11).value   # K: $/kWh base TOTAL
    wd = punta_weekdays(y, m)
    old_disch = min(USABLE_DAY, 1503 * (4 if (m <= 3 or m >= 11) else 2)) * days
    old_charge = old_disch / RTE
    old_arb = old_disch * rate_p - old_charge * rate_b
    new_disch = min(USABLE_DAY * wd, kwh_p)
    new_charge = new_disch / RTE
    new_arb = new_disch * rate_p - new_charge * rate_b
    # claw-back applies to ALL savings of the month (demand+arb+pv) at the bonif share
    # demand & pv per DISPATCH (read straight from sheet)
    wsd = wb["DISPATCH"]
    dem = wsd.cell(row=92 + i, column=13).value     # M: ahorro BESS demanda
    pv = wsd.cell(row=92 + i, column=15).value      # O: ahorro PV
    subtotal_real = None
    bon = bonif.get((y, m), 0.0)
    mem_subtotal = wb["BASELINE_COST"].cell(row=4 + i, column=9).value  # I: subtotal sin IVA
    bonif_share = abs(bon) / mem_subtotal if mem_subtotal else 0.0
    sav_corr = dem + pv + new_arb
    claw = sav_corr * bonif_share
    tot["old_d"] += old_disch; tot["new_d"] += new_disch
    tot["old_a"] += old_arb;   tot["new_a"] += new_arb
    tot["claw"] += claw; tot["dem"] += dem; tot["pv"] += pv
    print(f"{y}-{m:02d}  | {wd} | {kwh_p:>9,.0f} | {old_disch:>10,.0f} | {new_disch:>9,.0f} | "
          f"{old_arb:>11,.0f} | {new_arb:>10,.0f} | {bonif_share:.2%} | {claw:>8,.0f}")

print(f"\nTOTALES:")
print(f"  Descarga anual: vieja {tot['old_d']:,.0f} kWh -> corregida {tot['new_d']:,.0f} kWh ({tot['new_d']/tot['old_d']-1:+.1%})")
print(f"  Arbitraje:      viejo  ${tot['old_a']:,.0f} -> corregido ${tot['new_a']:,.0f}  (delta ${tot['new_a']-tot['old_a']:,.0f})")
print(f"  Demanda BESS (sin cambio): ${tot['dem']:,.0f}")
print(f"  PV (sin cambio):           ${tot['pv']:,.0f}")
print(f"  Claw-back bonificacion FP: -${tot['claw']:,.0f}")
gross_old = tot['dem'] + tot['pv'] + tot['old_a']
gross_new = tot['dem'] + tot['pv'] + tot['new_a'] - tot['claw']
print(f"  AHORRO BRUTO: viejo ${gross_old:,.0f} (25.2%) -> corregido ${gross_new:,.0f} ({gross_new/30157370.98:.1%} del baseline MEM)")
