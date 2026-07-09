# -*- coding: utf-8 -*-
# Build "Calculadora Ahorros PV-BESS (generica).xlsx" — generic, client-agnostic
# PV / BESS / hybrid savings calculator for CFE GDMTH, webapp-ready architecture.
#
# Replaces the intent of raw/data/780881200029.xlsx (client v1, errors E1-E5 per
# wiki/analyses/2026-06-11-780881200029-calculadora-audit.md) with the corrected
# v2 formula set, generalized:
#   - Any client: 12-month bill entry, profile-driven assumptions (giro, division)
#   - Three scenarios side by side: PV-only / BESS-only / Hibrido
#   - Regulatory regime auto-detection + savings caps:
#       < 0.7 MW  -> Generador Exento / medicion neta (PV credit capped at kWh consumed)
#       >= 0.7 MW -> Autoconsumo (CNE): only % instantaneous self-consumption credited,
#                    excedentes at TEXC (default $0, CFE-only), SAE backup mandate flagged
#   - FP correction as a separate additive lever (never blended)
#   - FINANCE: full doctrine — proyecto unlevered vs financiador, labeled
# Demo client preloaded: 780881200029 (Grupo Posadas Cancun) -> must reproduce the
# engine golden numbers (Ahorro hibrido $7,083,252 / 23.49% MEM; arb $790,048.75;
# claw -$84,528.03).
#
# Every formula is mirrored in python; run prints expected values for verification.
import math, calendar, sys
from datetime import date
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

ROOT = Path(__file__).resolve().parent.parent
V2 = ROOT / "780881200029 - Calculadora v2.xlsx"
OUT = ROOT / "Calculadora Ahorros PV-BESS (generica).xlsx"

# ---------------- demo data (from the validated v2 rebuild) ----------------
wb2 = openpyxl.load_workbook(V2)
wsB2 = wb2["BILLS"]
months = []
for r in range(2, 14):
    g = lambda c: wsB2.cell(row=r, column=c).value
    months.append(dict(
        per=g(1), days=g(2), wd=g(3), kwh_b=g(4), kwh_i=g(5), kwh_p=g(6),
        kw_b=g(8), kw_i=g(9), kw_p=g(10), cargo_fijo=g(11), dist=g(12),
        trans=g(13), cenace=g(14), gen_b=g(15), gen_i=g(16), gen_p=g(17),
        capacidad=g(18), scnmem=g(19), bonif=g(21), recibo=g(24)))
wsI2 = wb2["INPUTS"]
HREF = 839.41
yield_kwp = {m: wsI2.cell(row=1 + m, column=6).value / HREF for m in range(1, 13)}

P = dict(kwp=194.48, pv_degr=0.005, epc_usd_wp=0.75, fx=17.55, om_pv=60288.8,
         ppa=1.5, ppa_yrs=15, esc_ppa=0.05,
         bess_kwh=2940.0, bess_kw=1503.0, dod=0.96, rte=0.96, bess_degr=0.0125,
         epc_bess=280.0, om_bess=529200.0, comision=0.5, falla_meses=3, bess_yrs=15,
         fianza=0.139, wacc=0.12, esc_cfe=0.06, horizon=20,
         fc=0.57, dcon=2000.0, pct_ac=0.80, texc=0.0)
P["capex_pv"] = P["kwp"] * 1000 * P["epc_usd_wp"] * P["fx"]
P["capex_bess"] = P["bess_kwh"] * P["epc_bess"] * P["fx"]
P["deliv"] = P["bess_kwh"] * P["dod"] * math.sqrt(P["rte"])
P["falla"] = (12 - P["falla_meses"]) / 12

def festivos(y):
    def nth_monday(year, month, n):
        d = date(year, month, 1); off = (7 - d.weekday()) % 7
        return date(year, month, 1 + off + 7 * (n - 1))
    return {date(y,1,1), nth_monday(y,2,1), nth_monday(y,3,3), date(y,5,1),
            date(y,9,16), nth_monday(y,11,3), date(y,12,25)}
FEST = sorted(set().union(*[festivos(y) for y in (2025, 2026, 2027)]))

PUNTA_H = {  # horas punta por dia habil, por division x mes (aprox mes completo)
    "SIN": {m: (2 if 4 <= m <= 10 else 4) for m in range(1, 13)},
    "BC":  {m: (4 if 5 <= m <= 10 else 0) for m in range(1, 13)},
    "BCS": {m: (10 if 4 <= m <= 10 else 0) for m in range(1, 13)},
}
GIROS = [("Hotel/Turismo", 0.80), ("Industrial 24/7", 0.90), ("Industrial 1 turno", 0.65),
         ("Comercial/Retail", 0.75), ("Oficinas", 0.60), ("Otro", 0.75)]

# ---------------- python mirror ----------------
def mirror(kwp, autoc_pct_applies):
    rows = []
    for mo in months:
        y, m = mo["per"].year, mo["per"].month
        kwh_tot = mo["kwh_b"] + mo["kwh_i"] + mo["kwh_p"]
        mem = (mo["cargo_fijo"] + mo["dist"] + mo["trans"] + mo["cenace"] + mo["gen_b"]
               + mo["gen_i"] + mo["gen_p"] + mo["capacidad"] + mo["scnmem"])
        umbral = kwh_tot / (mo["days"] * 24 * P["fc"])
        r_cap = mo["capacidad"] / min(mo["kw_p"], umbral)
        r_t, r_c, r_s = mo["trans"]/kwh_tot, mo["cenace"]/kwh_tot, mo["scnmem"]/kwh_tot
        bnd_b = mo["gen_b"]/mo["kwh_b"] + r_t + r_c + r_s
        bnd_i = mo["gen_i"]/mo["kwh_i"] + r_t + r_c + r_s
        bnd_p = mo["gen_p"]/mo["kwh_p"] + r_t + r_c + r_s
        ph = PUNTA_H["SIN"][m]
        gen = kwp * yield_kwp[m]
        usable = gen * (P["pct_ac"] if autoc_pct_applies else 1.0)
        aprov = min(usable, mo["kwh_i"])
        exced = gen - aprov
        ah_pv = aprov * bnd_i + exced * P["texc"]
        umb_post = (kwh_tot - aprov) / (mo["days"] * 24 * P["fc"])
        f_pre = min(mo["kw_p"], umbral)
        f_prebess = min(mo["kw_p"], umb_post)
        dem_pv = (f_pre - f_prebess) * r_cap
        share = -mo["bonif"] / mem
        claw_pv = -(ah_pv + dem_pv) * share if mo["bonif"] < 0 else 0.0
        pv_only = ah_pv + dem_pv + claw_pv
        shave = min(P["bess_kw"], P["deliv"] / ph) if ph else 0.0
        resid = max(0.0, mo["kw_p"] - shave)
        f_post_h = min(resid, umb_post)
        dem_bess_h = (f_prebess - f_post_h) * r_cap
        desc = min(P["deliv"] * mo["wd"], mo["kwh_p"]) if ph else 0.0
        carga = desc / P["rte"]
        arb = desc * bnd_p - carga * bnd_b
        sub_h = ah_pv + dem_pv + dem_bess_h + arb
        claw_h = -sub_h * share if mo["bonif"] < 0 else 0.0
        hib = sub_h + claw_h
        f_post_b = min(resid, umbral)
        dem_bess_o = (f_pre - f_post_b) * r_cap
        sub_b = dem_bess_o + arb
        claw_b = -sub_b * share if mo["bonif"] < 0 else 0.0
        bess_only = sub_b + claw_b
        rows.append(dict(mem=mem, antes=mem + mo["bonif"], gen=gen, ah_pv=ah_pv,
                         dem_pv=dem_pv, claw_pv=claw_pv, pv_only=pv_only,
                         dem_bess_h=dem_bess_h, desc=desc, arb=arb, claw_h=claw_h, hib=hib,
                         dem_bess_o=dem_bess_o, claw_b=claw_b, bess_only=bess_only))
    return rows

mir = mirror(P["kwp"], False)
S = lambda k: sum(r[k] for r in mir)
an = dict(pv_en=S("ah_pv"), dem_pv=S("dem_pv"), dem_bess=S("dem_bess_h"), arb=S("arb"),
          claw_h=S("claw_h"), hib=S("hib"), pv_only=S("pv_only"), bess_only=S("bess_only"),
          mem=S("mem"), antes=S("antes"), gen=S("gen"))

def finance(base_pv, base_bess, base_claw, gen_anual, capex, has_pv, has_bess):
    fl_new, fl_proj = [-capex], [-capex]
    for yy in range(1, P["horizon"] + 1):
        esc = (1 + P["esc_cfe"]) ** yy
        dpv = (1 - P["pv_degr"]) ** yy
        dbe = (1 - P["bess_degr"]) ** yy
        sav_pv = base_pv * esc * dpv
        sav_bess = base_bess * esc * dbe * P["falla"]
        claw = base_claw * esc * (dpv + dbe) / 2
        bruto = sav_pv + sav_bess + claw
        ppa_pay = (gen_anual * dpv * P["ppa"] * (1 + P["esc_ppa"]) ** (yy - 1)) if (has_pv and yy <= P["ppa_yrs"]) else 0.0
        comis = sav_bess * P["comision"] if (has_bess and yy <= P["bess_yrs"]) else 0.0
        om_pv = P["om_pv"] * (1 + P["esc_cfe"]) ** (yy - 1) if (has_pv and yy <= P["ppa_yrs"]) else 0.0
        om_bess = P["om_bess"] * (1 + P["esc_cfe"]) ** (yy - 1) if (has_bess and yy <= P["bess_yrs"]) else 0.0
        fianza = ppa_pay * P["fianza"]
        fl_new.append(ppa_pay + comis - om_pv - om_bess - fianza)
        fl_proj.append(bruto - om_pv - om_bess)
    return fl_new, fl_proj

def irr(flows, lo=-0.95, hi=3.0):
    def npv(r): return sum(f / (1 + r) ** i for i, f in enumerate(flows))
    if npv(lo) * npv(hi) > 0: return float("nan")
    for _ in range(200):
        mid = (lo + hi) / 2
        if npv(lo) * npv(mid) <= 0: hi = mid
        else: lo = mid
    return (lo + hi) / 2

fl_new, fl_proj = finance(an["pv_en"] + an["dem_pv"], an["dem_bess"] + an["arb"], an["claw_h"],
                          an["gen"], P["capex_pv"] + P["capex_bess"], True, True)
exp_tir_proj, exp_tir_new = irr(fl_proj), irr(fl_new)

print("=== MIRROR (demo, hibrido debe = golden) ===")
print(f"Ahorro PV energia : {an['pv_en']:,.2f}  (golden 787,522.28)")
print(f"Dem PV+BESS (hib) : {an['dem_pv']+an['dem_bess']:,.2f}  (golden 5,590,209.47)")
print(f"Arbitraje         : {an['arb']:,.2f}  (golden 790,048.75)")
print(f"Claw-back hibrido : {an['claw_h']:,.2f}  (golden -84,528.03)")
print(f"AHORRO HIBRIDO    : {an['hib']:,.2f}  (golden 7,083,252) = {an['hib']/an['mem']:.2%} MEM / {an['hib']/an['antes']:.2%} real")
print(f"AHORRO PV-ONLY    : {an['pv_only']:,.2f}")
print(f"AHORRO BESS-ONLY  : {an['bess_only']:,.2f}")
print(f"TIR proyecto {exp_tir_proj:.2%} (esp ~34.9%) | TIR financiador {exp_tir_new:.2%} (esp ~15.0%)")
mir_ac = mirror(1000.0, True)
print(f"[test regimen] kWp=1000 autoconsumo 80%: hibrido {sum(r['hib'] for r in mir_ac):,.2f}")

# ---------------- workbook ----------------
wb = openpyxl.Workbook()
F_IN = Font(color="0000FF", name="Arial", size=10)          # input azul
F_FX = Font(color="000000", name="Arial", size=10)          # formula negra
F_LK = Font(color="008000", name="Arial", size=10)          # link otra hoja verde
F_H = Font(bold=True, name="Arial", size=10)
F_T = Font(bold=True, name="Arial", size=12)
FILL_H = PatternFill("solid", start_color="DDEBF7")
FILL_S = PatternFill("solid", start_color="FCE4D6")
FILL_W = PatternFill("solid", start_color="FFF2CC")
NF_M = "$#,##0;($#,##0);-"
NF_M2 = "$#,##0.00;($#,##0.00);-"
NF_K = "#,##0;(#,##0);-"
NF_K1 = "#,##0.0;(#,##0.0);-"
NF_P = "0.0%"
NF_R = "$0.0000"

def hdr(ws, row, headers, col0=1, fill=FILL_H):
    for c, h in enumerate(headers, col0):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = F_H; cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

# ---- AUX ----
wsX = wb.active; wsX.title = "AUX"
wsX["A1"] = "Giro"; wsX["B1"] = "% autoconsumo instantaneo (default)"
for i, (g_, p_) in enumerate(GIROS, 2):
    wsX.cell(row=i, column=1, value=g_); wsX.cell(row=i, column=2, value=p_).number_format = NF_P
wsX["D1"] = "Horas punta/dia habil"
for m in range(1, 13): wsX.cell(row=1, column=4 + m, value=m).font = F_H
for i, dv in enumerate(["SIN", "BC", "BCS"], 2):
    wsX.cell(row=i, column=4, value=dv).font = F_H
    for m in range(1, 13): wsX.cell(row=i, column=4 + m, value=PUNTA_H[dv][m])
wsX["R1"] = "Festivos (engine v2: 7 estatutarios/ano)"
for i, f in enumerate(FEST, 2): wsX.cell(row=i, column=18, value=f).number_format = "yyyy-mm-dd"
wsX["T1"] = "Listas"
for i, v in enumerate(["SIN", "BC", "BCS"], 2): wsX.cell(row=i, column=20, value=v)
for i, v in enumerate(["Si", "No"], 2): wsX.cell(row=i, column=21, value=v)
for i, v in enumerate(["PV", "BESS", "Hibrido"], 2): wsX.cell(row=i, column=22, value=v)
for r in wsX.iter_rows():
    for c in r:
        if c.font is None or not c.font.bold: c.font = F_FX
wsX["A1"].font = wsX["D1"].font = wsX["R1"].font = wsX["T1"].font = F_H
wsX.sheet_state = "hidden"

# ---- INPUTS ----
wsI = wb.create_sheet("INPUTS")
REF = {}
rows_def = [
    ("CLIENTE", None, None, None),
    ("Nombre del cliente", "cli_", "Grupo Posadas — Hotel Cancun (DEMO)", "texto libre"),
    ("RPU / No. de servicio", "rpu_", "780881200029", ""),
    ("Municipio / Estado", "mun_", "Benito Juarez, Q. Roo", "para rendimiento solar"),
    ("Division tarifaria", "div_", "SIN", "SIN / BC / BCS — define horarios punta"),
    ("Giro del cliente", "giro_", "Hotel/Turismo", "define % autoconsumo default"),
    ("Demanda contratada (kW)", "dcon_", P["dcon"], "del recibo CFE — valida SAE-CC"),
    ("Incluir correccion FP", "fptog_", "No", "palanca separada; solo si hay cargo FP"),
    ("Factor de carga GDMTH (FC)", "fc_", P["fc"], "A/158/2024 — NO editar salvo cambio regulatorio"),
    ("SISTEMA PV", None, None, None),
    ("Capacidad PV (kWp)", "kwp_", P["kwp"], ">=700 kWp activa regimen Autoconsumo (ver REGIMEN)"),
    ("Degradacion PV (%/ano)", "pvdeg_", P["pv_degr"], ""),
    ("EPC PV (USD/Wp)", "epcpv_", P["epc_usd_wp"], "default engine"),
    ("Tipo de cambio (MXN/USD)", "fx_", P["fx"], ""),
    ("CAPEX PV (MXN)", "capexpv_", "F:={kwp_}*1000*{epcpv_}*{fx_}", "kWp x 1000 x EPC x FX"),
    ("O&M PV (MXN/ano, ano 1)", "ompv_", P["om_pv"], "escala con tarifa CFE"),
    ("Tarifa PPA (MXN/kWh)", "ppa_", P["ppa"], "deal term"),
    ("Plazo PPA (anos)", "ppayrs_", P["ppa_yrs"], ""),
    ("Escalacion PPA (%/ano)", "escppa_", P["esc_ppa"], "desde ano 2"),
    ("BESS", None, None, None),
    ("Capacidad nominal (kWh)", "bkwh_", P["bess_kwh"], "0 = sin BESS"),
    ("Potencia (kW)", "bkw_", P["bess_kw"], "debe ser <= demanda contratada (SAE-CC)"),
    ("DoD", "dod_", P["dod"], ""),
    ("RTE", "rte_", P["rte"], "round-trip efficiency"),
    ("kWh entregables por ciclo", "deliv_", "F:={bkwh_}*{dod_}*SQRT({rte_})", "kWh x DoD x raiz(RTE)"),
    ("Degradacion BESS (%/ano)", "bdeg_", P["bess_degr"], ""),
    ("EPC BESS (USD/kWh)", "epcb_", P["epc_bess"], ""),
    ("CAPEX BESS (MXN)", "capexb_", "F:={bkwh_}*{epcb_}*{fx_}", ""),
    ("O&M BESS (MXN/ano, ano 1)", "omb_", P["om_bess"], "falla NO aplica al O&M (C5)"),
    ("Comision financiador (% ahorro BESS)", "comis_", P["comision"], "deal term"),
    ("Meses falla BESS/ano", "fallam_", P["falla_meses"], "disponibilidad, no fisica"),
    ("Factor falla", "falla_", "F:=(12-{fallam_})/12", "aplica solo al ahorro BESS"),
    ("Plazo servicio BESS (anos)", "byrs_", P["bess_yrs"], ""),
    ("REGIMEN Y CAPS", None, None, None),
    ("% autoconsumo instantaneo (default por giro)", "acdef_", "F:=VLOOKUP({giro_},AUX!$A$2:$B$7,2,0)", "supuesto — sin datos 15-min"),
    ("Override % autoconsumo (vacio = default)", "acovr_", "", "pegar dato medido si existe"),
    ("% autoconsumo efectivo", "pctac_", 'F:=IF({acovr_}="",{acdef_},{acovr_})', "usado si kWp >= 700"),
    ("Tarifa excedentes (MXN/kWh)", "texc_", P["texc"], "piso conservador $0; >=0.7 MW venta SOLO a CFE (MEM)"),
    ("FINANZAS", None, None, None),
    ("WACC", "wacc_", P["wacc"], ""),
    ("Escalacion tarifa CFE (%/ano)", "esc_", P["esc_cfe"], ""),
    ("Fianza (% del PPA)", "fianza_", P["fianza"], ""),
    ("Horizonte (anos)", "hor_", P["horizon"], "FINANCE usa 20"),
    ("Escenario a financiar", "escn_", "Hibrido", "PV / BESS / Hibrido"),
]
r = 2
for label, key, val, note in rows_def:
    if key: REF[key] = f"$B${r}"
    r += 1
r = 2
for label, key, val, note in rows_def:
    a = wsI.cell(row=r, column=1, value=label)
    if key is None:
        a.font = F_H; a.fill = FILL_H
        wsI.cell(row=r, column=2).fill = FILL_H; wsI.cell(row=r, column=3).fill = FILL_H
    else:
        a.font = F_FX
        is_f = isinstance(val, str) and val.startswith("F:")
        b = wsI.cell(row=r, column=2, value=val[2:].format(**REF) if is_f else val)
        b.font = F_FX if is_f else F_IN
        wsI.cell(row=r, column=3, value=note).font = Font(italic=True, size=8, name="Arial", color="808080")
    r += 1
wsI.cell(row=1, column=1, value="INPUTS — unico lugar con supuestos editables (azul = editable)").font = F_T
for key, fmt in [("dcon_", NF_K), ("kwp_", NF_K1), ("pvdeg_", NF_P), ("capexpv_", NF_M),
                 ("ompv_", NF_M), ("escppa_", NF_P), ("bkwh_", NF_K), ("bkw_", NF_K),
                 ("deliv_", NF_K1), ("bdeg_", NF_P), ("capexb_", NF_M), ("omb_", NF_M),
                 ("comis_", NF_P), ("falla_", "0.00"), ("acdef_", NF_P), ("pctac_", NF_P),
                 ("wacc_", NF_P), ("esc_", NF_P), ("fianza_", NF_P), ("fc_", "0.00")]:
    wsI[REF[key].replace("$", "")].number_format = fmt
# yield block
wsI.cell(row=1, column=5, value="Rendimiento PV").font = F_H
hdr(wsI, 2, ["Mes", "kWh/kWp"], col0=5)
for m in range(1, 13):
    wsI.cell(row=2 + m, column=5, value=m).font = F_FX
    y_ = wsI.cell(row=2 + m, column=6, value=yield_kwp[m])
    y_.font = F_IN; y_.number_format = "0.0"
t = wsI.cell(row=15, column=5, value="Anual"); t.font = F_H
a_ = wsI.cell(row=15, column=6, value="=SUM(F3:F14)"); a_.font = F_FX; a_.number_format = NF_K
wsI.cell(row=16, column=5, value="Fuente: Helioscope mensual (preferido) o indice municipal [[solar-yield-lookup]] — grado prefactibilidad").font = Font(italic=True, size=8, name="Arial", color="808080")
wsI.column_dimensions["A"].width = 38; wsI.column_dimensions["B"].width = 16
wsI.column_dimensions["C"].width = 46; wsI.column_dimensions["E"].width = 6
for nm, key in [("KWP", "kwp_"), ("DELIV", "deliv_"), ("BKW", "bkw_"), ("BKWH", "bkwh_"),
                ("RTE_", "rte_"), ("FALLA", "falla_"), ("COMIS", "comis_"), ("WACC_", "wacc_"),
                ("ESC", "esc_"), ("FC_", "fc_"), ("PCTAC", "pctac_"), ("TEXC", "texc_"),
                ("DIVN", "div_"), ("DCON", "dcon_"), ("FPTOG", "fptog_"), ("ESCN", "escn_")]:
    wb.defined_names.add(DefinedName(nm, attr_text=f"INPUTS!{REF[key]}"))
for nm, rng in [("LDIV", "AUX!$T$2:$T$4"), ("LSN", "AUX!$U$2:$U$3"), ("LESC", "AUX!$V$2:$V$4"),
                ("LGIRO", "AUX!$A$2:$A$7")]:
    wb.defined_names.add(DefinedName(nm, attr_text=rng))
for nm, key in [("LDIV", "div_"), ("LGIRO", "giro_"), ("LSN", "fptog_"), ("LESC", "escn_")]:
    dv = DataValidation(type="list", formula1=nm, allow_blank=False)
    wsI.add_data_validation(dv); dv.add(wsI[REF[key].replace("$", "")])
I = {k: f"INPUTS!{v}" for k, v in REF.items()}

# ---- BILLS ----
wsB = wb.create_sheet("BILLS")
hdrB = ["Periodo (dia 1 del mes)", "Dias del periodo", "Dias habiles punta", "kWh Base", "kWh Intermedio",
        "kWh Punta", "kWh Total", "kW Base", "kW Intermedio", "kW Punta", "Cargo Fijo $",
        "Distribucion $", "Transmision $", "CENACE $", "Energia Base $", "Energia Intermedio $",
        "Energia Punta $", "Capacidad $", "SCnMEM $", "Subtotal MEM $", "Bonif(-)/Cargo(+) FP $",
        "FP share", "Total calc c/IVA", "Recibo CFE $", "Diff %", "Check"]
hdr(wsB, 1, hdrB)
for i, mo in enumerate(months):
    r = 2 + i
    wsB.cell(row=r, column=1, value=mo["per"]).number_format = "mmm-yy"
    wsB.cell(row=r, column=2, value=mo["days"])
    wsB.cell(row=r, column=3, value=f"=NETWORKDAYS(A{r},EOMONTH(A{r},0),AUX!$R$2:$R$22)").font = F_FX
    for c, v in zip([4, 5, 6], [mo["kwh_b"], mo["kwh_i"], mo["kwh_p"]]):
        wsB.cell(row=r, column=c, value=v)
    wsB.cell(row=r, column=7, value=f"=D{r}+E{r}+F{r}").font = F_FX
    for c, v in zip(range(8, 20), [mo["kw_b"], mo["kw_i"], mo["kw_p"], mo["cargo_fijo"],
                                   mo["dist"], mo["trans"], mo["cenace"], mo["gen_b"],
                                   mo["gen_i"], mo["gen_p"], mo["capacidad"], mo["scnmem"]]):
        wsB.cell(row=r, column=c, value=v)
    wsB.cell(row=r, column=20, value=f"=SUM(K{r}:S{r})").font = F_FX
    wsB.cell(row=r, column=21, value=mo["bonif"])
    wsB.cell(row=r, column=22, value=f"=IF(T{r}=0,0,-U{r}/T{r})").font = F_FX
    wsB.cell(row=r, column=23, value=f"=(T{r}+U{r})*1.16").font = F_FX
    wsB.cell(row=r, column=24, value=mo["recibo"])
    wsB.cell(row=r, column=25, value=f"=IF(X{r}=0,0,W{r}/X{r}-1)").font = F_FX
    wsB.cell(row=r, column=26, value=f'=IF(X{r}=0,"sin recibo",IF(ABS(Y{r})<0.005,"OK","REVISAR"))').font = F_FX
    for c in [1, 2, 4, 5, 6, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 21, 24]:
        wsB.cell(row=r, column=c).font = F_IN
wsB.cell(row=14, column=1, value="TOTAL").font = F_H
for c in [4, 5, 6, 7, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 23, 24]:
    L = get_column_letter(c)
    cc = wsB.cell(row=14, column=c, value=f"=SUM({L}2:{L}13)"); cc.font = F_H
for r in range(2, 15):
    for c in range(11, 25):
        if c != 22: wsB.cell(row=r, column=c).number_format = NF_M
    wsB.cell(row=r, column=22).number_format = NF_P
    wsB.cell(row=r, column=25).number_format = NF_P
    for c in [4, 5, 6, 7]: wsB.cell(row=r, column=c).number_format = NF_K
wsB.cell(row=16, column=1, value="Validacion: (Subtotal MEM + Bonif/Cargo FP) x 1.16 debe cuadrar con el recibo impreso (<0.5%). Si dice REVISAR: error de transcripcion o error de CFE — detener y revisar.").font = Font(italic=True, size=8, name="Arial")
wsB.freeze_panes = "B2"

# ---- RATES ----
wsR = wb.create_sheet("RATES")
hdrR = ["Periodo", "Umbral kW", "Base fact. capacidad", "Base fact. distribucion", "$/kW Capacidad",
        "$/kW Distribucion", "$/kWh Trans", "$/kWh CENACE", "$/kWh SCnMEM", "$/kWh Energia B",
        "$/kWh Energia I", "$/kWh Energia P", "Bundled Base", "Bundled Intermedio", "Bundled Punta"]
hdr(wsR, 1, hdrR)
for r in range(2, 14):
    fx = [
        (1, f"=BILLS!A{r}"), (2, f"=IF(BILLS!B{r}=0,0,BILLS!G{r}/(BILLS!B{r}*24*FC_))"),
        (3, f"=IF(BILLS!J{r}=0,B{r},MIN(BILLS!J{r},B{r}))"),
        (4, f"=MIN(MAX(BILLS!H{r}:J{r}),B{r})"),
        (5, f"=IF(C{r}=0,0,BILLS!R{r}/C{r})"), (6, f"=IF(D{r}=0,0,BILLS!L{r}/D{r})"),
        (7, f"=IF(BILLS!G{r}=0,0,BILLS!M{r}/BILLS!G{r})"),
        (8, f"=IF(BILLS!G{r}=0,0,BILLS!N{r}/BILLS!G{r})"),
        (9, f"=IF(BILLS!G{r}=0,0,BILLS!S{r}/BILLS!G{r})"),
        (10, f"=IF(BILLS!D{r}=0,0,BILLS!O{r}/BILLS!D{r})"),
        (11, f"=IF(BILLS!E{r}=0,0,BILLS!P{r}/BILLS!E{r})"),
        (12, f"=IF(BILLS!F{r}=0,0,BILLS!Q{r}/BILLS!F{r})"),
        (13, f"=IF(BILLS!D{r}=0,N{r},J{r}+G{r}+H{r}+I{r})"),
        (14, f"=K{r}+G{r}+H{r}+I{r}"), (15, f"=IF(BILLS!F{r}=0,0,L{r}+G{r}+H{r}+I{r})")]
    for c, f in fx:
        cell = wsR.cell(row=r, column=c, value=f); cell.font = F_LK if "BILLS!" in f else F_FX
    wsR.cell(row=r, column=1).number_format = "mmm-yy"
    for c in [2, 3, 4]: wsR.cell(row=r, column=c).number_format = NF_K1
    for c in range(5, 7): wsR.cell(row=r, column=c).number_format = NF_M2
    for c in range(7, 16): wsR.cell(row=r, column=c).number_format = NF_R
wsR.cell(row=16, column=1, value="Tarifas implicitas del propio recibo (ground truth de su mes). Umbral = kWh/(dias x 24 x FC) [[demanda-facturable]]; capacidad sobre min(kW punta, umbral); distribucion sobre min(max(B,I,P), umbral). Si no hay punta (BC/BCS invierno): capacidad sobre umbral.").font = Font(italic=True, size=8, name="Arial")

# ---- REGIMEN (before CALC so AUTOC exists) ----
wsG = wb.create_sheet("REGIMEN")
wsG.cell(row=1, column=1, value="REGIMEN REGULATORIO — deteccion y caps automaticos").font = F_T
wsG.cell(row=2, column=1, value="Autoconsumo aplica (kWp >= 700)").font = F_FX
wsG.cell(row=2, column=2, value="=KWP>=700").font = F_FX
wb.defined_names.add(DefinedName("AUTOC", attr_text="REGIMEN!$B$2"))
reg_rows = [
    ("Regimen detectado", '=IF(KWP<700,"GENERADOR EXENTO — Medicion Neta (sin permiso; Art. 30 LSE; RES/142/2017)",IF(KWP<=20000,"AUTOCONSUMO interconectado 0.7–20 MW — permiso SIMPLIFICADO CNE (Num. 2.2 DACG)","AUTOCONSUMO >20 MW — permiso ORDINARIO CNE (Num. 2.3 DACG)"))'),
    ("Compensacion de energia", '=IF(KWP<700,"Neteo por periodo + creditos 12 meses; excedente expirado liquida a PML (no a tarifa)","Solo autoconsumo instantaneo acreditable; excedentes SOLO a CFE (MEM) o carta de renuncia — sin venta a terceros (Art. 32-II LSE)")'),
    ("% PV acreditable en el modelo", '=IF(KWP<700,"100% del kWh generado (tope: consumo del mes)",TEXT(PCTAC,"0%")&" autoconsumo instantaneo (supuesto por giro — editar en INPUTS)")'),
    ("Excedentes valuados a", '=TEXT(TEXC,"$0.00")&" /kWh (default $0 = piso conservador)"'),
]
for i, (a, f) in enumerate(reg_rows, 3):
    wsG.cell(row=i, column=1, value=a).font = F_H
    wsG.cell(row=i, column=2, value=f).font = F_FX
wsG.cell(row=8, column=1, value="OBLIGACIONES (checklist)").font = F_H; wsG.cell(row=8, column=1).fill = FILL_H
obls = [
    ("Permiso de generacion CNE", '=IF(KWP<700,"No requerido (exento)",IF(KWP<=20000,"REQUERIDO — simplificado, ~6-12 meses (OPE)","REQUERIDO — ordinario, ~12-24 meses"))'),
    ("Contrato de interconexion CFE", '="Requerido siempre que hay PV interconectado (clases BT/MT1/MT2; regla 80% del transformador)"'),
    ("Medidor bidireccional", '=IF(KWP<700,"Requerido (spec CFE G0000-48)","Requerido + representacion MEM")'),
    ("Respaldo SAE / CFE (intermitencia)", '=IF(KWP<700,"No aplica",IF(BKWH>0,"OBLIGATORIO — cubierto por el BESS del proyecto (Num. 4.5 DACG)","OBLIGATORIO Y NO CUBIERTO — agregar SAE o contratar respaldo CFE"))'),
    ("Registro anual de Autoconsumo (1T)", '=IF(KWP<700,"No aplica","REQUERIDO cada ano ante CNE (Art. 52 RLSE)")'),
    ("Uso minimo de la central (12 meses)", '=IF(KWP<700,"No aplica",">= 30% (renovable) — si no, el permiso termina (Num. 1.6-I DACG)")'),
    ("BESS — SAE-CC (sin inyeccion)", '=IF(BKWH=0,"Sin BESS","Sin permiso; notificacion 90 dias habiles; PROHIBIDO inyectar a la red (A/113/2024)")'),
]
for i, (a, f) in enumerate(obls, 9):
    wsG.cell(row=i, column=1, value=a).font = F_FX
    wsG.cell(row=i, column=2, value=f).font = F_FX
wsG.cell(row=17, column=1, value="ALERTAS").font = F_H; wsG.cell(row=17, column=1).fill = FILL_W
alerts = [
    ("Validacion de recibos", '=IF(COUNTIF(BILLS!Z2:Z13,"REVISAR")>0,"ALERTA: hay recibos que no cuadran (>0.5%) — corregir transcripcion antes de usar resultados","OK — los recibos capturados cuadran")'),
    ("Gate 100 kW (GDMTH/GDMTO)", '=IF(MAX(BILLS!H2:J13)<110,"ALERTA: demanda max < 110 kW — cerca del gate de 100 kW; revisar si la tarifa correcta es GDMTO","OK — claramente GDMTH")'),
    ("SAE-CC: potencia BESS vs contratada", '=IF(BKW>DCON,"ALERTA: BESS "&TEXT(BKW,"#,##0")&" kW > demanda contratada "&TEXT(DCON,"#,##0")&" kW — viola condicion SAE-CC","OK")'),
    ("Sobredimensionamiento PV", '=IF(SUM(CALC!G2:G13)>0,"ALERTA: "&TEXT(SUM(CALC!G2:G13),"#,##0")&" kWh/ano de excedentes valuados a "&TEXT(TEXC,"$0.00")&" — bajo medicion neta los creditos expiran a PML en 12 meses; ajustar kWp al consumo","OK — sin excedentes estructurales")'),
    ("Factor de potencia", '=IF(SUMIF(BILLS!U2:U13,">0")>0,"OPORTUNIDAD: cargos FP por "&TEXT(SUMIF(BILLS!U2:U13,">0"),"$#,##0")&"/ano — la correccion FP suele ser el ROI mas alto y es ortogonal a PV/BESS (activar en INPUTS)",IF(SUM(BILLS!U2:U13)<0,"Sitio con bonificacion FP — el claw-back ya esta modelado","Sin dato FP"))'),
    ("Division no-SIN", '=IF(DIVN="SIN","OK — modelo calibrado SIN","PRECAUCION: en "&DIVN&" el PV genera parcialmente en horario punta (no acreditado aqui = conservador) y las tarifas no-SIN son gap abierto del wiki; punta invierno = 0 h")'),
    ("PV >= 700 kWp", '=IF(AUTOC,"Regimen Autoconsumo activo: solo "&TEXT(PCTAC,"0%")&" del PV se acredita (autoconsumo instantaneo), excedentes a "&TEXT(TEXC,"$0.00")&"/kWh, permiso CNE y obligaciones arriba","OK — dentro del techo exento (<0.7 MW)")'),
]
for i, (a, f) in enumerate(alerts, 18):
    wsG.cell(row=i, column=1, value=a).font = F_FX
    wsG.cell(row=i, column=2, value=f).font = F_FX
wsG.cell(row=26, column=1, value="Fuentes wiki: [[scheme-comparison]] [[autoconsumo]] [[medicion-neta]] [[generador-exento]] [[sae-cc]] [[interconexion-cre]] [[demanda-facturable]]").font = Font(italic=True, size=8, name="Arial", color="808080")
wsG.column_dimensions["A"].width = 36; wsG.column_dimensions["B"].width = 110

# ---- CALC ----
wsC = wb.create_sheet("CALC")
hdrC = ["Periodo", "Mes", "h punta/dia", "Gen PV kWh", "Gen usable kWh (cap regimen)",
        "Gen aprovechada kWh (cap consumo)", "Excedentes kWh", "Ingreso excedentes $",
        "Ahorro energia PV $", "Umbral post-PV kW", "Fact. pre kW", "Fact. pre-BESS kW",
        "Ahorro dem. PV $", "Claw PV $", "AHORRO PV-ONLY $",
        "Shave kW", "Pico residual kW", "Fact. post hibrido kW", "Ahorro dem. BESS $ (hib)",
        "Descarga kWh", "Carga kWh", "Arbitraje $", "Subtotal hibrido $", "Claw hibrido $",
        "AHORRO HIBRIDO $", "Fact. post BESS-only kW", "Ahorro dem. BESS-only $",
        "Subtotal BESS-only $", "Claw BESS-only $", "AHORRO BESS-ONLY $",
        "Factura ANTES $ (real)", "Factura DESPUES $ (hib)", "Ahorro % (hib)"]
hdr(wsC, 1, hdrC)
# hidden helper: H col index for excedentes warning uses CALC!H
for r in range(2, 14):
    fx = [
        (1, f"=BILLS!A{r}"), (2, f"=MONTH(A{r})"),
        (3, f"=INDEX(AUX!$E$2:$P$4,MATCH(DIVN,AUX!$D$2:$D$4,0),B{r})"),
        (4, f"=KWP*INDEX(INPUTS!$F$3:$F$14,B{r})"),
        (5, f"=D{r}*IF(AUTOC,PCTAC,1)"),
        (6, f"=MIN(E{r},BILLS!E{r})"),
        (7, f"=D{r}-F{r}"), (8, f"=G{r}*TEXC"),
        (9, f"=F{r}*RATES!N{r}+H{r}"),
        (10, f"=IF(BILLS!B{r}=0,0,(BILLS!G{r}-F{r})/(BILLS!B{r}*24*FC_))"),
        (11, f"=RATES!C{r}"),
        (12, f"=IF(BILLS!J{r}=0,J{r},MIN(BILLS!J{r},J{r}))"),
        (13, f"=(K{r}-L{r})*RATES!E{r}"),
        (14, f"=IF(BILLS!U{r}<0,-(I{r}+M{r})*BILLS!V{r},0)"),
        (15, f"=I{r}+M{r}+N{r}"),
        (16, f"=IF(C{r}=0,0,MIN(BKW,DELIV/C{r}))"),
        (17, f"=MAX(0,BILLS!J{r}-P{r})"),
        (18, f"=MIN(Q{r},J{r})"),
        (19, f"=(L{r}-R{r})*RATES!E{r}"),
        (20, f"=IF(C{r}=0,0,MIN(DELIV*BILLS!C{r},BILLS!F{r}))"),
        (21, f"=T{r}/RTE_"),
        (22, f"=T{r}*RATES!O{r}-U{r}*RATES!M{r}"),
        (23, f"=I{r}+M{r}+S{r}+V{r}"),
        (24, f"=IF(BILLS!U{r}<0,-W{r}*BILLS!V{r},0)"),
        (25, f"=W{r}+X{r}"),
        (26, f"=MIN(Q{r},RATES!B{r})"),
        (27, f"=(K{r}-Z{r})*RATES!E{r}"),
        (28, f"=AA{r}+V{r}"),
        (29, f"=IF(BILLS!U{r}<0,-AB{r}*BILLS!V{r},0)"),
        (30, f"=AB{r}+AC{r}"),
        (31, f"=BILLS!T{r}+BILLS!U{r}"),
        (32, f"=AE{r}-Y{r}"),
        (33, f"=IF(AE{r}=0,0,Y{r}/AE{r})")]
    for c, f in fx:
        cell = wsC.cell(row=r, column=c, value=f)
        cell.font = F_LK if ("BILLS!" in f or "RATES!" in f or "INPUTS!" in f or "AUX!" in f) else F_FX
    wsC.cell(row=r, column=1).number_format = "mmm-yy"
wsC.cell(row=14, column=1, value="TOTAL ANO 0").font = F_H
for c in list(range(4, 10)) + list(range(13, 16)) + [19, 20, 21, 22, 23, 24, 25, 27, 28, 29, 30, 31, 32]:
    L = get_column_letter(c)
    cc = wsC.cell(row=14, column=c, value=f"=SUM({L}2:{L}13)"); cc.font = F_H
wsC.cell(row=14, column=33, value="=IF(AE14=0,0,Y14/AE14)").font = F_H
for r in range(2, 15):
    for c in [4, 5, 6, 7, 20, 21]: wsC.cell(row=r, column=c).number_format = NF_K
    for c in [3, 10, 11, 12, 16, 17, 18, 26]: wsC.cell(row=r, column=c).number_format = NF_K1
    for c in [8, 9, 13, 14, 15, 19, 22, 23, 24, 25, 27, 28, 29, 30, 31, 32]:
        wsC.cell(row=r, column=c).number_format = NF_M
    wsC.cell(row=r, column=33).number_format = NF_P
wsC.cell(row=16, column=1, value="Motor mensual — no editar (todo input vive en INPUTS/BILLS). Demanda: capacidad sobre min(kW punta, umbral) y el shave BESS solo via pico punta; PV no toca punta en SIN [[pv-savings-model]]. Descarga limitada a dias habiles punta y al kWh punta del recibo [[bess-savings-model]]. Umbral post-PV acopla PV+BESS [[pv-bess-combined]].").font = Font(italic=True, size=8, name="Arial")
wsC.freeze_panes = "B2"

# ---- RESULTADOS ----
wsD = wb.create_sheet("RESULTADOS")
wsD.cell(row=1, column=1, value="RESULTADOS — resumen anual y tabla titular").font = F_T
wsD.cell(row=2, column=1, value='=REGIMEN!B3').font = F_LK
hdr(wsD, 4, ["Escenario", "Ahorro bruto $/ano", "% de factura real", "Ahorro neto disponibilidad $/ano",
             "CAPEX MXN", "Payback simple (anos)"])
scen = [
    ("PV solo", "=CALC!O14", "=CALC!O14/CALC!AE14", "=CALC!O14", f"={I['capexpv_']}",
     f"=IF(CALC!O14=0,0,{I['capexpv_']}/D5)"),
    ("BESS solo", "=CALC!AD14", "=CALC!AD14/CALC!AE14", "=CALC!AD14-(CALC!AA14+CALC!V14)*(1-FALLA)",
     f"={I['capexb_']}", f"=IF(D6=0,0,{I['capexb_']}/D6)"),
    ("Hibrido (PV+BESS)", "=CALC!Y14", "=CALC!Y14/CALC!AE14",
     "=CALC!Y14-(CALC!S14+CALC!V14)*(1-FALLA)",
     f"={I['capexpv_']}+{I['capexb_']}", f"=IF(D7=0,0,({I['capexpv_']}+{I['capexb_']})/D7)"),
]
for i, row in enumerate(scen, 5):
    for c, f in enumerate(row, 1):
        cell = wsD.cell(row=i, column=c, value=f)
        cell.font = F_H if c == 1 else (F_LK if "CALC!" in str(f) else F_FX)
wsD.cell(row=8, column=1, value="+ Correccion FP (palanca separada)").font = F_H
wsD.cell(row=8, column=2, value='=IF(FPTOG="Si",SUMIF(BILLS!U2:U13,">0"),0)').font = F_LK
wsD.cell(row=8, column=3, value="=B8/CALC!AE14").font = F_FX
wsD.cell(row=8, column=4, value="=B8").font = F_FX
wsD.cell(row=9, column=1, value="TOTAL escenario elegido + FP").font = F_H
wsD.cell(row=9, column=2, value="=CHOOSE(MATCH(ESCN,LESC,0),B5,B6,B7)+B8").font = F_FX
wsD.cell(row=9, column=3, value="=B9/CALC!AE14").font = F_FX
wsD.cell(row=9, column=4, value="=CHOOSE(MATCH(ESCN,LESC,0),D5,D6,D7)+B8").font = F_FX
for r in range(5, 10):
    for c, fmt in [(2, NF_M), (3, NF_P), (4, NF_M), (5, NF_M), (6, "0.0")]:
        wsD.cell(row=r, column=c).number_format = fmt
wsD.cell(row=11, column=1, value='="TABLA TITULAR — escenario: "&ESCN&" (cambiar en INPUTS)"').font = F_H
hdr(wsD, 12, ["Mes", "kWh Consumo", "Generacion kWh", "kWh BESS desc.", "Factura ANTES $",
              "Factura DESPUES $", "Ahorro $", "Ahorro %"])
for i in range(12):
    r, cr = 13 + i, 2 + i
    fx = [(1, f"=CALC!A{cr}"), (2, f"=BILLS!G{cr}"),
          (3, f'=IF(ESCN="BESS",0,CALC!D{cr})'),
          (4, f'=IF(ESCN="PV",0,CALC!T{cr})'),
          (5, f"=CALC!AE{cr}"),
          (6, f"=E{r}-G{r}"),
          (7, f"=CHOOSE(MATCH(ESCN,LESC,0),CALC!O{cr},CALC!AD{cr},CALC!Y{cr})"),
          (8, f"=IF(E{r}=0,0,G{r}/E{r})")]
    for c, f in fx:
        cell = wsD.cell(row=r, column=c, value=f); cell.font = F_LK if "CALC!" in f or "BILLS!" in f else F_FX
    wsD.cell(row=r, column=1).number_format = "mmm-yy"
wsD.cell(row=25, column=1, value="TOTAL").font = F_H
for c, fmt in [(2, NF_K), (3, NF_K), (4, NF_K), (5, NF_M), (6, NF_M), (7, NF_M)]:
    L = get_column_letter(c)
    cc = wsD.cell(row=25, column=c, value=f"=SUM({L}13:{L}24)"); cc.font = F_H
wsD.cell(row=25, column=8, value="=IF(E25=0,0,G25/E25)").font = F_H
for r in range(13, 26):
    for c, fmt in [(2, NF_K), (3, NF_K), (4, NF_K), (5, NF_M), (6, NF_M), (7, NF_M), (8, NF_P)]:
        wsD.cell(row=r, column=c).number_format = fmt
wsD.cell(row=27, column=1, value="Niveles de hecho: tarifas y factura = del recibo (fuente primaria); ahorros = derivados del motor; % autoconsumo, rendimiento solar, curva intradia = SUPUESTOS (editar en INPUTS). Numeros titulares en BRUTO; neto de disponibilidad al lado [doctrina CLAUDE.md].").font = Font(italic=True, size=8, name="Arial")
for cdim, w in [("A", 22), ("B", 14), ("C", 14), ("D", 14), ("E", 16), ("F", 16), ("G", 14), ("H", 10)]:
    wsD.column_dimensions[cdim].width = w

# ---- FINANCE ----
wsF = wb.create_sheet("FINANCE")
wsF.cell(row=1, column=1, value='="FINANCE — proyeccion 20 anos, escenario: "&ESCN').font = F_T
base_defs = [
    ("Indice escenario", "=MATCH(ESCN,LESC,0)"),
    ("Base ahorro PV $/ano", "=CHOOSE(B3,CALC!I14+CALC!M14,0,CALC!I14+CALC!M14)"),
    ("Base ahorro BESS $/ano", "=CHOOSE(B3,0,CALC!AA14+CALC!V14,CALC!S14+CALC!V14)"),
    ("Base claw-back $/ano", "=CHOOSE(B3,CALC!N14*0,CALC!AC14,CALC!X14)"),
    ("Generacion anual kWh", "=IF(B3=2,0,CALC!D14)"),
    ("CAPEX escenario", f"=CHOOSE(B3,{I['capexpv_']},{I['capexb_']},{I['capexpv_']}+{I['capexb_']})"),
]
for i, (a, f) in enumerate(base_defs, 3):
    wsF.cell(row=i, column=1, value=a).font = F_FX
    wsF.cell(row=i, column=2, value=f).font = F_LK
    if i >= 4: wsF.cell(row=i, column=2).number_format = NF_M
# fix claw PV: CHOOSE arg1 should be CALC!N14*0? no: PV-only claw is CALC!N14? PV claw col is N (14)?  -> col 14 = "Claw PV $" => letter N. Correct mapping below.
wsF["B6"] = "=CHOOSE(B3,CALC!N14,CALC!AC14,CALC!X14)"
hdrF = ["Ano", "Esc CFE", "Degr PV", "Degr BESS", "Ahorro PV $", "Ahorro BESS $ (x falla)",
        "Claw-back $", "Ahorro Bruto $", "Pago PPA $", "Comision BESS $", "O&M PV $",
        "O&M BESS $", "Fianza $", "Flujo CLIENTE $", "Flujo FINANCIADOR $", "Flujo PROYECTO $",
        "Acum. financiador", "Acum. proyecto"]
hdr(wsF, 9, hdrF)
wsF.cell(row=10, column=1, value=0)
wsF.cell(row=10, column=15, value="=-$B$8").font = F_FX
wsF.cell(row=10, column=16, value="=-$B$8").font = F_FX
wsF.cell(row=10, column=17, value="=O10").font = F_FX
wsF.cell(row=10, column=18, value="=P10").font = F_FX
for yy in range(1, 21):
    r = 10 + yy
    fx = [(1, yy), (2, f"=(1+ESC)^A{r}"), (3, f"=(1-{I['pvdeg_']})^A{r}"),
          (4, f"=(1-{I['bdeg_']})^A{r}"),
          (5, f"=$B$4*B{r}*C{r}"), (6, f"=$B$5*B{r}*D{r}*FALLA"),
          (7, f"=$B$6*B{r}*(C{r}+D{r})/2"), (8, f"=E{r}+F{r}+G{r}"),
          (9, f"=IF(AND($B$4>0,A{r}<={I['ppayrs_']}),$B$7*C{r}*{I['ppa_']}*(1+{I['escppa_']})^(A{r}-1),0)"),
          (10, f"=IF(AND($B$5>0,A{r}<={I['byrs_']}),F{r}*COMIS,0)"),
          (11, f"=IF(AND($B$4>0,A{r}<={I['ppayrs_']}),{I['ompv_']}*(1+ESC)^(A{r}-1),0)"),
          (12, f"=IF(AND($B$5>0,A{r}<={I['byrs_']}),{I['omb_']}*(1+ESC)^(A{r}-1),0)"),
          (13, f"=I{r}*{I['fianza_']}"),
          (14, f"=H{r}-I{r}-J{r}"), (15, f"=I{r}+J{r}-K{r}-L{r}-M{r}"),
          (16, f"=H{r}-K{r}-L{r}"), (17, f"=Q{r-1}+O{r}"), (18, f"=R{r-1}+P{r}")]
    for c, f in fx:
        wsF.cell(row=r, column=c, value=f).font = F_FX
for r in range(10, 31):
    for c in range(5, 19): wsF.cell(row=r, column=c).number_format = NF_M
    for c in [2, 3, 4]: wsF.cell(row=r, column=c).number_format = "0.000"
inds = [
    ("TIR PROYECTO (unlevered) — 'vale la pena construirlo'", "=IRR(P10:P30)", NF_P),
    ("TIR FINANCIADOR — depende de terminos del deal", "=IRR(O10:O30)", NF_P),
    ("VPN proyecto @ WACC", "=NPV(WACC_,P11:P30)+P10", NF_M),
    ("VPN financiador @ WACC", "=NPV(WACC_,O11:O30)+O10", NF_M),
    ("Payback proyecto (anos)", '=COUNTIF(R10:R30,"<0")', "0"),
    ("Payback financiador (anos)", '=COUNTIF(Q10:Q30,"<0")', "0"),
]
wsF.cell(row=32, column=1, value="INDICADORES (perspectivas separadas — nunca mezclar)").font = F_H
for i, (a, f, fmt) in enumerate(inds, 33):
    wsF.cell(row=i, column=1, value=a).font = F_FX
    cc = wsF.cell(row=i, column=2, value=f); cc.font = F_FX; cc.number_format = fmt
wsF.column_dimensions["A"].width = 46
wsF.cell(row=40, column=1, value="Doctrina: titulares en bruto; TIR proyecto = decision de construir; TIR financiador = deal-specific (PPA, comision, fianza). Pago PPA y comision solo si el escenario incluye esa tecnologia.").font = Font(italic=True, size=8, name="Arial")

# B7 = generacion anual (renombrar refs): move gen anual to B7 and capex to B8
wsF["A3"], wsF["B3"] = "Indice escenario", "=MATCH(ESCN,LESC,0)"
wsF["A4"], wsF["B4"] = "Base ahorro PV $/ano", "=CHOOSE(B3,CALC!I14+CALC!M14,0,CALC!I14+CALC!M14)"
wsF["A5"], wsF["B5"] = "Base ahorro BESS $/ano", "=CHOOSE(B3,0,CALC!AA14+CALC!V14,CALC!S14+CALC!V14)"
wsF["A6"], wsF["B6"] = "Base claw-back $/ano", "=CHOOSE(B3,CALC!N14,CALC!AC14,CALC!X14)"
wsF["A7"], wsF["B7"] = "Generacion anual kWh", "=IF(B3=2,0,CALC!D14)"
wsF["A8"], wsF["B8"] = "CAPEX escenario", f"=CHOOSE(B3,{I['capexpv_']},{I['capexb_']},{I['capexpv_']}+{I['capexb_']})"
for rr in range(4, 9):
    wsF.cell(row=rr, column=2).font = F_LK
    wsF.cell(row=rr, column=2).number_format = NF_M if rr in (4, 5, 6, 8) else NF_K

# ---- README ----
ws0 = wb.create_sheet("README", 0)
readme = [
    ("CALCULADORA DE AHORROS PV / BESS / HIBRIDO — CFE GDMTH (generica)", F_T),
    ("Sucesora corregida y generalizada de la calculadora v1 (raw/data/780881200029.xlsx). Demo precargado: 780881200029 — validado al peso vs el engine (golden: ahorro hibrido $7,083,252 = 23.5% del MEM).", None),
    ("", None),
    ("COMO USAR (3 pasos):", F_H),
    ("  1. INPUTS — captura el perfil del cliente (giro, division, demanda contratada) y el sistema (kWp, BESS). Solo celdas AZULES son editables; los supuestos se auto-proponen por perfil y todo es override-able.", None),
    ("  2. BILLS — transcribe los 12 recibos CFE (celdas azules). La columna Check valida cada recibo: (MEM + bonif FP) x 1.16 vs total impreso. Si dice REVISAR, hay error de captura o error de CFE — no continuar.", None),
    ("  3. RESULTADOS — resumen anual de los 3 escenarios (PV / BESS / Hibrido) + tabla titular mensual. FINANCE da TIR/VPN/payback del escenario elegido en INPUTS. REGIMEN muestra el regimen legal, obligaciones y alertas.", None),
    ("", None),
    ("CAPS REGULATORIOS AUTOMATICOS:", F_H),
    ("  < 700 kWp  — Generador Exento / medicion neta: 100% del PV acreditable, con tope al consumo del mes (el excedente genera creditos que expiran a PML en 12 meses — aqui valuado a tarifa excedentes, default $0).", None),
    ("  >= 700 kWp — Autoconsumo (CNE): solo el % de autoconsumo instantaneo se acredita (supuesto por giro), excedentes SOLO a CFE (default $0), permiso CNE, registro anual, uso minimo 30%, y respaldo SAE obligatorio para PV intermitente.", None),
    ("  BESS — SAE-CC: sin permiso, inyeccion a red prohibida, potencia <= demanda contratada. La descarga solo cicla dias habiles con punta y no puede desplazar mas kWh punta que los del recibo.", None),
    ("", None),
    ("CODIGO DE COLOR: AZUL = input editable | NEGRO = formula | VERDE = link entre hojas.", F_H),
    ("NIVELES DE HECHO: recibo = fuente primaria | ahorro = derivado | % autoconsumo, rendimiento, curva = supuesto.", None),
    ("", None),
    ("MAPA PARA LA WEBAPP (una hoja = un modulo):", F_H),
    ("  INPUTS -> formulario de captura con defaults por perfil | BILLS -> grid de captura con validacion | RATES+CALC -> motor deterministico (logica = wiki CFE Brain) | REGIMEN -> motor de reglas/alertas | RESULTADOS -> dashboard | FINANCE -> modulo financiero.", None),
    ("", None),
    ("LIMITES (decir siempre): curva intradia modelada (no datos 15-min) — el shave es prefactibilidad, no bancable; mecanica de creditos <0.7 MW bajo LSE 2025 sin re-confirmar (gap abierto); tarifas no-SIN = gap; rendimiento por indice municipal = prefactibilidad.", None),
]
for i, (txt, ft) in enumerate(readme, 1):
    cc = ws0.cell(row=i, column=1, value=txt)
    cc.font = ft or Font(name="Arial", size=10)
    cc.alignment = Alignment(wrap_text=False)
ws0.column_dimensions["A"].width = 200

for ws in [wsI, wsB, wsR, wsC, wsG, wsD, wsF]:
    for row in ws.iter_rows():
        for c in row:
            if c.font is None: c.font = F_FX

wb.save(OUT)
print(f"\nWorkbook escrito: {OUT}")
