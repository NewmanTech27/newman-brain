# GEPP monthly savings reconstruction (best-case scenario, 2026-06-10 method)
# Replicates wiki/analyses/2026-06-10-gepp-portfolio-project-check.md §Best-case at MONTHLY resolution:
#   PV  : kWp x monthly yield (Solar_index_geografico.csv) x intermedia rate (Tarifas 2026 avg);
#         Ixtlahuacan zafra months (Dec-May) valued 0.9x (displaces Tala-discounted energy)
#   BESS: punta shave (filed kW), monthly umbral-capped per [[demanda-facturable]] (FC 0.57),
#         valued at capacidad rate; SIN arbitrage penalty -$835/kWh-yr (Tototlan-calibrated)
#   FP  : printed monthly Bonificacion/Cargo FP from the workbook (Acapulco, Cancun)
# Data: raw/Perfil de Consumo Eléctrico GEPP 2025-26 (1).xlsx — 2025 block (cols T..AE), Total Planta rows 6-27.
# Output: per-site monthly markdown tables + Proplasa per-meter ENEL/CFE split + validation vs filed annuals.
import openpyxl, csv, sys, calendar, unicodedata

sys.stdout.reconfigure(encoding="utf-8")
WB = r"C:\Users\vidan\Desktop\CFE Brain\raw\Perfil de Consumo Eléctrico GEPP 2025-26 (1).xlsx"
SOLAR = r"C:\Users\vidan\Desktop\CFE Brain\raw\assets\Solar_index_geografico.csv"
FC = 0.57
ARB_PENALTY_KWH_YR = 835.0   # $/kWh-yr, Tototlan engine calibration
MES = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]

def norm(s):
    return unicodedata.normalize("NFD", str(s)).encode("ascii","ignore").decode().lower()

SITES = [
    dict(sheet="IXT", name="Ixtlahuacán", tarifa_key="ixtlahuacan", pv=2530, bkw=850, bkwh=1700,
         yield_q=("jalisco","ixtlahuac"), zafra={12,1,2,3,4,5}, fp=False),
    dict(sheet="ACA", name="Acapulco", tarifa_key="acapulco", pv=1020, bkw=320, bkwh=640,
         yield_q=("guerrero","acapulco"), zafra=set(), fp=True),
    dict(sheet="CAN", name="Cancún", tarifa_key="cancun", pv=700, bkw=0, bkwh=0,
         yield_q=("quintana roo","benito ju"), zafra=set(), fp=True),
    dict(sheet="PR1", name="Proplasa PR1", tarifa_key="planta preforma", pv=1860, bkw=360, bkwh=720,
         yield_q=("mexico","cuautitlan izcalli"), zafra=set(), fp=False),
    dict(sheet="PR2", name="Proplasa PR2", tarifa_key="planta preforma", pv=5220, bkw=1220, bkwh=2440,
         yield_q=("mexico","cuautitlan izcalli"), zafra=set(), fp=False),
    dict(sheet="TAP", name="Proplasa TAP", tarifa_key="planta tapa", pv=3430, bkw=790, bkwh=1580,
         yield_q=("mexico","cuautitlan izcalli"), zafra=set(), fp=False),
]

# ---------- solar monthly yields ----------
def load_yields():
    out = {}
    with open(SOLAR, encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    for s in SITES:
        est, cid = s["yield_q"]
        best = None
        for r in rows:
            if est in norm(r["estado"]) and cid in norm(r["ciudad"]):
                best = r; break
        if not best:
            raise SystemExit(f"yield row not found for {s['name']} ({est},{cid})")
        keys = ["kWh/kWp_Jan","kWh/kWp_Feb","kWh/kWp_Mar","kWh/kWp_Apr","kWh/kWp_May","kWh/kWp_Jun",
                "kWh/kWp_Jul","kWh/kWp_Ago","kWh/kWp_Sep","kWh/kWp_Oct","kWh/kWp_Nov","kWh/kWp_Dec"]
        s["yield_m"] = [float(best[k]) for k in keys]
        s["yield_city"] = best["ciudad"]
        s["yield_yr"] = float(best["kWh/kWp_anual"])
    return out

# ---------- tarifas (2026 avg of available months, cols F..Q) ----------
def load_tarifas(wb):
    ws = wb["Tarifas"]
    plants = {}   # norm(planta) -> {concepto: rate}
    for row in ws.iter_rows(min_row=4, max_col=17):
        planta, concepto = row[1].value, row[4].value
        if not planta or not concepto:
            continue
        vals = [c.value for c in row[5:17] if isinstance(c.value, (int, float))]
        if not vals:
            continue
        plants.setdefault(norm(planta), {})[norm(concepto)] = sum(vals)/len(vals)
    return plants

def find_rates(plants, key):
    hits = [p for p in plants if key in p]
    if not hits:
        raise SystemExit(f"tarifa plant not found: {key}  (have: {sorted(plants)[:5]}...)")
    p = plants[hits[0]]
    return hits[0], p.get("intermedia"), p.get("capacidad")

# ---------- site sheet 2025 block (cols T..AE = 20..31), Total Planta rows ----------
ROWS = dict(base=6, inter=7, punta=8, dmax_b=9, dmax_i=10, dmax_p=11, fp_cargo=20, recibo=26)
CFE_ROWS = dict(base=35, inter=36, punta=37)

def month_vals(ws, row, cols):
    out = []
    for c in cols:
        v = ws.cell(row=row, column=c).value
        out.append(float(v) if isinstance(v, (int, float)) else 0.0)
    return out

def main():
    load_yields()
    wb = openpyxl.load_workbook(WB, data_only=True, read_only=True)
    plants = load_tarifas(wb)
    cols25 = list(range(20, 32))   # T..AE
    cols26 = list(range(3, 7))     # C..F  (2026 Ene-Abr)
    port = dict(pv=0.0, bess=0.0, fp=0.0)

    for s in SITES:
        ws = wb[s["sheet"]]
        pname, r_int, r_cap = find_rates(plants, s["tarifa_key"])
        kwh = [a+b+c for a,b,c in zip(month_vals(ws, ROWS["base"], cols25),
                                       month_vals(ws, ROWS["inter"], cols25),
                                       month_vals(ws, ROWS["punta"], cols25))]
        dmax_p = month_vals(ws, ROWS["dmax_p"], cols25)
        fp_m = month_vals(ws, ROWS["fp_cargo"], cols25) if s["fp"] else [0.0]*12
        gen_total = bess_total = fp_total = 0.0
        bind = 0
        print(f"\n=== {s['name']}  (hoja {s['sheet']}, tarifas '{pname}': int ${r_int:.3f}/kWh, cap ${r_cap:.1f}/kW-mes; "
              f"PV {s['pv']} kWp @ {s['yield_city']} {s['yield_yr']:.0f} kWh/kWp; BESS {s['bkw']} kW/{s['bkwh']} kWh) ===")
        print("| Mes | Consumo MWh | Dem.punta kW | Umbral kW | Gen PV MWh | Ahorro PV $ | Shave ef. kW | Ahorro BESS $ | Ahorro FP $ | Total mes $ |")
        print("|---|---|---|---|---|---|---|---|---|---|")
        arb_m = ARB_PENALTY_KWH_YR * s["bkwh"] / 12.0
        rows_out = []
        for i in range(12):
            d = calendar.monthrange(2025, i+1)[1]
            umbral = kwh[i] / (d * FC * 24.0)
            gen = s["pv"] * s["yield_m"][i]
            zaf = 0.9 if (i+1) in s["zafra"] else 1.0
            pv_val = gen * r_int * zaf
            billed_b = min(dmax_p[i], umbral)
            billed_a = min(max(dmax_p[i] - s["bkw"], 0.0), umbral)
            eff = max(0.0, billed_b - billed_a)
            bess_val = (eff * r_cap - arb_m) if s["bkw"] else 0.0
            if umbral < dmax_p[i]:
                bind += 1
            tot = pv_val + bess_val + fp_m[i]
            gen_total += gen; bess_total += bess_val; fp_total += fp_m[i]; port["pv"] += pv_val
            rows_out.append((MES[i], kwh[i]/1000, dmax_p[i], umbral, gen/1000, pv_val, eff, bess_val, fp_m[i], tot))
        port["bess"] += bess_total; port["fp"] += fp_total
        pv_sum = sum(r[5] for r in rows_out); tot_sum = sum(r[9] for r in rows_out)
        for r in rows_out:
            print(f"| {r[0]} | {r[1]:,.0f} | {r[2]:,.0f} | {r[3]:,.0f} | {r[4]:,.1f} | {r[5]:,.0f} | {r[6]:,.0f} | {r[7]:,.0f} | {r[8]:,.0f} | {r[9]:,.0f} |")
        print(f"| **Año** | {sum(r[1] for r in rows_out):,.0f} |  |  | {gen_total/1000:,.1f} | **{pv_sum:,.0f}** |  | **{bess_total:,.0f}** | **{fp_total:,.0f}** | **{tot_sum:,.0f}** |")
        print(f"umbral binds (umbral < Dem.punta): {bind}/12")

        # Proplasa / IXT supplier split (Total Planta vs CFE block), 2025 + 2026 YTD
        if s["sheet"] in ("PR1","PR2","TAP","IXT"):
            cfe25 = [a+b+c for a,b,c in zip(month_vals(ws, CFE_ROWS["base"], cols25),
                                             month_vals(ws, CFE_ROWS["inter"], cols25),
                                             month_vals(ws, CFE_ROWS["punta"], cols25))]
            aa25 = [max(t-c, 0.0) for t,c in zip(kwh, cfe25)]
            tot26 = [a+b+c for a,b,c in zip(month_vals(ws, ROWS["base"], cols26),
                                             month_vals(ws, ROWS["inter"], cols26),
                                             month_vals(ws, ROWS["punta"], cols26))]
            cfe26 = [a+b+c for a,b,c in zip(month_vals(ws, CFE_ROWS["base"], cols26),
                                             month_vals(ws, CFE_ROWS["inter"], cols26),
                                             month_vals(ws, CFE_ROWS["punta"], cols26))]
            share25 = [f"{a/t*100:.0f}%" if t else "-" for a,t in zip(aa25, kwh)]
            share26 = [f"{max(t-c,0)/t*100:.0f}%" if t else "-" for t,c in zip(tot26, cfe26)]
            print(f"split autoabasto 2025 (% del kWh del mes): {' '.join(share25)}")
            print(f"split autoabasto 2026 Ene-Abr: {' '.join(share26)}")

    print("\n=== VALIDATION vs filed annuals (2026-06-10 best-case) ===")
    print(f"PV   base: {port['pv']:,.0f}  (filed 44,740,000  delta {port['pv']/44.74e6-1:+.1%})")
    print(f"BESS base: {port['bess']:,.0f} (filed  9,870,000  delta {port['bess']/9.87e6-1:+.1%})")
    print(f"FP   base: {port['fp']:,.0f}  (filed  2,310,000  delta {port['fp']/2.31e6-1:+.1%})")
    print(f"TOTAL    : {port['pv']+port['bess']+port['fp']:,.0f} (filed 56,930,000)")

main()
