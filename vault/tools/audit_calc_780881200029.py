# Audit extractor for raw/data/780881200029.xlsx (Calculadora BESS PV).
# Usage: python audit_calc_780881200029.py SHEET [min_row max_row max_col] [--formulas]
# Prints non-empty cells compactly; values by default, formulas with --formulas.
import openpyxl, sys

sys.stdout.reconfigure(encoding="utf-8")
WB = r"C:\Users\vidan\Desktop\CFE Brain\raw\data\780881200029.xlsx"

args = [a for a in sys.argv[1:] if a != "--formulas"]
formulas = "--formulas" in sys.argv
sheet = args[0]
r1 = int(args[1]) if len(args) > 1 else 1
r2 = int(args[2]) if len(args) > 2 else 120
c2 = int(args[3]) if len(args) > 3 else 30

wb = openpyxl.load_workbook(WB, data_only=not formulas, read_only=True)
if sheet == "LIST":
    for ws in wb.worksheets:
        print(ws.title, ws.max_row, ws.max_column)
    sys.exit()
ws = wb[sheet]
print(f"# {sheet} max_row={ws.max_row} max_col={ws.max_column} rows {r1}-{r2} cols<= {c2} {'FORMULAS' if formulas else 'VALUES'}")
for row in ws.iter_rows(min_row=r1, max_row=r2, max_col=c2):
    cells = []
    for c in row:
        v = c.value
        if v is None or v == "":
            continue
        if isinstance(v, float):
            v = f"{v:,.4g}" if abs(v) < 1 else f"{v:,.2f}".rstrip("0").rstrip(".")
        s = str(v)
        if len(s) > 60:
            s = s[:57] + "..."
        cells.append(f"{c.coordinate}={s}")
    if cells:
        print(" | ".join(cells))
