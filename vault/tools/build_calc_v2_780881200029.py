# Build "780881200029 - Calculadora v2.xlsx" — modular, corrected rebuild of the client's
# Calculadora BESS PV (raw/data/780881200029.xlsx).
#
# Corrections vs v1:
#   C1 BILLS carries the FP bonificacion (v1 left it blank -> +1.2% validation gap; v2 ~0.0%)
#   C2 Arbitrage: discharge only punta weekdays (Mon-Fri minus festivos), capped at the
#      bill's punta kWh (v1 cycled 365 d/yr uncapped -> arb overstated ~35%)
#   C3 Bonificacion claw-back: savings shrink the bill -> the ~1.2% FP credit shrinks too
#   C4 O&M PV escalates (1+esc)^(yr-1) (v1 exponent bug: ^(factor-1) ~= flat)
#   C5 Falla factor applies to BESS savings only, not to O&M BESS (v1 multiplied O&M too)
#   C6 One perspective per column in FINANCE; TIR/VPN labeled (proyecto vs financiador)
#
# Structure (one concern per sheet): README / INPUTS / BILLS / RATES / SAVINGS / FINANCE.
# Every formula is mirrored in python; the script prints the expected values (Excel must
# recalc to the same numbers when opened).
import openpyxl, sys, math, calendar
from datetime import date
from pathlib import Path
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Font

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, r"C:\Users\vidan\Desktop\CFE Brain\tools")
from cfe_savings import extract

V1 = r"C:\Users\vidan\Desktop\CFE Brain\raw\data\780881200029.xlsx"
OUT = r"C:\Users\vidan\Desktop\CFE Brain\780881200029 - Calculadora v2.xlsx"
BILLS_DIR = r"C:\Users\vidan\Desktop\CFE Brain\raw\bills\780881200029"

# ---------------- source data ----------------
def festivos(y):
    def nth_monday(year, month, n):
        d = date(year, month, 1); off = (7 - d.weekday()) % 7
        return date(year, month, 1 + off + 7 * (n - 1))
    return {date(y,1,1), nth_monday(y,2,1), nth_monday(y,3,3), date(y,5,1),
            date(y,9,16), nth_monday(y,11,3), date(y,12,25)}

def punta_weekdays(y, m):
    f = festivos(y)
    return sum(1 for d in range(1, calendar.monthrange(y, m)[1] + 1)
               if date(y, m, d).weekday() < 5 and date(y, m, d) not in f)

wb1 = openpyxl.load_workbook(V1, data_only=True, read_only=True)
tr = wb1["TARIFF_RATES"]
pi = wb1["PROJECT_INPUTS"]

MONTH_KEY = {"ABR":(2025,4),"MAY":(2025,5),"JUN":(2025,6),"JUL":(2025,7),"AGO":(2025,8),
             "SEP":(2025,9),"OCT":(2025,10),"NOV":(2025,11),"DIC":(2025,12),
             "ENE":(2026,1),"FEB":(2026,2),"MAR":(2026,3)}
bonif = {}
for pdf in Path(BILLS_DIR).glob("*.pdf"):
    b = extract.parse_bill(str(pdf))
    bonif[MONTH_KEY[pdf.stem.split()[-1][:3]]] = b.get("bonif_fp", 0.0) or 0.0

months = []
for i in range(12):
    r1, r2 = 6 + i, 21 + i
    per = tr.cell(row=r1, column=1).value
    g = lambda c, r=r1: tr.cell(row=r, column=c).value
    g2 = lambda c, r=r2: tr.cell(row=r, column=c).value
    months.append(dict(
        y=per.year, m=per.month, days=g(2), wd=punta_weekdays(per.year, per.month),
        kwh_b=g(7), kwh_i=g(8), kwh_p=g(9), kw_b=g(11), kw_i=g(12), kw_p=g(13),
        cargo_fijo=g2(3), dist=g2(4), trans=g2(5), cenace=g2(6),
        gen_b=g2(7), gen_i=g2(8), gen_p=g2(9), capacidad=g2(10), scnmem=g2(11),
        recibo=g2(14), bonif=bonif[(per.year, per.month)],
    ))

# Helioscope monthlies (PROJECT_INPUTS B46:D57, ENE..DIC, for the 839.41 kWp reference)
HELIO_REF_KWP = pi.cell(row=45, column=3).value            # 839.41
helio = {m: pi.cell(row=45 + m, column=4).value for m in range(1, 13)}

P = dict(kwp=194.48, pv_degr=0.005, epc_usd_wp=0.75, fx=17.55, om_pv=60288.8,
         ppa=1.5, ppa_yrs=15, esc_ppa=0.05,
         bess_kwh=2940.0, bess_kw=1503.0, dod=0.96, rte=0.96, bess_degr=0.0125,
         epc_bess=280.0, om_bess=529200.0, comision=0.5, falla_meses=3, bess_yrs=15,
         fianza=0.139, wacc=0.12, esc_cfe=0.06, horizon=20)
P["capex_pv"] = P["kwp"] * 1000 * P["epc_usd_wp"] * P["fx"]
P["capex_bess"] = P["bess_kwh"] * P["epc_bess"] * P["fx"]
P["deliv"] = P["bess_kwh"] * P["dod"] * math.sqrt(P["rte"])   # kWh deliverable per cycle
P["falla"] = (12 - P["falla_meses"]) / 12

FC = 0.57

# ---------------- python mirror (ground truth for verification) ----------------
mir = []
for mo in months:
    kwh_tot = mo["kwh_b"] + mo["kwh_i"] + mo["kwh_p"]
    mem = (mo["cargo_fijo"] + mo["dist"] + mo["trans"] + mo["cenace"] + mo["gen_b"]
           + mo["gen_i"] + mo["gen_p"] + mo["capacidad"] + mo["scnmem"])
    umbral = kwh_tot / (mo["days"] * 24 * FC)
    basis_cap = min(mo["kw_p"], umbral)
    basis_dist = min(max(mo["kw_b"], mo["kw_i"], mo["kw_p"]), umbral)
    r_cap = mo["capacidad"] / basis_cap
    r_trans, r_cen, r_scn = mo["trans"]/kwh_tot, mo["cenace"]/kwh_tot, mo["scnmem"]/kwh_tot
    bnd_b = mo["gen_b"]/mo["kwh_b"] + r_trans + r_cen + r_scn
    bnd_i = mo["gen_i"]/mo["kwh_i"] + r_trans + r_cen + r_scn
    bnd_p = mo["gen_p"]/mo["kwh_p"] + r_trans + r_cen + r_scn
    gen = P["kwp"] * helio[mo["m"]] / HELIO_REF_KWP
    ahorro_pv = gen * bnd_i
    umbral_post = (kwh_tot - gen) / (mo["days"] * 24 * FC)
    billed_pre = min(mo["kw_p"], umbral)
    billed_prebess = min(mo["kw_p"], umbral_post)
    dem_pv = (billed_pre - billed_prebess) * r_cap
    ph = 4 if (mo["m"] <= 3 or mo["m"] >= 11) else 2
    shave = min(P["bess_kw"], P["deliv"] / ph)
    residual = max(0.0, mo["kw_p"] - shave)
    billed_post = min(residual, umbral_post)
    dem_bess = (billed_prebess - billed_post) * r_cap
    desc = min(P["deliv"] * mo["wd"], mo["kwh_p"])
    carga = desc / P["rte"]
    arb = desc * bnd_p - carga * bnd_b
    sub = ahorro_pv + dem_pv + dem_bess + arb
    bonif_share = -mo["bonif"] / mem
    claw = -sub * bonif_share
    total = sub + claw
    antes = mem + mo["bonif"]
    mir.append(dict(mo, kwh_tot=kwh_tot, mem=mem, gen=gen, ahorro_pv=ahorro_pv,
                    dem_pv=dem_pv, dem_bess=dem_bess, desc=desc, arb=arb, claw=claw,
                    total=total, antes=antes, check=(mem + mo["bonif"]) * 1.16 - mo["recibo"]))

A = lambda k: sum(r[k] for r in mir)
an = dict(pv=A("ahorro_pv"), dem=A("dem_bess") + A("dem_pv"), arb=A("arb"),
          claw=A("claw"), total=A("total"), mem=A("mem"), antes=A("antes"), gen=A("gen"))

# 20-yr finance mirror
def finance():
    rows, fl_cli, fl_new, fl_proj = [], [-0.0], [-(P["capex_pv"] + P["capex_bess"])], [-(P["capex_pv"] + P["capex_bess"])]
    for yy in range(1, P["horizon"] + 1):
        esc = (1 + P["esc_cfe"]) ** yy
        dpv = (1 - P["pv_degr"]) ** yy
        dbe = (1 - P["bess_degr"]) ** yy
        sav_pv = an["pv"] * esc * dpv
        sav_bess = (an["dem"] + an["arb"]) * esc * dbe * P["falla"]
        claw = an["claw"] * esc * (dpv + dbe) / 2          # proportional, midpoint degr
        bruto = sav_pv + sav_bess + claw
        ppa_pay = (an["gen"] * dpv * P["ppa"] * (1 + P["esc_ppa"]) ** (yy - 1)) if yy <= P["ppa_yrs"] else 0.0
        comis = sav_bess * P["comision"] if yy <= P["bess_yrs"] else 0.0
        om_pv = P["om_pv"] * (1 + P["esc_cfe"]) ** (yy - 1) if yy <= P["ppa_yrs"] else 0.0
        om_bess = P["om_bess"] * (1 + P["esc_cfe"]) ** (yy - 1) if yy <= P["bess_yrs"] else 0.0
        fianza = ppa_pay * P["fianza"]
        cli = bruto - ppa_pay - comis
        new = ppa_pay + comis - om_pv - om_bess - fianza
        proj = bruto - om_pv - om_bess
        rows.append((yy, bruto, ppa_pay, comis, om_pv, om_bess, fianza, cli, new, proj))
        fl_cli.append(cli); fl_new.append(new); fl_proj.append(proj)
    return rows, fl_cli, fl_new, fl_proj

def irr(flows, lo=-0.95, hi=2.0):
    def npv(r): return sum(f / (1 + r) ** i for i, f in enumerate(flows))
    if npv(lo) * npv(hi) > 0: return float("nan")
    for _ in range(200):
        mid = (lo + hi) / 2
        if npv(lo) * npv(mid) <= 0: hi = mid
        else: lo = mid
    return (lo + hi) / 2

def npv12(flows): return sum(f / (1 + P["wacc"]) ** i for i, f in enumerate(flows))
def payback(flows):
    c = 0.0
    for i, f in enumerate(flows):
        c += f
        if c >= 0: return i
    return None

frows, fl_cli, fl_new, fl_proj = finance()
ind = dict(tir_proj=irr(fl_proj), tir_new=irr(fl_new), vpn_new=npv12(fl_new),
           vpn_proj=npv12(fl_proj), pb_new=payback(fl_new), pb_proj=payback(fl_proj))

# ---------------- report (audit numbers) ----------------
print("=== v2 PYTHON MIRROR — expected workbook values ===")
print(f"validacion por mes (calc vs recibo): max |diff| = {max(abs(r['check']) for r in mir):,.2f} MXN")
print(f"Bill ANTES anual (MEM, comparable v1): {an['mem']:,.2f}")
print(f"Bill ANTES anual real (con bonif FP):  {an['antes']:,.2f}")
print(f"Ahorro PV:        {an['pv']:,.2f}   (v1: 787,522.28)")
print(f"Ahorro demanda:   {an['dem']:,.2f}   (v1: 5,590,209.47)")
print(f"Ahorro arbitraje: {an['arb']:,.2f}   (v1: 1,216,237.59)")
print(f"Claw-back bonif:  {an['claw']:,.2f}   (v1: no modelado)")
print(f"AHORRO BRUTO año-0: {an['total']:,.2f}  = {an['total']/an['mem']:.2%} del MEM, {an['total']/an['antes']:.2%} del real")
print(f"\nTIR proyecto (unlevered): {ind['tir_proj']:.2%}  payback {ind['pb_proj']} años  VPN@12% {ind['vpn_proj']:,.0f}")
print(f"TIR financiador:          {ind['tir_new']:.2%}  payback {ind['pb_new']} años  VPN@12% {ind['vpn_new']:,.0f}")
print(f"(v1 reportaba TIR 19.02%, VPN 7,834,725, payback 6 — todos del financiador, sin etiquetar)")

# ---------------- build workbook ----------------
wb = openpyxl.Workbook()
bold = Font(bold=True)

def sheet(name, idx=None):
    ws = wb.create_sheet(name) if name not in wb.sheetnames else wb[name]
    return ws

ws0 = wb.active; ws0.title = "README"
readme = [
    ["780881200029 — Calculadora PV+BESS v2 (rebuild modular y corregido)"],
    ["Generado por tools/build_calc_v2_780881200029.py — cada fórmula tiene espejo en python."],
    [""],
    ["MÓDULOS (una hoja = una responsabilidad):"],
    ["  INPUTS   — todos los supuestos editables (sin constantes enterradas)"],
    ["  BILLS    — 12 recibos transcritos + bonificación FP; valida (MEM+bonif)×1.16 vs recibo ≈ 0.0%"],
    ["  RATES    — tarifas unitarias implícitas por mes (umbral, bases facturables correctas)"],
    ["  SAVINGS  — ahorro mensual por palanca: PV, demanda BESS, arbitraje corregido, claw-back bonif"],
    ["  FINANCE  — proyección 20 años; flujos cliente / financiador / proyecto SEPARADOS y etiquetados"],
    [""],
    ["CORRECCIONES vs calculadora v1:"],
    ["  C1 Bonificación FP transcrita — el 'gap' de +1.2% en la validación v1 era esta omisión"],
    ["  C2 Arbitraje: la batería solo cicla días hábiles punta (lun–vie menos festivos) y no puede"],
    ["     desplazar más kWh punta que los del recibo. v1: 365 ciclos/año sin tope → +$426k/año irreales"],
    ["  C3 Claw-back: el ahorro reduce el recibo → la bonificación FP (~1.2%) se reduce en proporción"],
    ["  C4 O&M PV escala (1+esc)^(año−1); v1 tenía ^(factor−1) → O&M casi plano (subestimado)"],
    ["  C5 Factor falla (0.75) aplica al ahorro BESS, no al O&M BESS"],
    ["  C6 TIR/VPN etiquetados: proyecto (unlevered) vs financiador; v1 mezclaba perspectivas"],
    [""],
    ["SIN CAMBIO (lógica v1 correcta): shave plano usable/horas-punta (conservador, garantizable);"],
    ["  capacidad sobre min(kW punta, umbral); distribución sobre min(max(B,I,P), umbral);"],
    ["  PV 100% intermedio (SIN); atribución de demanda PV vs BESS vía umbral post-PV."],
    [""],
    ["Grado: prefactibilidad sobre curva modelada; datos 15-min harían el shave bancable."],
]
for r, row in enumerate(readme, 1):
    ws0.cell(row=r, column=1, value=row[0])
ws0.cell(row=1, column=1).font = bold

# INPUTS — rows assigned programmatically; REF registry eliminates hand-counted refs
wsI = sheet("INPUTS")
REF = {}
inputs_rows = [
    ("PV", None, None, None),
    ("Capacidad PV (kWp)", "kwp_", P["kwp"], "kWp"),
    ("Referencia Helioscope (kWp)", "href_", HELIO_REF_KWP, "kWp del diseño Helioscope"),
    ("Degradación PV (%/año)", "pvdeg_", P["pv_degr"], ""),
    ("EPC PV (USD/Wp)", "epcpv_", P["epc_usd_wp"], ""),
    ("Tipo de cambio (MXN/USD)", "fx_", P["fx"], ""),
    ("CAPEX PV (MXN)", "capexpv_", "F:={kwp_}*1000*{epcpv_}*{fx_}", "= kWp×1000×EPC×FX"),
    ("O&M PV (MXN/año, año 1)", "ompv_", P["om_pv"], "escala con tarifa CFE"),
    ("Tarifa PPA (MXN/kWh)", "ppa_", P["ppa"], ""),
    ("Plazo PPA (años)", "ppayrs_", P["ppa_yrs"], ""),
    ("Escalación PPA (%/año)", "escppa_", P["esc_ppa"], "desde año 2"),
    ("BESS", None, None, None),
    ("Capacidad nominal (kWh)", "bkwh_", P["bess_kwh"], ""),
    ("Potencia (kW)", "bkw_", P["bess_kw"], ""),
    ("DoD", "dod_", P["dod"], ""),
    ("RTE", "rte_", P["rte"], ""),
    ("kWh entregables por ciclo", "deliv_", "F:={bkwh_}*{dod_}*SQRT({rte_})", "= kWh×DoD×√RTE"),
    ("Degradación BESS (%/año)", "bdeg_", P["bess_degr"], ""),
    ("EPC BESS (USD/kWh)", "epcb_", P["epc_bess"], ""),
    ("CAPEX BESS (MXN)", "capexb_", "F:={bkwh_}*{epcb_}*{fx_}", ""),
    ("O&M BESS (MXN/año, año 1)", "omb_", P["om_bess"], "NO se multiplica por falla (C5)"),
    ("Comisión financiador (% ahorro BESS)", "comis_", P["comision"], ""),
    ("Meses falla BESS/año", "fallam_", P["falla_meses"], ""),
    ("Factor falla", "falla_", "F:=(12-{fallam_})/12", "aplica solo al ahorro BESS"),
    ("Plazo servicio BESS (años)", "byrs_", P["bess_yrs"], ""),
    ("FINANZAS", None, None, None),
    ("WACC", "wacc_", P["wacc"], ""),
    ("Escalación tarifa CFE (%/año)", "esc_", P["esc_cfe"], ""),
    ("Fianza (% del PPA)", "fianza_", P["fianza"], ""),
    ("Horizonte (años)", "hor_", P["horizon"], ""),
]
# first pass: assign rows
r = 2
for label, key, val, note in inputs_rows:
    if key:
        REF[key] = f"$B${r}"
    r += 1
# second pass: write (formulas resolve via REF)
r = 2
for label, key, val, note in inputs_rows:
    wsI.cell(row=r, column=1, value=label).font = bold if key is None else Font()
    if key:
        if isinstance(val, str) and val.startswith("F:"):
            val = val[2:].format(**REF)
        wsI.cell(row=r, column=2, value=val)
        wsI.cell(row=r, column=3, value=note)
    r += 1
wsI.cell(row=1, column=1, value="INPUTS — único lugar con supuestos editables").font = bold
# Helioscope monthly block (cols E:F)
wsI.cell(row=1, column=5, value="Helioscope kWh/mes (ref kWp)").font = bold
for m in range(1, 13):
    wsI.cell(row=1 + m, column=5, value=m)
    wsI.cell(row=1 + m, column=6, value=helio[m])

NAMES = {"KWP": REF["kwp_"], "HREF": REF["href_"], "DELIV": REF["deliv_"], "BKW": REF["bkw_"],
         "RTE_": REF["rte_"], "FALLA": REF["falla_"], "COMIS": REF["comis_"],
         "WACC_": REF["wacc_"], "ESC": REF["esc_"]}
for nm, ref in NAMES.items():
    wb.defined_names.add(DefinedName(nm, attr_text=f"INPUTS!{ref}"))
I = {k: f"INPUTS!{v}" for k, v in REF.items()}   # for direct formula refs

# BILLS
wsB = sheet("BILLS")
hdrB = ["Periodo", "Días", "Días hábiles punta", "kWh Base", "kWh Inter", "kWh Punta", "kWh Total",
        "kW Base", "kW Inter", "kW Punta", "Cargo Fijo $", "Distribución $", "Transmisión $",
        "CENACE $", "Gen B $", "Gen I $", "Gen P $", "Capacidad $", "SCnMEM $", "Subtotal MEM $",
        "Bonif FP $", "Bonif % MEM", "Total calc c/IVA", "Recibo CFE $", "Diff %", "Check"]
for c, h in enumerate(hdrB, 1):
    wsB.cell(row=1, column=c, value=h).font = bold
for r, mo in enumerate(months, 2):
    wsB.cell(row=r, column=1, value=date(mo["y"], mo["m"], 1))
    vals = [mo["days"], mo["wd"], mo["kwh_b"], mo["kwh_i"], mo["kwh_p"]]
    for c, v in enumerate(vals, 2):
        wsB.cell(row=r, column=c, value=v)
    wsB.cell(row=r, column=7, value=f"=D{r}+E{r}+F{r}")
    for c, v in zip(range(8, 20), [mo["kw_b"], mo["kw_i"], mo["kw_p"], mo["cargo_fijo"],
                                   mo["dist"], mo["trans"], mo["cenace"], mo["gen_b"],
                                   mo["gen_i"], mo["gen_p"], mo["capacidad"], mo["scnmem"]]):
        wsB.cell(row=r, column=c, value=v)
    wsB.cell(row=r, column=20, value=f"=SUM(K{r}:S{r})")
    wsB.cell(row=r, column=21, value=mo["bonif"])
    wsB.cell(row=r, column=22, value=f"=-U{r}/T{r}")
    wsB.cell(row=r, column=23, value=f"=(T{r}+U{r})*1.16")
    wsB.cell(row=r, column=24, value=mo["recibo"])
    wsB.cell(row=r, column=25, value=f"=W{r}/X{r}-1")
    wsB.cell(row=r, column=26, value=f'=IF(ABS(Y{r})<0.005,"OK","CHECK")')
wsB.cell(row=14, column=1, value="TOTAL").font = bold
for c in [4,5,6,7,11,12,13,14,15,16,17,18,19,20,21,23,24]:
    L = openpyxl.utils.get_column_letter(c)
    wsB.cell(row=14, column=c, value=f"=SUM({L}2:{L}13)")

# RATES
wsR = sheet("RATES")
hdrR = ["Periodo", "Umbral kW (FC 0.57)", "Base fact. capacidad", "Base fact. distribución",
        "$/kW Capacidad", "$/kW Distribución", "$/kWh Trans", "$/kWh CENACE", "$/kWh SCnMEM",
        "$/kWh Gen B", "$/kWh Gen I", "$/kWh Gen P", "Bundled Base", "Bundled Inter", "Bundled Punta"]
for c, h in enumerate(hdrR, 1):
    wsR.cell(row=1, column=c, value=h).font = bold
for r in range(2, 14):
    wsR.cell(row=r, column=1, value=f"=BILLS!A{r}")
    wsR.cell(row=r, column=2, value=f"=BILLS!G{r}/(BILLS!B{r}*24*0.57)")
    wsR.cell(row=r, column=3, value=f"=MIN(BILLS!J{r},B{r})")
    wsR.cell(row=r, column=4, value=f"=MIN(MAX(BILLS!H{r}:J{r}),B{r})")
    wsR.cell(row=r, column=5, value=f"=BILLS!R{r}/C{r}")
    wsR.cell(row=r, column=6, value=f"=BILLS!L{r}/D{r}")
    wsR.cell(row=r, column=7, value=f"=BILLS!M{r}/BILLS!G{r}")
    wsR.cell(row=r, column=8, value=f"=BILLS!N{r}/BILLS!G{r}")
    wsR.cell(row=r, column=9, value=f"=BILLS!S{r}/BILLS!G{r}")
    wsR.cell(row=r, column=10, value=f"=BILLS!O{r}/BILLS!D{r}")
    wsR.cell(row=r, column=11, value=f"=BILLS!P{r}/BILLS!E{r}")
    wsR.cell(row=r, column=12, value=f"=BILLS!Q{r}/BILLS!F{r}")
    wsR.cell(row=r, column=13, value=f"=J{r}+G{r}+H{r}+I{r}")
    wsR.cell(row=r, column=14, value=f"=K{r}+G{r}+H{r}+I{r}")
    wsR.cell(row=r, column=15, value=f"=L{r}+G{r}+H{r}+I{r}")

# SAVINGS
wsS = sheet("SAVINGS")
hdrS = ["Periodo", "Gen PV kWh", "Ahorro PV $", "Umbral post-PV", "Fact. pre (kW)",
        "Fact. pre-BESS (kW)", "Ahorro dem. PV $", "Horas punta", "Shave kW", "Pico residual kW",
        "Fact. post (kW)", "Ahorro dem. BESS $", "Descarga kWh (corr.)", "Carga kWh",
        "Ahorro arbitraje $", "Subtotal $", "Claw-back bonif $", "AHORRO TOTAL $",
        "Factura ANTES $ (real)", "Factura DESPUÉS $", "Ahorro %"]
for c, h in enumerate(hdrS, 1):
    wsS.cell(row=1, column=c, value=h).font = bold
for i, r in enumerate(range(2, 14)):
    m = months[i]["m"]
    wsS.cell(row=r, column=1, value=f"=BILLS!A{r}")
    wsS.cell(row=r, column=2, value=f"=KWP*INPUTS!$F${1 + m}/HREF")
    wsS.cell(row=r, column=3, value=f"=B{r}*RATES!N{r}")
    wsS.cell(row=r, column=4, value=f"=(BILLS!G{r}-B{r})/(BILLS!B{r}*24*0.57)")
    wsS.cell(row=r, column=5, value=f"=MIN(BILLS!J{r},RATES!B{r})")
    wsS.cell(row=r, column=6, value=f"=MIN(BILLS!J{r},D{r})")
    wsS.cell(row=r, column=7, value=f"=(E{r}-F{r})*RATES!E{r}")
    wsS.cell(row=r, column=8, value=4 if (m <= 3 or m >= 11) else 2)
    wsS.cell(row=r, column=9, value=f"=MIN(BKW,DELIV/H{r})")
    wsS.cell(row=r, column=10, value=f"=MAX(0,BILLS!J{r}-I{r})")
    wsS.cell(row=r, column=11, value=f"=MIN(J{r},D{r})")
    wsS.cell(row=r, column=12, value=f"=(F{r}-K{r})*RATES!E{r}")
    wsS.cell(row=r, column=13, value=f"=MIN(DELIV*BILLS!C{r},BILLS!F{r})")
    wsS.cell(row=r, column=14, value=f"=M{r}/RTE_")
    wsS.cell(row=r, column=15, value=f"=M{r}*RATES!O{r}-N{r}*RATES!M{r}")
    wsS.cell(row=r, column=16, value=f"=C{r}+G{r}+L{r}+O{r}")
    wsS.cell(row=r, column=17, value=f"=-P{r}*BILLS!V{r}")
    wsS.cell(row=r, column=18, value=f"=P{r}+Q{r}")
    wsS.cell(row=r, column=19, value=f"=BILLS!T{r}+BILLS!U{r}")
    wsS.cell(row=r, column=20, value=f"=S{r}-R{r}")
    wsS.cell(row=r, column=21, value=f"=R{r}/S{r}")
wsS.cell(row=14, column=1, value="TOTAL AÑO 0").font = bold
for c in [2,3,7,12,13,15,16,17,18,19,20]:
    L = openpyxl.utils.get_column_letter(c)
    wsS.cell(row=14, column=c, value=f"=SUM({L}2:{L}13)")
wsS.cell(row=14, column=21, value="=R14/S14")
wsS.cell(row=16, column=1, value="Comparación v1 → v2 (anual):").font = bold
cmp_rows = [("Ahorro PV", 787522.28, "=C14"), ("Demanda BESS (+PV)", 5590209.47, "=G14+L14"),
            ("Arbitraje", 1216237.59, "=O14"), ("Claw-back bonif", 0.0, "=Q14"),
            ("TOTAL", 7593969.34, "=R14")]
wsS.cell(row=17, column=2, value="v1").font = bold
wsS.cell(row=17, column=3, value="v2").font = bold
for r, (lbl, v1v, f2) in enumerate(cmp_rows, 18):
    wsS.cell(row=r, column=1, value=lbl)
    wsS.cell(row=r, column=2, value=v1v)
    wsS.cell(row=r, column=3, value=f2)

# FINANCE
wsF = sheet("FINANCE")
hdrF = ["Año", "Esc CFE", "Degr PV", "Degr BESS", "Ahorro PV $", "Ahorro BESS $ (×falla)",
        "Claw-back $", "Ahorro Bruto $", "Pago PPA $", "Comisión BESS $", "O&M PV $",
        "O&M BESS $", "Fianza $", "Flujo CLIENTE $", "Flujo FINANCIADOR $", "Flujo PROYECTO $",
        "Acum. financiador"]
for c, h in enumerate(hdrF, 1):
    wsF.cell(row=6, column=c, value=h).font = bold
wsF.cell(row=7, column=1, value=0)
wsF.cell(row=7, column=15, value=f"=-({I['capexpv_']}+{I['capexb_']})")
wsF.cell(row=7, column=16, value=f"=-({I['capexpv_']}+{I['capexb_']})")
wsF.cell(row=7, column=17, value="=O7")
for yy in range(1, 21):
    r = 7 + yy
    wsF.cell(row=r, column=1, value=yy)
    wsF.cell(row=r, column=2, value=f"=(1+ESC)^A{r}")
    wsF.cell(row=r, column=3, value=f"=(1-{I['pvdeg_']})^A{r}")
    wsF.cell(row=r, column=4, value=f"=(1-{I['bdeg_']})^A{r}")
    wsF.cell(row=r, column=5, value=f"=SAVINGS!$C$14*B{r}*C{r}")
    wsF.cell(row=r, column=6, value=f"=(SAVINGS!$G$14+SAVINGS!$L$14+SAVINGS!$O$14)*B{r}*D{r}*FALLA")
    wsF.cell(row=r, column=7, value=f"=SAVINGS!$Q$14*B{r}*(C{r}+D{r})/2")
    wsF.cell(row=r, column=8, value=f"=E{r}+F{r}+G{r}")
    wsF.cell(row=r, column=9, value=f"=IF(A{r}<={I['ppayrs_']},SAVINGS!$B$14*C{r}*{I['ppa_']}*(1+{I['escppa_']})^(A{r}-1),0)")
    wsF.cell(row=r, column=10, value=f"=IF(A{r}<={I['byrs_']},F{r}*COMIS,0)")
    wsF.cell(row=r, column=11, value=f"=IF(A{r}<={I['ppayrs_']},{I['ompv_']}*(1+ESC)^(A{r}-1),0)")
    wsF.cell(row=r, column=12, value=f"=IF(A{r}<={I['byrs_']},{I['omb_']}*(1+ESC)^(A{r}-1),0)")
    wsF.cell(row=r, column=13, value=f"=I{r}*{I['fianza_']}")
    wsF.cell(row=r, column=14, value=f"=H{r}-I{r}-J{r}")
    wsF.cell(row=r, column=15, value=f"=I{r}+J{r}-K{r}-L{r}-M{r}")
    wsF.cell(row=r, column=16, value=f"=H{r}-K{r}-L{r}")
    wsF.cell(row=r, column=17, value=f"=Q{r-1}+O{r}")
ind_rows = [
    ("TIR PROYECTO (unlevered)", "=IRR(P7:P27)", f"{ind['tir_proj']:.2%}"),
    ("TIR FINANCIADOR", "=IRR(O7:O27)", f"{ind['tir_new']:.2%}"),
    ("VPN proyecto @WACC", "=NPV(WACC_,P8:P27)+P7", f"{ind['vpn_proj']:,.0f}"),
    ("VPN financiador @WACC", "=NPV(WACC_,O8:O27)+O7", f"{ind['vpn_new']:,.0f}"),
    ("Payback financiador (años, simple)", "=COUNTIF(Q7:Q27,\"<0\")", f"{ind['pb_new']}"),
]
wsF.cell(row=1, column=1, value="INDICADORES (etiquetados — C6)").font = bold
for r, (lbl, f, expected) in enumerate(ind_rows, 2):
    wsF.cell(row=r, column=1, value=lbl)
    wsF.cell(row=r, column=2, value=f)
    wsF.cell(row=r, column=4, value=f"esperado: {expected}")

wb.save(OUT)
print(f"\nWorkbook escrito: {OUT}")
