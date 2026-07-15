#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_books_cs.py — Build the two GEPP rev4 "Solar-Charge BESS" workbooks by
ZIP/XML surgery on the delivered rev3 books (charts + drawings preserved byte-identical).

Per site sheet (7 total), the surgery is MINIMAL — everything financial (J,K,M,P,R,
totals, C41-C55, projection 94-139, Resumen, Supuestos) stays LIVE and recomputes:
  * Section 2:  C26 (Op1 kWp, ixt/aca/can only) ; C31=rev3*k_op1 ; D31=rev3*k_op2.
  * Monthly Op1 rows 60-71 / Op2 rows 77-88:
      - D/E/F  -> injected self-basis %Base/%Int/%Punta VALUES (motor).
      - O      -> LIVE formula referencing the new carga-base/solar cols:
                    Op1  O{r}=W{r}*{col}$14                (c_oport = texc ~= 0)
                    Op2  O{r}=W{r}*{col}$14+V{r}*{col}$15  (c_oport = t_int)
                  ({col} = B..M per-month tariff column, preserving rev3 semantics).
      - U/V/W  -> NEW value columns: Exceso solar / Carga solar BESS / Carga base (kWh).
  * Row 73 note rewritten for the solar-charge dispatch.
  * Typical day (invierno 152-175, verano 188-211): col D (pv) & col F (BESS carga)
    replaced with motor solar-first placement; cols C/E/G/H untouched (live).
  * fullCalcOnLoad=1 ; <dimension> extended to W219.

Reads motor_cs.json (precomputed dispatch) + solve_k_cs.json (PPA k). READ-ONLY on rev3.
Rerunnable.
"""
import zipfile, re, os, json, importlib.util

HERE = os.path.dirname(os.path.abspath(__file__))
REV3 = '/home/mario/CFE Brain/work/gepp-deck'
import openpyxl

MOTOR = json.load(open(os.path.join(HERE, 'motor_cs.json')))
KSOLVE = json.load(open(os.path.join(HERE, 'solve_k_cs.json')))
K_OP1 = KSOLVE['op1']['k_selfbasis']
K_OP2 = KSOLVE['op2']['k_selfbasis']

NEW_KWP_OP1 = {'ixt': 5170.0, 'aca': 1748.0, 'can': 1270.0}

BOOK4_IN = os.path.join(REV3, 'GEPP - Propuesta PV+BESS 4 Sitios (2026-07) rev3 - Esc 4pct TIR 14.xlsx')
BOOKP_IN = os.path.join(REV3, 'GEPP - Propuesta Proplasa (3 sitios) (2026-07) rev3 - Esc 4pct TIR 14.xlsx')
BOOK4_OUT = os.path.join(HERE, 'GEPP - Solucion Energetica Solar-Charge BESS - 4 Sitios.xlsx')
BOOKP_OUT = os.path.join(HERE, 'GEPP - Solucion Energetica Solar-Charge BESS - Proplasa.xlsx')

# sitekey -> sheet name
SHEET = {'ixt': 'Ixtlahuacán', 'aca': 'Acapulco', 'can': 'Cancún', 'pro': 'Proplasa',
         'pr1': 'Preforma 1', 'pr2': 'Preforma 2', 'tap': 'Tapa'}
BOOK_SITES = {BOOK4_OUT: ['ixt', 'aca', 'can', 'pro'], BOOKP_OUT: ['pr1', 'pr2', 'tap']}

NOTE73 = ("Nueva regla de despacho (Carga Solar): el BESS se carga PRIMERO con el "
          "excedente solar del mediodía (columna V) y sólo el faltante se toma de la red "
          "en horario base 0–5 h (columna W); U = excedente solar disponible. Descarga "
          "idéntica (rasurado plano en punta). O (costo de carga) = Carga base × tarifa "
          "base (Op1); en Op2 se suma Carga solar × tarifa intermedia (crédito neto no "
          "aprovechado). Q y L provienen de la simulación horaria al tamaño de diseño y "
          "escalan con MIN(1, E_útil/E_diseño). Prefactibilidad — no bancable.")

HDR = {'U': 'Exceso solar (kWh)', 'V': 'Carga solar BESS (kWh)', 'W': 'Carga base (kWh)'}


def fnum(x):
    return repr(float(x))


def repl_value(xml, ref, newval):
    pat = re.compile(r'(<c r="' + ref + r'"[^>]*>)<v>[^<]*</v>(</c>)')
    xml2, n = pat.subn(lambda m: m.group(1) + '<v>' + fnum(newval) + '</v>' + m.group(2), xml)
    assert n == 1, f'{ref}: value patch matched {n}'
    return xml2


def repl_formula(xml, ref, formula, cached):
    pat = re.compile(r'<c r="' + ref + r'"([^>]*)>.*?</c>', re.S)
    def rep(m):
        return f'<c r="{ref}"{m.group(1)}><f>{formula}</f><v>{fnum(cached)}</v></c>'
    xml2, n = pat.subn(rep, xml, count=1)
    assert n == 1, f'{ref}: formula patch matched {n}'
    return xml2


def insert_cells(xml, rownum, cells_xml):
    pat = re.compile(r'(<row r="' + str(rownum) + r'"[^>]*>.*?)(</row>)', re.S)
    xml2, n = pat.subn(lambda m: m.group(1) + cells_xml + m.group(2), xml, count=1)
    assert n == 1, f'row {rownum}: insert matched {n}'
    return xml2


def hdr_cells(rownum):
    out = ''
    for col in ('U', 'V', 'W'):
        out += f'<c r="{col}{rownum}" s="2" t="inlineStr"><is><t>{HDR[col]}</t></is></c>'
    return out


def uvw_cells(rownum, u, v, w):
    return (f'<c r="U{rownum}" s="5"><v>{fnum(u)}</v></c>'
            f'<c r="V{rownum}" s="5"><v>{fnum(v)}</v></c>'
            f'<c r="W{rownum}" s="5"><v>{fnum(w)}</v></c>')


def surger_sheet(xml, sitekey):
    site = MOTOR['sites'][sitekey]
    wb3 = _REV3CACHE[sitekey]
    # --- section 2 ---
    if sitekey in NEW_KWP_OP1:
        xml = repl_value(xml, 'C26', NEW_KWP_OP1[sitekey])
    c31 = wb3['C31'] * K_OP1
    d31 = wb3['D31'] * K_OP2
    xml = repl_value(xml, 'C31', c31)
    xml = repl_value(xml, 'D31', d31)
    # --- monthly blocks ---
    for op, base_row in [('op1', 60), ('op2', 77)]:
        months = site[op]['months']
        for i in range(12):
            m = months[i]
            r = base_row + i
            col = chr(ord('B') + i)                     # per-month tariff column
            xml = repl_value(xml, f'D{r}', m['D'])
            xml = repl_value(xml, f'E{r}', m['E'])
            xml = repl_value(xml, f'F{r}', m['F'])
            if op == 'op1':
                formula = f'W{r}*{col}$14'
            else:
                formula = f'W{r}*{col}$14+V{r}*{col}$15'
            xml = repl_formula(xml, f'O{r}', formula, m['O_new_mxn'])
            xml = insert_cells(xml, r, uvw_cells(r, m['exceso_kwh'],
                                                 m['carga_solar_kwh'], m['carga_base_kwh']))
        # header row for U/V/W
        xml = insert_cells(xml, base_row - 1, hdr_cells(base_row - 1))
    # --- row 73 note ---
    pat = re.compile(r'<c r="A73"([^>]*) t="s"><v>[^<]*</v></c>')
    xml, n = pat.subn(lambda mm: f'<c r="A73"{mm.group(1)} t="inlineStr"><is><t>{NOTE73}</t></is></c>', xml)
    assert n == 1, f'A73 note matched {n}'
    # --- typical day: col D (pv), col F (BESS carga) ---
    td = site['typical_day']
    for season, row0 in [('invierno', 152), ('verano', 188)]:
        prof = td[season]
        for h in range(24):
            r = row0 + h
            xml = repl_value(xml, f'D{r}', prof[h]['pv'])
            xml = repl_value(xml, f'F{r}', prof[h]['chg'])
    # --- dimension ---
    xml = xml.replace('<dimension ref="A1:T219"/>', '<dimension ref="A1:W219"/>')
    return xml


_REV3CACHE = {}
def load_rev3_section2(bookpath, sites):
    wb = openpyxl.load_workbook(bookpath, data_only=True)
    for sk in sites:
        ws = wb[SHEET[sk]]
        _REV3CACHE[sk] = {'C31': ws['C31'].value, 'D31': ws['D31'].value}


def build(book_in, book_out, sites):
    load_rev3_section2(book_in, sites)
    z = zipfile.ZipFile(book_in)
    smap = _sheet_file_map(z)
    target_files = {smap[SHEET[sk]]: sk for sk in sites}
    with zipfile.ZipFile(book_out, 'w', zipfile.ZIP_DEFLATED) as zo:
        for item in z.infolist():
            data = z.read(item.filename)
            if item.filename in target_files:
                xml = data.decode('utf-8')
                xml = surger_sheet(xml, target_files[item.filename])
                data = xml.encode('utf-8')
            elif item.filename == 'xl/workbook.xml':
                xml = data.decode('utf-8')
                if 'fullCalcOnLoad' not in xml:
                    if '<calcPr' in xml:
                        xml = re.sub(r'<calcPr([^>]*)/>', r'<calcPr\1 fullCalcOnLoad="1"/>', xml)
                    else:
                        xml = xml.replace('</workbook>', '<calcPr fullCalcOnLoad="1"/></workbook>')
                    data = xml.encode('utf-8')
            zo.writestr(item, data)
    z.close()
    print('wrote', os.path.basename(book_out))


def _sheet_file_map(z):
    wbxml = z.read('xl/workbook.xml').decode('utf-8')
    rels = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')
    ridmap = dict(re.findall(r'<Relationship Id="([^"]+)"[^>]*Target="(worksheets/[^"]+)"', rels))
    out = {}
    for mm in re.finditer(r'<sheet [^>]*name="([^"]+)"[^>]*r:id="([^"]+)"', wbxml):
        name, rid = mm.group(1), mm.group(2)
        if rid in ridmap:
            out[name] = 'xl/' + ridmap[rid]
    return out


if __name__ == '__main__':
    print(f'k_op1(self)={K_OP1:.10f}  k_op2(self)={K_OP2:.10f}')
    build(BOOK4_IN, BOOK4_OUT, BOOK_SITES[BOOK4_OUT])
    build(BOOKP_IN, BOOKP_OUT, BOOK_SITES[BOOKP_OUT])
    print('DONE')
