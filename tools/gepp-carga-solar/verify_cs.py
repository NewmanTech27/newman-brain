#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
verify_cs.py — Verification gates for the rev4 "Carga Solar" books (DESIGN.md §5).

GATE 1  Old-mode replica: motor with kWp_old + solar-charge OFF reproduces rev3
        R72 (monthly total) and C52 (TIR BESS via projection replica) <0.1%, 7 sheets x 2 ops.
        C52/C55 reference = the rev3 book RECALCULATED by LibreOffice (scratchpad copies),
        NOT the delivered cache: the delivered rev3 books carry STALE projection caches
        (e.g. ixt C52 cache 11.363% vs live-formula 10.462%; C55 cache 128.6M vs 103.5M)
        — a pre-existing rev3 condition (cache predates a late Supuestos change), flagged
        in the run output. R72/C44 caches ARE consistent.
GATE 2  Energy identities per month (motor_cs.json): V+W == N/RTE ; V <= U ;
        recalced sheet N rows == motor N ; UVW cells in book == motor.
GATE 3  Book sanity: LibreOffice-recalculated *_values.xlsx C44/C51/C52/C55 (+R72, M72)
        vs the Python projection replica at solved k, <0.5%.

Run with venv-dd python (openpyxl). READ-ONLY.
"""
import json, os, sys
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import dispatch_cs as D
import solve_k_cs as K

MOTOR = json.load(open(os.path.join(HERE, 'motor_cs.json')))
KS = json.load(open(os.path.join(HERE, 'solve_k_cs.json')))
RTE = D.RTE

REV3 = {
    '4S': '/home/mario/CFE Brain/work/gepp-deck/GEPP - Propuesta PV+BESS 4 Sitios (2026-07) rev3 - Esc 4pct TIR 14.xlsx',
    'PP': '/home/mario/CFE Brain/work/gepp-deck/GEPP - Propuesta Proplasa (3 sitios) (2026-07) rev3 - Esc 4pct TIR 14.xlsx'}
NEWV = {
    '4S': os.path.join(HERE, 'GEPP - Solucion Energetica Solar-Charge BESS - 4 Sitios_values.xlsx'),
    'PP': os.path.join(HERE, 'GEPP - Solucion Energetica Solar-Charge BESS - Proplasa_values.xlsx')}
NEWF = {
    '4S': os.path.join(HERE, 'GEPP - Solucion Energetica Solar-Charge BESS - 4 Sitios.xlsx'),
    'PP': os.path.join(HERE, 'GEPP - Solucion Energetica Solar-Charge BESS - Proplasa.xlsx')}
SHEET = {'ixt': ('4S', 'Ixtlahuacán'), 'aca': ('4S', 'Acapulco'), 'can': ('4S', 'Cancún'),
         'pro': ('4S', 'Proplasa'), 'pr1': ('PP', 'Preforma 1'), 'pr2': ('PP', 'Preforma 2'),
         'tap': ('PP', 'Tapa')}
OPCOL = {'op1': 'C', 'op2': 'D'}
TOTROW = {'op1': 72, 'op2': 89}
BASEROW = {'op1': 60, 'op2': 77}

fails = []
def chk(label, got, exp, tol):
    if exp == 0:
        ok = abs(got) < 1e-6
        rel = abs(got)
    else:
        rel = abs(got - exp) / abs(exp)
        ok = rel <= tol
    if not ok:
        fails.append(f'{label}: got {got!r} expected {exp!r} (rel {rel:.4%} > {tol:.2%})')
    return ok

REV3R = {'4S': os.path.join(HERE, '_rev3_recalc', 'rev3_4S_recalc.xlsx'),
         'PP': os.path.join(HERE, '_rev3_recalc', 'rev3_PP_recalc.xlsx')}
for _p in REV3R.values():
    if not os.path.exists(_p):
        sys.exit('missing ' + _p + ' — run ./recalc.sh first')

wb3 = {k: openpyxl.load_workbook(p, data_only=True) for k, p in REV3.items()}
wb3r = {k: openpyxl.load_workbook(p, data_only=True) for k, p in REV3R.items()}
wbv = {k: openpyxl.load_workbook(p, data_only=True) for k, p in NEWV.items()}
wbf = {k: openpyxl.load_workbook(p, data_only=True) for k, p in NEWF.items()}

B = D.books()

# ---------------- GATE 1: old-mode replica ----------------
print('GATE 1 — old-mode replica (kWp_old, solar OFF) vs rev3 LO-recalced, tol 0.1%')
print('  NOTE: delivered rev3 caches for C52/C55 are STALE vs their own live formulas:')
for sk, (bk, sh) in SHEET.items():
    c_cache = wb3[bk][sh]['C52'].value
    c_live = wb3r[bk][sh]['C52'].value
    print(f'    {sk}: C52 cache {c_cache:.4%} vs live {c_live:.4%} | '
          f'C55 cache {wb3[bk][sh]["C55"].value:,.0f} vs live {wb3r[bk][sh]["C55"].value:,.0f}')
for sk, (bk, sh) in SHEET.items():
    ws = wb3r[bk][sh]
    for op in ('op1', 'op2'):
        kwp_old = B[sk][op]['26']
        ann = D.motor_site(sk, op, kwp_old, solar_charge=False)['annual']
        r_ref = ws[f'R{TOTROW[op]}'].value
        chk(f'G1 {sk} {op} R72', ann['R72_new'], r_ref, 0.001)
        # C52 (TIR BESS) replica: projection from rev3 anchors at k=1, old motor
        c = B[sk][op]
        capex_pv = kwp_old * 1000 * K.S['B5'] * K.S['B4']
        A = dict(C38=c['38'], C39_base=c['39'] - c['38'], C42=ann['K72_self'],
                 C43=ann['M72_new'], C44=ann['ahorro_bess_bruto'], C26=kwp_old,
                 C29=c['29'], C32=K.S['B17'], N19=c['N19'], price0=c['31'],
                 price=c['31'], capex_pv=capex_pv, capex_bess=c['36'], C51_kwp=kwp_old)
        R = K.recompute(A)
        c52_ref = ws[f'{OPCOL[op]}52'].value
        chk(f'G1 {sk} {op} C52', R['C52'], c52_ref, 0.001)
print(('  PASS (28 checks)' if not fails else f'  FAIL {len(fails)}'))
n_after_g1 = len(fails)

# ---------------- GATE 2: energy identities ----------------
print('GATE 2 — energy identities per month + injected UVW/N vs motor')
nchk = 0
for sk, (bk, sh) in SHEET.items():
    wsv = wbv[bk][sh]
    for op in ('op1', 'op2'):
        months = MOTOR['sites'][sk][op]['months']
        for i, m in enumerate(months):
            r = BASEROW[op] + i
            lbl = f'G2 {sk} {op} m{i}'
            # identity V+W = N/RTE
            chk(f'{lbl} V+W=N/RTE', m['carga_solar_kwh'] + m['carga_base_kwh'],
                m['N_kwh'] / RTE, 0.001); nchk += 1
            # V <= U
            if m['carga_solar_kwh'] > m['exceso_kwh'] + 0.51:
                fails.append(f'{lbl}: V {m["carga_solar_kwh"]} > U {m["exceso_kwh"]}')
            nchk += 1
            # sheet N (live) still equals motor N
            chk(f'{lbl} sheetN', wsv[f'N{r}'].value or 0.0, m['N_kwh'], 0.001); nchk += 1
            # injected U/V/W present & correct in the book
            for col, key in (('U', 'exceso_kwh'), ('V', 'carga_solar_kwh'), ('W', 'carga_base_kwh')):
                chk(f'{lbl} {col}', wsv[f'{col}{r}'].value or 0.0, m[key], 0.0005); nchk += 1
            # O live formula recalced == motor O_new
            chk(f'{lbl} O', wsv[f'O{r}'].value or 0.0, m['O_new_mxn'], 0.005); nchk += 1
print(('  PASS' if len(fails) == n_after_g1 else f'  FAIL {len(fails)-n_after_g1}') + f' ({nchk} checks)')
n_after_g2 = len(fails)

# ---------------- GATE 3: LO-recalced book vs Python projection ----------------
print('GATE 3 — LO full-recalc vs Python replica at solved k, tol 0.5%')
KVAL = {'op1': KS['op1']['k_selfbasis'], 'op2': KS['op2']['k_selfbasis']}
summary = {}
for sk, (bk, sh) in SHEET.items():
    wsv = wbv[bk][sh]
    ws3 = wb3[bk][sh]
    for op in ('op1', 'op2'):
        A = K.anchors(sk, op, price_mult=KVAL[op])
        R = K.recompute(A)
        col = OPCOL[op]
        tr = TOTROW[op]
        ws3r = wb3r[bk][sh]
        got_R72 = wsv[f'R{tr}'].value
        exp_ann = MOTOR['sites'][sk][op]['annual']
        chk(f'G3 {sk} {op} R72', got_R72, exp_ann['R72_new_mxn'], 0.005)
        chk(f'G3 {sk} {op} C44', wsv[f'{col}44'].value, A['C44'], 0.005)
        chk(f'G3 {sk} {op} C51', wsv[f'{col}51'].value, R['C51'], 0.005)
        chk(f'G3 {sk} {op} C52', wsv[f'{col}52'].value, R['C52'], 0.005)
        chk(f'G3 {sk} {op} C55', wsv[f'{col}55'].value, R['C55'], 0.005)
        chk(f'G3 {sk} {op} C42(K72)', wsv[f'{col}42'].value, A['C42'], 0.005)
        chk(f'G3 {sk} {op} C43(M72)', wsv[f'{col}43'].value, A['C43'], 0.005)
        chk(f'G3 {sk} {op} C31(price)', wsv[f'{col}31'].value, A['price'], 0.0001)
        if op == 'op1':
            summary[sk] = dict(
                price_rev3=A['price0'], price_rev4=A['price'],
                ahorro_bess_rev3=ws3[f'{col}44'].value, ahorro_bess_rev4=wsv[f'{col}44'].value,
                tir_bess_rev3=ws3r[f'{col}52'].value, tir_bess_rev4=wsv[f'{col}52'].value,
                npv_cliente_rev3_cached=ws3r[f'{col}55'].value, npv_cliente_rev4=wsv[f'{col}55'].value)
print(('  PASS' if len(fails) == n_after_g2 else f'  FAIL {len(fails)-n_after_g2}') + ' (112 checks)')

print()
if fails:
    print(f'== {len(fails)} FAILURES ==')
    for f in fails[:60]:
        print(' ', f)
    sys.exit(1)
print('== ALL GATES PASS ==')
print()
print('site  ppa_rev3->rev4      ahorroBESS r3->r4 (MXN)   TIR_BESS r3->r4     NPVcli r3(live)->r4')
for sk, s in summary.items():
    print(f"{sk:4s} {s['price_rev3']:.4f}->{s['price_rev4']:.4f}   "
          f"{s['ahorro_bess_rev3']:>12,.0f}->{s['ahorro_bess_rev4']:>12,.0f}   "
          f"{s['tir_bess_rev3']:.3%}->{s['tir_bess_rev4']:.3%}   "
          f"{s['npv_cliente_rev3_cached']:>13,.0f}->{s['npv_cliente_rev4']:>13,.0f}")
