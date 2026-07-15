#!/usr/bin/env python3
# Extract GEPP "Carga Solar" (rev4) data from the RECALCULATED values books into
# gepp_data_cs.json, shaped exactly like gepp_data_tir14.json so the MODEL splice works.
# Dispatch (inv/ver .prof) is OVERRIDDEN with motor_cs.json typical_day curves.
import openpyxl, json

SP = "."
B4 = f"{SP}/GEPP - Solucion Energetica Solar-Charge BESS - 4 Sitios_values.xlsx"
BP = f"{SP}/GEPP - Solucion Energetica Solar-Charge BESS - Proplasa_values.xlsx"

def cellrows(ws):
    rows = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                rows.setdefault(c.row, {})[c.column_letter] = c.value
    return rows

def find_row(rows, prefix, col='A'):
    for r in sorted(rows):
        v = rows[r].get(col)
        if isinstance(v, str) and v.strip().startswith(prefix):
            return r
    return None

def num(v):
    return round(v, 6) if isinstance(v, (int, float)) else None

MONCOLS = list("BCDEFGHIJKLM")

def extract_site(ws):
    rows = cellrows(ws)
    out = {}
    out['title'] = rows[1]['A']
    r0 = find_row(rows, '1. DATOS BASE')
    hdr = r0 + 1
    out['months'] = [rows[hdr].get(c) for c in MONCOLS]
    def series(label):
        for r in range(hdr, hdr + 20):
            v = rows.get(r, {}).get('A', '')
            if isinstance(v, str) and v.startswith(label):
                return [num(rows[r].get(c)) for c in MONCOLS], num(rows[r].get('N'))
        return None, None
    for key, label in [('cbase','Consumo Base'),('cint','Consumo Intermedia'),('cpunta','Consumo Punta'),
                       ('ctotal','Consumo Total'),('dpunta','Dem. máx. en punta'),('dmax','Dem. máx. mensual'),
                       ('gasto','Gasto actual facturado')]:
        s, t = series(label)
        out[key], out[key+'_tot'] = s, t
    r2 = find_row(rows, '2. SISTEMA PROPUESTO')
    r3 = find_row(rows, '3. MODELO MENSUAL')
    sys = {}
    for r in range(r2+1, r3):
        a = rows.get(r, {}).get('A')
        if isinstance(a, str) and a.strip() and not a.startswith('—'):
            sys[a.strip()] = [rows[r].get('C'), rows[r].get('D')]
    out['sys'] = {k: [num(x) if isinstance(x,(int,float)) else x for x in v] for k, v in sys.items()}
    hdr3 = r3 + 1
    mm = []
    r = hdr3 + 1
    while True:
        b = rows.get(r, {}).get('B')
        if b == 'TOTAL' or b is None: break
        mm.append({'pv': num(rows[r].get('M')), 'bess': num(rows[r].get('R'))})
        r += 1
    out['mm1'] = mm
    def proj(section):
        rs = find_row(rows, section)
        if rs is None: return None
        h = rs + 1
        res = []
        r = h + 1
        while True:
            b = rows.get(r, {}).get('B')
            if b is None: break
            res.append({'yr': b,
                'pleno': num(rows[r].get('C')), 'autoab': num(rows[r].get('D')),
                'real': num(rows[r].get('E')), 'apv': num(rows[r].get('F')),
                'abess': num(rows[r].get('G')), 'ppa': num(rows[r].get('H')),
                'part': num(rows[r].get('I')), 'neto': num(rows[r].get('J')),
                'acum': num(rows[r].get('K')), 'pct': num(rows[r].get('L'))})
            r += 1
        return res
    out['proj1'] = proj('5. PROYECCIÓN 20 AÑOS — OPCIÓN 1')
    out['proj2'] = proj('6. PROYECCIÓN 20 AÑOS — OPCIÓN 2')
    def dispatch(prefix):
        rs = find_row(rows, prefix)
        if rs is None: return None
        h = rs + 1
        prof = []
        for r in range(h+1, h+25):
            rr = rows.get(r, {})
            prof.append({'h': rr.get('A'), 'per': rr.get('B'), 'carga': num(rr.get('C')),
                         'pv': num(rr.get('D')), 'chg': num(rr.get('F')), 'dis': num(rr.get('G')),
                         'neta': num(rr.get('H'))})
        lect = {}
        for r in range(h+25, h+35):
            a = rows.get(r, {}).get('A', '')
            e = rows.get(r, {}).get('E')
            if isinstance(a, str) and e is not None:
                lect[a.strip()] = num(e)
        return {'prof': prof, 'lect': lect}
    out['inv'] = dispatch('8. DESPACHO DÍA TÍPICO — INVIERNO')
    out['ver'] = dispatch('9. DESPACHO DÍA TÍPICO — VERANO')
    return out

def extract_resumen(wb):
    ws = wb['Resumen']
    rows = cellrows(ws)
    res = {'rows': [], 'portfolio': {}}
    site = None
    for r in sorted(rows):
        rr = rows[r]
        a, b = rr.get('A'), rr.get('B')
        if a == 'PORTAFOLIO' or (site == 'PORTAFOLIO' and b):
            site = 'PORTAFOLIO'
            if b:
                res['portfolio'][b] = {k: num(rr.get(c)) for k, c in
                    [('kwp','C'),('bkw','D'),('bkwh','E'),('gen','F'),('gasto','G'),('autoab','H'),
                     ('bruto','I'),('ppa','J'),('part','K'),('neto','L')]}
            continue
        if b and isinstance(b, str) and b.startswith('Op'):
            if a: site = a
            res['rows'].append({'site': site, 'op': b, **{k: num(rr.get(c)) for k, c in
                [('kwp','C'),('bkw','D'),('bkwh','E'),('gen','F'),('gasto','G'),('autoab','H'),
                 ('bruto','I'),('ppa','J'),('part','K'),('neto','L'),('pctb','M'),('pctn','N'),
                 ('ppakwh','O'),('tirpv','P'),('tirbess','Q')]}})
    return res

wb4 = openpyxl.load_workbook(B4, data_only=True)
wbp = openpyxl.load_workbook(BP, data_only=True)

data = {'resumen4': extract_resumen(wb4), 'resumenP': extract_resumen(wbp), 'sites': {}}
SHEETS = [('ixt', wb4, 'Ixtlahuacán'), ('aca', wb4, 'Acapulco'), ('can', wb4, 'Cancún'),
          ('pro', wb4, 'Proplasa'), ('pr1', wbp, 'Preforma 1'), ('pr2', wbp, 'Preforma 2'),
          ('tap', wbp, 'Tapa')]
for key, wb, sn in SHEETS:
    data['sites'][key] = extract_site(wb[sn])

# ---- OVERRIDE dispatch prof with motor_cs.json typical_day curves ----
motor = json.load(open(f"{SP}/motor_cs.json"))
for key, _, _ in SHEETS:
    td = motor['sites'][key]['typical_day']
    for season, mkey in [('inv','invierno'), ('ver','verano')]:
        prof = []
        for p in td[mkey]:
            prof.append({'h': p['h'], 'per': p['per'], 'carga': num(p['carga']),
                         'pv': num(p['pv']), 'chg': num(p['chg']), 'dis': num(p['dis']),
                         'neta': num(p['neta'])})
        data['sites'][key][season] = {'prof': prof,
                                      'lect': data['sites'][key][season].get('lect', {}) if data['sites'][key].get(season) else {}}

# ---- superficie sweep (standalone basis for per-site claims + union curve) ----
sweep = json.load(open(f"{SP}/sweep_cs_ext.json"))
data['sweep'] = {'meta': sweep['meta'], 'sites': {}}
for key in [s[0] for s in SHEETS]:
    s = sweep['sites'][key]
    st = s.get('standalone', {})
    data['sweep']['sites'][key] = {
        'kwp_design': s['kwp_design'], 'kwp_rev3': s['kwp_rev3'],
        'm2_disponibles': s.get('m2_disponibles'),
        'kwp_cap_regulatorio': s.get('kwp_cap_regulatorio'),
        # union curve (illustrative shape; has van + ahorro + remanente)
        'curve': [{'kwp': c['kwp'], 'm2': c['m2'], 'ahorro': c['ahorro_cliente_anio1_mxn'],
                   'van': c['van_20a_mxn'], 'rem': c['exceso_remanente_kwh']} for c in s['curve']],
        'union_design': s['design_point'], 'union_opt': s['opt_point'],
        # standalone = defensible per-site bound
        'standalone': {
            'kwp_opt': st.get('kwp_opt'), 'm2_opt': st.get('m2_opt'),
            'm2_faltantes': st.get('m2_faltantes'), 'opt_at_cap': st.get('opt_at_regulatory_cap'),
            'design': st.get('design_point'), 'opt': st.get('opt_point'),
            'delta_ahorro': st.get('delta_ahorro_anio1_opt_vs_design'),
        }
    }

# ---- solar-charge annual blocks (for cobertura bono per site) ----
data['solarcharge'] = {}
for key in [s[0] for s in SHEETS]:
    a = motor['sites'][key]['op1']['annual']
    data['solarcharge'][key] = {
        'kwp_rev3': a['kwp_rev3'], 'kwp_rev4': a['kwp_rev4'], 'gen_kwh': a['gen_kwh'],
        'exceso_kwh': a['exceso_kwh'], 'carga_solar_kwh': a['carga_solar_kwh'],
        'carga_base_kwh': a['carga_base_kwh'], 'pct_carga_solar': a['pct_carga_solar'],
        'solar_charge_bonus_mxn': a['solar_charge_bonus_mxn'],
        'price_ppa_rev3': a['price_ppa_rev3'], 'price_ppa_rev4': a['price_ppa_rev4'],
    }
data['meta_cs'] = {'k': motor['meta']['k'], 'new_kwp_op1': motor['meta']['new_kwp_op1'],
                   'm2_per_kwp': motor['meta']['m2_per_kwp']}

json.dump(data, open(f"{SP}/gepp_data_cs.json", 'w'), ensure_ascii=False)
print('wrote gepp_data_cs.json', len(json.dumps(data))//1024, 'KB')
# quick sanity print
for row in data['resumen4']['rows']:
    if row['op'] and row['op'].startswith('Op'):
        print(f"  {row['site']:12s} {row['op']:22s} kwp={row['kwp']} gen={row['gen']} neto={row['neto']} ppakwh={row['ppakwh']} tirpv={row['tirpv']} tirbess={row['tirbess']}")
print('portfolio4:', {k:(v['gen'],v['neto']) for k,v in data['resumen4']['portfolio'].items()})
